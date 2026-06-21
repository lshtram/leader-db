"""Maddison Project Database 2023 source loader for the Chronicle slice.

This module owns the Maddison-specific Chronicle source loader
(:class:`MaddisonSource` + :func:`load_maddison_source`). It mirrors
the Stage 2 adapter's narrow-frame schema so the lookup code path is
identical regardless of whether the data came from the processed
parquet or the raw xlsx.

The loader honours the Maddison 2023 -> 2022 proxy mapping (the 2023
release ends at 2022) so the row builder can attach the
``proxy_year_used`` flag whenever a 2023+ target-year request is
filled by Maddison.
"""

from __future__ import annotations

import logging
from dataclasses import dataclass, field
from pathlib import Path

import pandas as pd

from ..paths import processed_dir, raw_dir
from .constants import MADDISON_PROXY_REQUESTED_YEAR, MADDISON_PROXY_YEAR

_logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# MaddisonSource
# ---------------------------------------------------------------------------


@dataclass(frozen=True)
class MaddisonSource:
    """In-memory Maddison Project 2023 slice keyed by ``(iso3, year)``.

    The loader prefers the Stage 2 narrow parquet
    (``data/processed/maddison_project/maddison_project_country_year.parquet``)
    because the parquet read is O(scan) on a small file. When the
    parquet is missing we fall back to the raw xlsx
    (``data/raw/maddison_project/mpd2023.xlsx``) by reading the
    ``Full data`` sheet once with ``openpyxl.read_only=True``.

    Attributes:
        parquet_path: Path to the narrow parquet (if used).
        xlsx_path: Path to the raw xlsx (used as fallback).
        frame: Narrow ``DataFrame`` with the three catalog
            indicators (``maddison_project_gdp_per_capita_2011_intl``,
            ``maddison_project_population_thousands``,
            ``maddison_project_gdp_total_2011_intl_derived``) per
            ``(countrycode, year)``.
    """

    parquet_path: Path | None
    xlsx_path: Path | None
    frame: pd.DataFrame = field(default_factory=pd.DataFrame)

    def lookup(
        self, iso3: str, year: int
    ) -> tuple[float | None, float | None, float | None, int, bool]:
        """Return ``(gdppc, pop, derived_gdp_total, source_year_used, is_proxy)``.

        Honors the 2023 -> 2022 proxy mapping (the 2023 release ends
        at 2022). When ``year >= MADDISON_PROXY_REQUESTED_YEAR`` and
        ``is_proxy=True``, ``source_year_used`` is
        ``MADDISON_PROXY_YEAR`` and the row builder should attach the
        ``proxy_year_used`` flag.

        Returns ``(None, None, None, year, False)`` when no Maddison
        data exists for the requested ``(iso3, year)`` pair.
        """
        if self.frame.empty:
            return (None, None, None, year, False)

        effective_year = year
        is_proxy = False
        if year >= MADDISON_PROXY_REQUESTED_YEAR:
            effective_year = MADDISON_PROXY_YEAR
            is_proxy = True

        # The narrow frame columns are: countrycode, year, country,
        # region, variable_name, raw_column, raw_value,
        # normalized_value. We pivot to a wide lookup so the
        # per-(iso3, year) row is materialized once.
        mask = (
            (self.frame["countrycode"] == iso3)
            & (self.frame["year"] == effective_year)
        )
        matches = self.frame.loc[mask]
        if matches.empty:
            return (None, None, None, year, False)

        def _value_for(variable: str) -> float | None:
            row = matches.loc[matches["variable_name"] == variable]
            if row.empty:
                return None
            v = row.iloc[0]["normalized_value"]
            return None if pd.isna(v) else float(v)

        gdppc = _value_for("maddison_project_gdp_per_capita_2011_intl")
        pop = _value_for("maddison_project_population_thousands")
        derived = _value_for("maddison_project_gdp_total_2011_intl_derived")
        return (gdppc, pop, derived, effective_year, is_proxy)


# ---------------------------------------------------------------------------
# Loader
# ---------------------------------------------------------------------------


def load_maddison_source(
    *,
    parquet_path: Path | None = None,
    xlsx_path: Path | None = None,
    iso3_scope: tuple[str, ...] = (),
) -> MaddisonSource:
    """Load the Maddison Project slice from the processed parquet
    (preferred) or the raw xlsx (fallback).

    Parameters
    ----------
    parquet_path:
        Optional override for the narrow parquet. Default: the
        canonical processed-artifact path. When the parquet is
        absent the loader falls back to the raw xlsx.
    xlsx_path:
        Optional override for the raw xlsx. Default: the canonical
        raw data-lake path. Used when the parquet is absent.
    iso3_scope:
        ISO3 keys to keep. When empty (the default) the loader
        returns the full Maddison slice so the row builder can still
        resolve countries that fall back to Maddison at runtime
        (e.g. for full-country runs).
    """
    iso3_filter = set(iso3_scope) if iso3_scope else set()

    if parquet_path is None:
        parquet_path = (
            processed_dir("maddison_project")
            / "maddison_project_country_year.parquet"
        )

    if xlsx_path is None:
        xlsx_path = raw_dir("maddison_project") / "mpd2023.xlsx"

    if parquet_path.is_file():
        df = pd.read_parquet(parquet_path)
        if iso3_filter:
            df = df[df["countrycode"].isin(iso3_filter)].copy()
        return MaddisonSource(
            parquet_path=parquet_path,
            xlsx_path=xlsx_path,
            frame=df,
        )

    if xlsx_path.is_file():
        return _load_maddison_from_xlsx(
            xlsx_path=xlsx_path, iso3_filter=iso3_filter,
        )

    _logger.warning(
        "Maddison Project neither parquet (%s) nor xlsx (%s) is "
        "available; Maddison-backed population / GDP fields will be "
        "empty with missing_* flags.",
        parquet_path,
        xlsx_path,
    )
    return MaddisonSource(
        parquet_path=parquet_path,
        xlsx_path=xlsx_path,
        frame=pd.DataFrame(
            columns=[
                "countrycode",
                "year",
                "country",
                "region",
                "variable_name",
                "raw_column",
                "raw_value",
                "normalized_value",
            ]
        ),
    )


def _load_maddison_from_xlsx(  # noqa: PLR0915
    *, xlsx_path: Path, iso3_filter: set[str]
) -> MaddisonSource:
    """Read the Maddison raw xlsx and emit the long-format frame.

    The frame matches the Stage 2 narrow parquet schema so the
    :class:`MaddisonSource` lookup code path is identical for
    parquet-loaded and xlsx-loaded instances. The reading is
    single-pass over the ``Full data`` sheet (one openpyxl pass,
    ~5 s on the 4.9 MB / 131,144-row bundle).
    """
    import openpyxl

    wb = openpyxl.load_workbook(
        xlsx_path, read_only=True, data_only=True,
    )
    try:
        if "Full data" not in wb.sheetnames:
            raise ValueError(
                f"Maddison Project xlsx {xlsx_path} has no 'Full data' "
                f"sheet. Available sheets: {wb.sheetnames}."
            )
        ws = wb["Full data"]
        long_records: list[dict[str, object]] = []
        header: list[str] | None = None
        required_cols = {
            "countrycode", "country", "region", "year", "gdppc", "pop",
        }
        for row in ws.iter_rows(values_only=True):
            row_list = list(row)
            if header is None:
                header = [str(c) if c is not None else "" for c in row_list]
                if required_cols - set(header):
                    raise ValueError(
                        f"Maddison xlsx {xlsx_path} missing required "
                        f"columns: {sorted(required_cols - set(header))}."
                    )
                continue
            if all(c is None for c in row_list):
                continue
            rec = dict(zip(header, row_list, strict=False))
            cc_raw = rec.get("countrycode")
            if not isinstance(cc_raw, str) or not cc_raw:
                continue
            iso3 = cc_raw.strip()
            if iso3_filter and iso3 not in iso3_filter:
                continue
            try:
                year = int(rec.get("year")) if rec.get("year") is not None else None  # type: ignore[arg-type]
            except (TypeError, ValueError):
                continue
            if year is None:
                continue
            country = str(rec.get("country") or "")
            region = str(rec.get("region") or "")
            gdppc_cell = rec.get("gdppc")
            pop_cell = rec.get("pop")
            gdppc_present = (
                gdppc_cell is not None
                and not (isinstance(gdppc_cell, float) and pd.isna(gdppc_cell))
            )
            pop_present = (
                pop_cell is not None
                and not (isinstance(pop_cell, float) and pd.isna(pop_cell))
            )
            if gdppc_present:
                long_records.append(
                    {
                        "countrycode": iso3,
                        "year": year,
                        "country": country,
                        "region": region,
                        "variable_name": "maddison_project_gdp_per_capita_2011_intl",
                        "raw_column": "gdppc",
                        "raw_value": str(gdppc_cell),
                        "normalized_value": float(gdppc_cell),  # type: ignore[arg-type]
                    }
                )
            if pop_present:
                long_records.append(
                    {
                        "countrycode": iso3,
                        "year": year,
                        "country": country,
                        "region": region,
                        "variable_name": "maddison_project_population_thousands",
                        "raw_column": "pop",
                        "raw_value": str(pop_cell),
                        "normalized_value": float(pop_cell),  # type: ignore[arg-type]
                    }
                )
            if gdppc_present and pop_present:
                derived_total = (
                    float(gdppc_cell)  # type: ignore[arg-type]
                    * float(pop_cell)  # type: ignore[arg-type]
                    * 1000.0
                )
                long_records.append(
                    {
                        "countrycode": iso3,
                        "year": year,
                        "country": country,
                        "region": region,
                        "variable_name": "maddison_project_gdp_total_2011_intl_derived",
                        "raw_column": "__derived_gdp_total__",
                        "raw_value": f"{derived_total:.6f}",
                        "normalized_value": derived_total,
                    }
                )
    finally:
        wb.close()

    if not long_records:
        df = pd.DataFrame(
            columns=[
                "countrycode",
                "year",
                "country",
                "region",
                "variable_name",
                "raw_column",
                "raw_value",
                "normalized_value",
            ]
        )
    else:
        df = pd.DataFrame.from_records(long_records)
        df = df.sort_values(
            by=["year", "countrycode", "variable_name"],
            ascending=[True, True, True],
            kind="mergesort",
        ).reset_index(drop=True)

    return MaddisonSource(
        parquet_path=None,
        xlsx_path=xlsx_path,
        frame=df,
    )


__all__ = ["MaddisonSource", "load_maddison_source"]
