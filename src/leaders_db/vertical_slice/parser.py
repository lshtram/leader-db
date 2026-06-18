"""Client matrix xlsx parser for the 2023 vertical slice.

Implements the parser contract from architecture doc §5:

    load_vertical_slice_client_rows(path, sheet, year, iso3_scope)
        -> list[ClientSliceRow]

The parser is deliberately tiny: it reads the client xlsx with
:mod:`openpyxl`, takes rows from the documented zero-based offset
onward, and returns one :class:`ClientSliceRow` per scoped country.

The parser never writes to the database. The orchestrator uses the
returned rows to seed countries, country-years, leaders, ruler spells,
ruler years, and ruler scores.
"""

from __future__ import annotations

from collections.abc import Sequence
from pathlib import Path

import openpyxl
from pydantic import BaseModel, Field, field_validator

from ..normalize.countries import normalize_country_name
from .constants import SLICE_ISO3_BY_CLIENT_NAME


class ClientSliceRow(BaseModel):
    """One client matrix row scoped to the 2023 vertical slice.

    Mirrors the contract from architecture doc §5. All fields are
    populated by :func:`load_vertical_slice_client_rows`; downstream
    code may rely on ``client_scores`` returning ints in ``0..10`` (or
    ``None``) per the architecture.
    """

    iso3: str = Field(..., min_length=3, max_length=3)
    country_name: str = Field(..., min_length=1)
    population_raw: int | None = Field(default=None, ge=0)
    leader_name: str = Field(..., min_length=1)
    year_started_raw: int | None = Field(default=None, ge=0)
    client_scores: dict[str, int | None] = Field(default_factory=dict)
    source_row_number: int = Field(..., ge=1)

    @field_validator("iso3")
    @classmethod
    def _upper_iso3(cls, value: str) -> str:
        return value.upper()

    @field_validator("client_scores")
    @classmethod
    def _scores_in_range(cls, value: dict[str, int | None]) -> dict[str, int | None]:
        for key, score in value.items():
            if score is None:
                continue
            if not isinstance(score, int):
                raise ValueError(
                    f"client score {key!r} must be int or None, got {type(score).__name__}"
                )
            if not 0 <= score <= 10:
                raise ValueError(
                    f"client score {key!r}={score} outside 0..10"
                )
        return value


# ---------------------------------------------------------------------------
# Column map per architecture doc §5 (zero-based pandas positions)
# ---------------------------------------------------------------------------
#
# Layout (verified against the real client xlsx):
#
#   xlsx row 6 onward (pandas zero-based row 5 onward) are the data rows.
#   xlsx rows 1..5 are title / metadata / column-name hints; the parser
#   starts reading at pandas row index 5 (xlsx row 6).
#
# Column mapping (zero-based):
#
#     country:           3
#     population:        4
#     leader:           14
#     year_started:     15
#     social_wellbeing: 22
#     integrity:        23

_DATA_START_ROW: int = 5
_COL_COUNTRY: int = 3
_COL_POPULATION: int = 4
_COL_LEADER: int = 14
_COL_YEAR_STARTED: int = 15
_COL_SOCIAL_WELLBEING: int = 22
_COL_INTEGRITY: int = 23

_SLICE_CATEGORY_COLUMNS: dict[str, int] = {
    "social_wellbeing": _COL_SOCIAL_WELLBEING,
    "integrity": _COL_INTEGRITY,
}


def load_vertical_slice_client_rows(
    *,
    path: Path,
    sheet: str,
    year: int,
    iso3_scope: Sequence[str],
) -> list[ClientSliceRow]:
    """Parse selected client rows from the client xlsx.

    Implements the parser contract from architecture doc §5. Returns one
    :class:`ClientSliceRow` per ISO3 in ``iso3_scope``. Raises when a
    scoped country is not present in the xlsx (single-row requirement).

    The parser accepts ``None`` for any client-score cell (the slice
    must not invent scores) and validates that any non-None score is
    an int in ``0..10``.
    """
    iso3_set = {code.upper() for code in iso3_scope}
    if not iso3_set:
        return []

    workbook = openpyxl.load_workbook(
        filename=str(path), read_only=True, data_only=True
    )
    try:
        if sheet not in workbook.sheetnames:
            raise ValueError(
                f"sheet {sheet!r} not in workbook (sheets: {workbook.sheetnames})"
            )
        ws = workbook[sheet]
        rows = ws.iter_rows(
            min_row=_DATA_START_ROW + 1,  # openpyxl rows are 1-based
            values_only=True,
        )

        matched: dict[str, ClientSliceRow] = {}
        source_row_index = _DATA_START_ROW  # zero-based pandas row index
        for raw_row in rows:
            source_row_index += 1
            country_raw = raw_row[_COL_COUNTRY] if len(raw_row) > _COL_COUNTRY else None
            if not country_raw:
                continue
            normalized = normalize_country_name(str(country_raw))
            iso3 = SLICE_ISO3_BY_CLIENT_NAME.get(normalized)
            if iso3 is None or iso3 not in iso3_set:
                continue

            leader_raw = raw_row[_COL_LEADER] if len(raw_row) > _COL_LEADER else None
            leader_name = str(leader_raw).strip() if leader_raw is not None else ""
            if not leader_name:
                # Architecture §5: leader string must be non-empty.
                raise ValueError(
                    f"client row for {iso3} (xlsx row {source_row_index + 1}) "
                    f"has an empty leader cell at column {_COL_LEADER + 1}"
                )

            population_raw = (
                _coerce_int(raw_row[_COL_POPULATION])
                if len(raw_row) > _COL_POPULATION else None
            )
            year_started_raw = (
                _coerce_int(raw_row[_COL_YEAR_STARTED])
                if len(raw_row) > _COL_YEAR_STARTED else None
            )

            client_scores: dict[str, int | None] = {}
            for cat_key, col_idx in _SLICE_CATEGORY_COLUMNS.items():
                cell_value = raw_row[col_idx] if len(raw_row) > col_idx else None
                client_scores[cat_key] = _coerce_int(cell_value)

            if iso3 in matched:
                existing = matched[iso3]
                raise ValueError(
                    f"client xlsx {path.name!r} contains duplicate scoped "
                    f"country rows for ISO3 {iso3}: first at xlsx row "
                    f"{existing.source_row_number + 1}, second at xlsx row "
                    f"{source_row_index + 1}. Architecture §5 requires exactly "
                    "one row per scoped ISO3."
                )
            matched[iso3] = ClientSliceRow(
                iso3=iso3,
                country_name=str(country_raw).strip(),
                population_raw=population_raw,
                leader_name=leader_name,
                year_started_raw=year_started_raw,
                client_scores=client_scores,
                source_row_number=source_row_index,
            )
    finally:
        workbook.close()

    missing = sorted(iso3_set - set(matched))
    if missing:
        raise ValueError(
            f"client xlsx {path.name!r} is missing scoped countries: {missing}. "
            "Architecture §5 requires exactly one row per scoped ISO3."
        )

    # Stable order: sort by the iso3_scope order so tests are deterministic.
    scope_order = {code: i for i, code in enumerate(sorted(iso3_set))}
    return [matched[iso] for iso in sorted(matched, key=lambda c: scope_order[c])]


def _coerce_int(cell_value: object) -> int | None:
    """Coerce an openpyxl cell value to ``int`` or ``None``.

    Excel cells can come back as ``None``, ``int``, ``float`` (e.g.
    ``128.0`` for a population), or ``str`` (e.g. ``"-"``). The parser
    accepts None silently; non-None values must parse to a non-negative
    integer in the slice's contract.
    """
    if cell_value is None:
        return None
    if isinstance(cell_value, bool):
        return int(cell_value)
    if isinstance(cell_value, int):
        return _non_negative_int(cell_value)
    return _coerce_int_other(cell_value)


def _coerce_int_other(value: object) -> int | None:
    if isinstance(value, float):
        if _is_nan(value) or value < 0:
            return None
        return int(value)
    if isinstance(value, str):
        return _coerce_int_from_string(value)
    return None


def _non_negative_int(value: int) -> int | None:
    return value if value >= 0 else None


def _is_nan(value: float) -> bool:
    # NaN is the only float that is not equal to itself.
    return value != value  # noqa: PLR0124 (intentional self-equality NaN test)


def _coerce_int_from_string(text: object) -> int | None:
    if not isinstance(text, str):
        return None
    stripped = text.strip()
    if not stripped or stripped in {"-", "—", "n/a", "N/A"}:
        return None
    try:
        parsed = float(stripped)
    except ValueError:
        return None
    if parsed < 0:
        return None
    return int(parsed)


__all__ = ["ClientSliceRow", "load_vertical_slice_client_rows"]
