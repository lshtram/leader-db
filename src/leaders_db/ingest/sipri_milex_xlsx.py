"""Stage 2 -- SIPRI milex xlsx read (REQ-SRC-002).

This module is the SIPRI milex xlsx-read half of the adapter. It
owns:

- :func:`_read_xlsx_one_sheet` -- open one catalog indicator
  sheet with openpyxl in read-only mode, detect the per-sheet
  header row (variable position 6, 7, or 8), build the year-to-
  column map, and return the country rows with the year cells
  extracted.
- :func:`read_sipri_milex` -- the read orchestrator. Calls
  :func:`_read_xlsx_one_sheet` once per catalog indicator,
  builds a long-format frame per indicator, concatenates them,
  and pivots to wide format (one row per ``(country, year)``, one
  column per catalog ``variable_name``).

The 4 catalog sheets have variable header-row positions -- the
read function detects the header row dynamically by scanning for
the first row where col 0 is the literal string ``"Country"``
(row 6 for ``Share of GDP`` / ``Constant (2024) US$``; row 7 for
``Per capita``; row 8 for ``Share of Govt. spending``). The 5
data sheets interleave 15 region/sub-region labels (e.g.
``"Africa"``, ``"Americas"``) with the ~177 country names; the
read function filters out the region labels via
:data:`_SIPRI_MILEX_REGION_LABELS`. The 3 missing-value tokens
(``"..."``, ``"xxx"``, ``""``) are coerced to ``None``; the
original cell text is preserved in
``df.attrs["_sipri_milex_raw_long"]`` for the ``raw_value`` audit
trail.

Constants live in :mod:`leaders_db.ingest.sipri_milex_io` to break
the import cycle. The DB helpers live in
:mod:`leaders_db.ingest.sipri_milex_db`. The orchestrator lives
in :mod:`leaders_db.ingest.sipri_milex`.
"""

from __future__ import annotations

import datetime
from pathlib import Path

import openpyxl
import pandas as pd

from .sipri_milex_io import (
    _SIPRI_MILEX_REGION_LABELS,
    default_xlsx_path,
    load_indicator_catalog,
)

__all__ = ["read_sipri_milex"]

#: Header-row detection literal. The first row whose col 0 equals
#: this string is the year/header row. A rename in a future SIPRI
#: release (e.g. ``"country"`` lowercase) would silently break the
#: detection; we add a defensive ``ValueError`` if no row matches.
_HEADER_TOKEN: str = "Country"

#: Year-range filter for header-row year-column detection. Integers
#: outside this range are treated as non-year columns (e.g. the
#: ``Reporting year`` column in ``Share of Govt. spending``). The
#: ``+5`` buffer accommodates a future 2030 release; the lower bound
#: matches the oldest SIPRI year (1949).
_YEAR_MIN: int = 1940
_YEAR_MAX_BUFFER: int = 5  # year max = current_year + 5


# ---------------------------------------------------------------------------
# Single-sheet helpers
# ---------------------------------------------------------------------------


def _find_header_row(
    ws: openpyxl.worksheet.worksheet.Worksheet,
) -> int:
    """Find the row index where col 0 is the literal ``"Country"``.

    The SIPRI xlsx header-row position varies per sheet (6, 7, or
    8); the read function detects it dynamically by scanning for
    the first row where col 0 is the literal string ``"Country"``.
    Returns the 1-indexed row number (openpyxl convention).
    Raises :class:`ValueError` if no row matches (defensive fix
    for a future SIPRI release that renames the header).
    """
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if row and row[0] == _HEADER_TOKEN:
            return i
    raise ValueError(
        f"SIPRI milex sheet {ws.title!r} has no row with col 0 == "
        f"{_HEADER_TOKEN!r}. The sheet layout may have changed in a "
        "newer SIPRI release. Check the raw xlsx and update the read "
        "function."
    )


def _build_year_to_col_map(
    header_row: list[object],
) -> dict[int, int]:
    """Build the year-to-column-position map for a header row.

    Year cells are integers in the year range
    ``[_YEAR_MIN, current_year + _YEAR_MAX_BUFFER]``; non-year
    columns (Notes, Currency, Reporting year) are skipped.
    Returns ``{year: zero_indexed_col}`` for every year detected
    in the header.
    """
    current_year = datetime.datetime.now().year
    year_max = current_year + _YEAR_MAX_BUFFER
    year_to_col: dict[int, int] = {}
    for col_idx, cell in enumerate(header_row):
        # Defensive: bool is a subclass of int in Python; skip it.
        if isinstance(cell, bool):
            continue
        if isinstance(cell, int) and _YEAR_MIN <= cell <= year_max:
            year_to_col[cell] = col_idx
    return year_to_col


# ---------------------------------------------------------------------------
# Single-sheet reader
# ---------------------------------------------------------------------------


def _read_xlsx_one_sheet(
    xlsx_path: Path,
    sheet_name: str,
    *,
    wanted_years: set[int] | None = None,
) -> tuple[
    dict[int, int],
    list[tuple[str, int, list[object]]],
    list[str],
]:
    """Read one SIPRI milex indicator sheet.

    Per-sheet processing: detect the header row, build the
    year-to-column map, walk the data rows, filter region labels,
    and extract the year cells. Returns the year map, the data
    rows, and a list of region labels found.

    Args:
        xlsx_path: absolute path to the SIPRI milex xlsx.
        sheet_name: xlsx sheet name (one of the catalog's
            ``raw_column`` values).
        wanted_years: optional set of years to keep. If ``None``,
            all years in the header are kept. If a set, only those
            years are kept in the year-to-col map.

    Returns:
        A 3-tuple ``(year_to_col, rows, regions_found)``:

        1. ``year_to_col``: ``{year: zero_indexed_col}`` for every
           year in the header (after the ``wanted_years`` filter).
        2. ``rows``: list of ``(country, row_idx, row_values)``
           tuples in sheet order, one per country row. Region rows
           are excluded.
        3. ``regions_found``: list of region / sub-region names
           found in the input (preserved as the audit trail).

    Raises:
        KeyError: if the sheet is absent from the xlsx.
        ValueError: if no row matches ``"Country"``.
    """
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    try:
        if sheet_name not in wb.sheetnames:
            raise KeyError(
                f"SIPRI milex xlsx {xlsx_path} has no sheet named "
                f"{sheet_name!r}. Available sheets: {wb.sheetnames}. "
                "The catalog's `raw_column` may have drifted from "
                "the live xlsx."
            )
        ws = wb[sheet_name]
        header_row_idx = _find_header_row(ws)
        # Walk all rows; collect the header row and the data rows.
        header_row: list[object] = []
        rows: list[tuple[str, int, list[object]]] = []
        regions_found: list[str] = []
        for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if i == header_row_idx:
                header_row = list(row)
                continue
            if i <= header_row_idx:
                continue
            if row is None or all(c is None for c in row):
                # Blank separator row right after the header. Skip
                # it; the data rows resume after.
                continue
            country = row[0]
            if not isinstance(country, str):
                continue
            country_name = country.strip()
            if not country_name:
                continue
            if country_name in _SIPRI_MILEX_REGION_LABELS:
                regions_found.append(country_name)
                continue
            rows.append((country_name, i, list(row)))
        year_to_col = _build_year_to_col_map(header_row)
        if wanted_years is not None:
            year_to_col = {
                y: col for y, col in year_to_col.items()
                if y in wanted_years
            }
        return year_to_col, rows, regions_found
    finally:
        wb.close()


# ---------------------------------------------------------------------------
# Read orchestrator
# ---------------------------------------------------------------------------


def read_sipri_milex(
    *,
    year: int | None = None,
    xlsx_path: Path | None = None,
    catalog_path: Path | None = None,
) -> pd.DataFrame:
    """Read SIPRI milex from the xlsx and pivot to wide format.

    Steps (full description in
    ``docs/architecture/sipri_milex.md`` §3.3):

    1. Load the catalog.
    2. Open the xlsx at ``xlsx_path`` (default: data-lake path).
    3. For each catalog indicator, open the sheet, detect the
       header row, build the year-to-column map, walk the data
       rows (filtering region labels), and emit
       ``(country, year, variable_name, value)`` long rows.
    4. Concatenate per-indicator long frames.
    5. Pivot to wide format: one row per ``(country, year)``, one
       column per catalog ``variable_name``. Coerce ``year`` to
       ``int`` and indicator columns to ``float`` (NaN for
       missing).
    6. Attach ``df.attrs["regions_covered"]`` (sorted list of
       region names found in the input) and
       ``df.attrs["country_count"]`` (distinct country names in
       the wide frame).

    Args:
        year: filter to a single year (e.g. ``2023``). Default: all
            years present in the xlsx (1949-2025 for Share of GDP
            / Constant USD; 1988-2025 for Per capita / Share of
            Govt. spending).
        xlsx_path: override the input xlsx. Default: data-lake
            path.
        catalog_path: override the indicator catalog. Default:
            checked-in.

    Returns:
        A wide pandas DataFrame with columns ``country`` (display
        name), ``year`` (int), and one column per catalog
        indicator. SIPRI milex has no ISO3 column; the country
        column carries the raw display name (Stage 3 resolves it
        to ISO3 via ``country_aliases.csv``).

    Raises:
        FileNotFoundError: if the xlsx is missing.
        KeyError: if a catalog ``raw_column`` sheet name is absent
            from the xlsx.
        ValueError: if a sheet's header row cannot be found.
    """
    path = xlsx_path or default_xlsx_path()
    if not path.is_file():
        raise FileNotFoundError(f"SIPRI milex xlsx not found: {path}")

    specs = load_indicator_catalog(catalog_path=catalog_path)

    long_frames: list[pd.DataFrame] = []
    all_regions: set[str] = set()

    for spec in specs:
        sheet_name = spec.raw_column
        year_to_col, country_rows, regions_found = _read_xlsx_one_sheet(
            path, sheet_name,
        )
        all_regions.update(regions_found)

        if year is not None:
            wanted_years = {int(year)}
        else:
            wanted_years = set(year_to_col.keys())

        records: list[dict[str, object]] = []
        for country_name, _row_idx, row_values in country_rows:
            for one_year in sorted(wanted_years):
                col_idx = year_to_col.get(one_year)
                if col_idx is None or col_idx >= len(row_values):
                    # Sheet does not have this year (e.g. the
                    # Per capita sheet only goes back to 1988; if
                    # the caller asked for 1950, the column is
                    # missing). Skip.
                    continue
                # Store the raw cell (not the coerced value) in
                # the long frame so the audit trail preserves
                # the original SIPRI tokens ("...", "xxx", "").
                # The wide pivot's ``pd.to_numeric(errors="coerce")``
                # turns the SIPRI missing strings into NaN;
                # the raw long frame is the source of truth for
                # the literal cell text.
                cell = row_values[col_idx]
                records.append(
                    {
                        "country": country_name,
                        "year": one_year,
                        "variable_name": spec.variable_name,
                        "value": cell,
                    }
                )
        if records:
            long_frames.append(pd.DataFrame.from_records(records))

    if not long_frames:
        # Nothing to return -- emit an empty frame with the
        # expected columns so downstream code does not have to
        # special-case an empty result.
        wide = pd.DataFrame(
            columns=["country", "year", "variable_name", "value"],
        )
        wide = wide.drop(columns=["variable_name", "value"])
        wide.attrs["regions_covered"] = sorted(all_regions)
        wide.attrs["country_count"] = 0
        return wide

    long_df = pd.concat(long_frames, ignore_index=True)

    # Pivot to wide format: one row per (country, year), one
    # column per variable_name.
    wide = long_df.pivot_table(
        index=["country", "year"],
        columns="variable_name",
        values="value",
        aggfunc="first",
    )
    wide = wide.reset_index()

    # Type coercion. ``year`` is already int (we constructed it
    # that way). Indicator columns may be numeric, string, or
    # None; coerce to float and let ``pd.to_numeric`` produce NaN
    # for the non-numeric cells.
    wide["year"] = wide["year"].astype(int)
    for col in list(wide.columns):
        if col in {"country", "year"}:
            continue
        wide[col] = pd.to_numeric(wide[col], errors="coerce").astype(float)

    # Preserve the pre-coercion long frame in ``df.attrs`` so the
    # DB write can recover the original cell text for the
    # ``raw_value`` audit-trail column. SIPRI's missing-data
    # convention is the literal string ``"..."`` / ``"xxx"`` /
    # ``""``; after ``pd.to_numeric(errors="coerce")`` those cells
    # become NaN. The audit trail must keep the literal token so
    # future readers of the source_observations table can tell
    # what the xlsx actually said. The attr key is private
    # (underscore prefix) so it does not leak into the parquet.
    wide.attrs["_sipri_milex_raw_long"] = long_df.copy()
    # SIPRI-specific audit fields.
    wide.attrs["regions_covered"] = sorted(all_regions)
    wide.attrs["country_count"] = int(wide["country"].nunique())

    return wide
