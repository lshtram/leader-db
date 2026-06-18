"""Stage 2 -- World Bank WGI xlsx read (REQ-SRC-002).

This module is the WGI xlsx-read half of the WGI adapter. It owns:

- :func:`_read_xlsx_one_sheet` -- open one WGI indicator sheet with
  openpyxl in read-only mode, walk the per-sheet header (row 14 = year
  row, row 15 = stat-type row), and return the year-to-Estimate-column
  map plus the country rows.
- :func:`read_wgi` -- the read orchestrator. Calls
  :func:`_read_xlsx_one_sheet` once per catalog indicator, builds a
  long-format frame per indicator, concatenates them, and pivots to
  wide format (one row per ``(iso3, year)``, one column per catalog
  ``variable_name``).

WGI xlsx layout (verified live against the 2023 Update release):
7 sheets total -- an ``Introduction`` title sheet and 6 indicator
sheets (``VoiceandAccountability``, ``Political StabilityNoViolence``
with a literal space, ``GovernmentEffectiveness``, ``RegulatoryQuality``,
``RuleofLaw``, ``ControlofCorruption``). Per indicator sheet, rows
1-13 are title/legend/disclaimer, row 14 is the year row (6 cols
repeated per year), row 15 is the stat-type row, and rows 16+ are
country data (214 real countries in the live xlsx; 5 in the test
fixture). The ``Estimate`` column is at position 2, 8, 14, ... in
zero-indexed terms (one per year, 6 stats per year). Missing data is
the literal string ``"#N/A"``.

Constants (the source key, the WGI attribution, the
:class:`IndicatorSpec` dataclass, the catalog loader, the path
helpers, and the parquet writer) live in :mod:`leaders_db.ingest.wgi_io`
to break the import cycle. The DB helpers live in
:mod:`leaders_db.ingest.wgi_db` and :mod:`leaders_db.ingest.wgi_db_helpers`.
The orchestrator that ties everything together lives in
:mod:`leaders_db.ingest.wgi`.
"""

from __future__ import annotations

from pathlib import Path

import openpyxl
import pandas as pd

from .wgi_io import default_xlsx_path, load_indicator_catalog

__all__ = ["read_wgi"]


# ---------------------------------------------------------------------------
# Single-sheet reader
# ---------------------------------------------------------------------------


def _read_xlsx_one_sheet(
    xlsx_path: Path, sheet_name: str
) -> tuple[dict[int, int], list[tuple[str, int, list[object]]]]:
    """Read one WGI indicator sheet and return (year_to_estimate_col, rows).

    Args:
        xlsx_path: absolute path to the WGI xlsx.
        sheet_name: xlsx sheet name (one of the catalog's ``raw_column``
            values; e.g. ``"VoiceandAccountability"``).

    Returns:
        A 2-tuple:

        1. ``year_to_estimate_col``: mapping ``{year: zero_indexed_col}``
           for every year present in the sheet, pointing to the
           ``Estimate`` column for that year.
        2. ``rows``: list of ``(iso3, year, [estimates_for_all_years])``
           triples in sheet order. The estimate for a specific year is
           taken from the index ``year_to_estimate_col[year]`` of the
           third element.

    Raises:
        KeyError: if the sheet is absent from the xlsx (the catalog's
            ``raw_column`` has drifted from the live xlsx).
    """
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    try:
        if sheet_name not in wb.sheetnames:
            raise KeyError(
                f"WGI xlsx {xlsx_path} has no sheet named {sheet_name!r}. "
                f"Available sheets: {wb.sheetnames}. The catalog's "
                "`raw_column` may have drifted from the live xlsx."
            )
        ws = wb[sheet_name]

        # Read row 14 (year row) and row 15 (stat row) in one pass.
        year_row: list[object] = []
        stat_row: list[object] = []
        for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if i == 14:
                year_row = list(row)
            elif i == 15:
                stat_row = list(row)
                break  # we have everything we need from the header

        if not year_row or not stat_row:
            raise ValueError(
                f"WGI sheet {sheet_name!r} is missing the year (row 14) "
                "or stat (row 15) header row."
            )

        # Build the year-to-Estimate-column map. The Estimate stat
        # appears once per year block (every 6 columns starting at col
        # index 2). We scan the stat row and for every "Estimate" cell
        # we record the year from the corresponding cell in year_row.
        year_to_estimate_col: dict[int, int] = {}
        for col_idx, stat in enumerate(stat_row):
            if stat == "Estimate":
                year_cell = year_row[col_idx]
                if year_cell is None:
                    continue
                year_to_estimate_col[int(year_cell)] = col_idx

        # Walk country rows (row 16+) and extract (iso3, year,
        # [estimates_for_all_years]) triples. We do this in one pass so
        # we only iterate the sheet once. The estimate for year ``y``
        # is ``row[year_to_estimate_col[y]]``.
        rows: list[tuple[str, int, list[object]]] = []
        for i, row in enumerate(ws.iter_rows(min_row=16, values_only=True), start=16):
            # Defensive: stop at the first all-None row (some WGI
            # builds have a trailing blank).
            if row is None or all(c is None for c in row):
                continue
            iso3 = row[1] if len(row) > 1 else None
            if not iso3:
                # Country row with no ISO3: skip (defensive; the live
                # xlsx has no such rows but the read is robust to it).
                continue
            rows.append((str(iso3), int(i), list(row)))
        return year_to_estimate_col, rows
    finally:
        wb.close()


# ---------------------------------------------------------------------------
# Read orchestrator
# ---------------------------------------------------------------------------


def read_wgi(
    *,
    year: int | None = None,
    xlsx_path: Path | None = None,
    catalog_path: Path | None = None,
) -> pd.DataFrame:
    """Read WGI from the xlsx and pivot to wide format.

    Steps:

    1. Load the catalog.
    2. Open the xlsx at ``xlsx_path`` (default: data-lake path).
    3. For each catalog row (one per indicator):
       a. Open the sheet named in ``raw_column``.
       b. Read row 14 (year row) and row 15 (stat row) to build the
          year-to-Estimate-column map.
       c. For each country row (rows 16..229 in the live xlsx): extract
          ``Code`` (col 2) and the ``Estimate`` cell for each year.
          Coerce ``"#N/A"`` -> ``None`` (and any other missing sentinel
          per the WGI/V-Dem/WDI convention); coerce numeric cells to
          ``float``.
       d. Append (iso3, year, variable_name, value) rows to a long
          frame.
    4. Concatenate per-indicator long frames.
    5. Pivot to wide format (one row per ``(iso3, year)``, one column
       per catalog ``variable_name``).
    6. Coerce the ``year`` column to ``int`` and the indicator columns
       to ``float`` (NaN for absent values).

    Args:
        year: filter to a single year (e.g. ``2022``). Default: all years
            present in the xlsx (1996-2022, 24 distinct years).
        xlsx_path: override the input xlsx. Default: data-lake path.
        catalog_path: override the indicator catalog. Default: checked-in.

    Returns:
        A pandas DataFrame with columns ``iso3``, ``year``, then one
        column per catalog indicator (named with the ``variable_name``).
        ``year`` is integer. Indicator columns are float (``NaN`` = missing).
        WGI does not return aggregate codes, so the returned DataFrame
        contains all rows the xlsx holds for the requested year(s).

    Raises:
        FileNotFoundError: if the xlsx is missing.
        KeyError: if a catalog ``raw_column`` sheet name is absent from
            the xlsx (i.e. the WGI release dropped or renamed a sheet).
    """
    path = xlsx_path or default_xlsx_path()
    if not path.is_file():
        raise FileNotFoundError(f"WGI xlsx not found: {path}")

    specs = load_indicator_catalog(catalog_path=catalog_path)

    long_frames: list[pd.DataFrame] = []
    for spec in specs:
        sheet_name = spec.raw_column
        year_to_col, country_rows = _read_xlsx_one_sheet(path, sheet_name)

        # If the caller asked for a single year, drop the other years
        # early to keep the long frame small.
        if year is not None:
            wanted_years = {int(year)}
        else:
            wanted_years = set(year_to_col.keys())

        records: list[dict[str, object]] = []
        for iso3, _row_idx, row_values in country_rows:
            for one_year in wanted_years:
                col_idx = year_to_col.get(one_year)
                if col_idx is None or col_idx >= len(row_values):
                    continue
                cell = row_values[col_idx]
                records.append(
                    {
                        "iso3": iso3,
                        "year": one_year,
                        "variable_name": spec.variable_name,
                        "value": cell,
                    }
                )
        if records:
            long_frames.append(pd.DataFrame.from_records(records))

    if not long_frames:
        # Nothing to return -- emit an empty frame with the expected
        # columns so downstream code does not have to special-case an
        # empty result.
        return pd.DataFrame(
            columns=["iso3", "year", "variable_name", "value"]
        ).drop(columns=["variable_name", "value"])

    long_df = pd.concat(long_frames, ignore_index=True)

    # Pivot to wide format: one row per (iso3, year), one column per
    # variable_name.
    wide = long_df.pivot_table(
        index=["iso3", "year"],
        columns="variable_name",
        values="value",
        aggfunc="first",
    )
    wide = wide.reset_index()

    # Type coercion. ``year`` is already int (we constructed it that
    # way). Indicator columns may be numeric, string (``"#N/A"``), or
    # ``None``; coerce to float and let ``pd.to_numeric`` produce NaN
    # for the non-numeric cells.
    wide["year"] = wide["year"].astype(int)
    for col in list(wide.columns):
        if col in {"iso3", "year"}:
            continue
        wide[col] = pd.to_numeric(wide[col], errors="coerce").astype(float)

    # Preserve the pre-coercion long frame in ``df.attrs`` so the DB
    # write can recover the original cell text for the ``raw_value``
    # audit-trail column. WGI's missing-data convention is the literal
    # string ``"#N/A"``; after ``pd.to_numeric(errors="coerce")`` those
    # cells become NaN, which the ``_raw_value_to_string`` helper would
    # render as ``"nan"``. The audit trail must keep the literal
    # ``"#N/A"`` so future readers of the source_observations table can
    # tell what the xlsx actually said. The attr key is private
    # (underscore prefix) so it does not leak into the parquet.
    wide.attrs["_wgi_raw_long"] = long_df.copy()

    return wide
