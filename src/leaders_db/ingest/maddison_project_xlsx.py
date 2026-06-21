"""Stage 2 -- Maddison Project Database 2023 xlsx read.

This module is the xlsx-read half of the Maddison Project adapter.
It owns:

- :func:`read_maddison_project` -- the read orchestrator. Opens the
  ``Full data`` sheet with ``openpyxl.read_only=True``, validates
  the required column set, walks the data rows in a single pass,
  applies the year filter, computes the derived GDP total for the
  catalog's derived indicator, and returns the narrow-format
  DataFrame (one row per ``(countrycode, year, variable_name)``
  triple with the ``raw_value`` preserved).

Maddison xlsx layout (verified live 2026-06-20 against the 4.9 MB
release): the workbook has 7 sheets (``Notes``, ``Sources``,
``GDPpc``, ``Population``, ``Full data``, ``Regional data``,
``Maddison original sources``). The Stage 2 contract reads ONLY the
``Full data`` sheet, which has 6 columns: ``countrycode``,
``country``, ``region``, ``year``, ``gdppc``, ``pop``. The other
sheets are the same data restructured as per-indicator wide tabs
and are not used by the Stage 2 adapter.

Long-format output:

The Maddison xlsx is already in long format per ``(countrycode,
year)`` -- one row per ``(countrycode, year)`` with the two
indicator columns (``gdppc`` and ``pop``) in cells. The Stage 2
adapter reshapes this to the canonical narrow ``(countrycode,
year, variable_name, raw_value, normalized_value)`` shape (one row
per non-NULL catalog indicator per country-year) so the DB write
matches the WGI / BTI / CIRIGHTS pattern. The derived total GDP
indicator (``maddison_project_gdp_total_2011_intl_derived``) is
emitted ONLY when both ``gdppc`` and ``pop`` are non-NULL for the
same country-year.

The derived GDP computation:

``derived_gdp_total = gdppc * pop * 1000`` (the ``*1000`` lifts
``pop`` from thousands of persons to absolute persons). The
literal ``*1000`` is the documented Maddison convention (pop is in
thousands). The value is labelled "derived Maddison historical real
GDP" in the ``unit`` column -- NOT current USD.

Constants (the source key, the Maddison attribution, the
:class:`IndicatorSpec` dataclass, the catalog loader, the path
helpers, and the parquet writer) live in
:mod:`leaders_db.ingest.maddison_project_io` to break the import
cycle. The DB helpers live in :mod:`leaders_db.ingest.maddison_project_db`
and :mod:`leaders_db.ingest.maddison_project_db_helpers`. The
orchestrator that ties everything together lives in
:mod:`leaders_db.ingest.maddison_project`.
"""

from __future__ import annotations

from pathlib import Path

import openpyxl
import pandas as pd

from .maddison_project_io import (
    MADDISON_PROJECT_DERIVED_GDP_RAW_COLUMN,
    MADDISON_PROJECT_SHEET_NAME,
    MADDISON_PROJECT_XLSX_COLUMNS,
    default_xlsx_path,
    load_indicator_catalog,
)

__all__ = ["read_maddison_project"]


# ---------------------------------------------------------------------------
# Read orchestrator
# ---------------------------------------------------------------------------


def read_maddison_project(
    *,
    year: int | None = None,
    xlsx_path: Path | None = None,
    catalog_path: Path | None = None,
) -> pd.DataFrame:
    """Read the Maddison Project xlsx and return the narrow-format
    frame.

    Steps:

    1. Load the catalog (3 indicators: ``gdppc``, ``pop``, and the
       derived ``__derived_gdp_total__``).
    2. Open the ``Full data`` sheet with ``openpyxl.read_only=True``.
    3. Validate the required 6 columns are present in the header
       row. Raise :class:`ValueError` with the missing columns if
       the live release drifts.
    4. Walk the data rows in a single pass; apply the year filter
       if ``year=`` was passed; coerce ``countrycode`` to str,
       ``year`` to int, ``gdppc`` / ``pop`` to float-or-NaN.
    5. Compute the derived GDP total for each country-year row
       where both ``gdppc`` and ``pop`` are non-NaN; otherwise the
       derived row is omitted (no observation emitted).
    6. Build the long-format frame: one row per
       ``(countrycode, year, variable_name, raw_value, normalized_value)``
       triple for the two raw indicators + one row for the derived
       indicator (only when derivable). The ``raw_value`` is the
       verbatim cell text (e.g. ``"1234.5"`` or ``"nan"``); the
       ``normalized_value`` is the numeric value (or NaN for
       missing).
    7. Sort by ``(year ASC, countrycode ASC, variable_name ASC)``
       so insertion order is fully deterministic.

    Args:
        year: filter to a single year (e.g. ``2022``). Default: all
            years present in the xlsx (1-2022, ~131,144 rows).
        xlsx_path: override the input xlsx. Default: data-lake path.
        catalog_path: override the catalog. Default: checked-in.

    Returns:
        A pandas DataFrame with columns ``countrycode``, ``year``,
        ``variable_name``, ``raw_value``, ``normalized_value``,
        plus the audit-trail columns ``country``, ``region``,
        ``raw_column``. Long format: one row per non-NULL catalog
        indicator per country-year (so 2 rows per country-year
        when both ``gdppc`` and ``pop`` are present, 3 rows when
        the derived total is also derivable, fewer when one or
        more cells are NaN). The 2022 row for a country with both
        ``gdppc`` and ``pop`` present emits 3 rows (gdppc + pop +
        derived total).

    Raises:
        FileNotFoundError: if the xlsx is missing.
        ValueError: if the ``Full data`` sheet is missing or any
            of the 6 required columns are absent from the header.
    """
    path = xlsx_path or default_xlsx_path()
    if not path.is_file():
        raise FileNotFoundError(
            f"Maddison Project xlsx not found: {path}"
        )

    specs = load_indicator_catalog(catalog_path=catalog_path)
    # Build the raw_column -> spec lookup once so the per-row loop
    # can resolve the spec without an inner-list scan.
    specs_by_raw_column: dict[str, object] = {
        s.raw_column: s for s in specs
    }
    if MADDISON_PROJECT_DERIVED_GDP_RAW_COLUMN not in specs_by_raw_column:
        raise ValueError(
            "Maddison Project catalog is missing the derived GDP "
            "total indicator (raw_column "
            f"{MADDISON_PROJECT_DERIVED_GDP_RAW_COLUMN!r}). "
            "Re-add it to src/leaders_db/ingest/catalogs/"
            "maddison_project.csv."
        )

    wb = openpyxl.load_workbook(
        path, read_only=True, data_only=True,
    )
    try:
        if MADDISON_PROJECT_SHEET_NAME not in wb.sheetnames:
            raise ValueError(
                f"Maddison Project xlsx {path} has no sheet named "
                f"{MADDISON_PROJECT_SHEET_NAME!r}. Available sheets: "
                f"{wb.sheetnames}. The release's sheet name may "
                "have drifted from the expected value."
            )
        ws = wb[MADDISON_PROJECT_SHEET_NAME]

        rows: list[dict[str, object]] = []
        header: list[str] | None = None
        required_cols = set(MADDISON_PROJECT_XLSX_COLUMNS)
        for row in ws.iter_rows(values_only=True):
            row_list = list(row)
            if header is None:
                header = [str(c) if c is not None else "" for c in row_list]
                missing_cols = required_cols - set(header)
                if missing_cols:
                    raise ValueError(
                        f"Maddison Project xlsx {path} is missing "
                        f"required columns: {sorted(missing_cols)}. "
                        f"Header was: {header}"
                    )
                continue
            # Defensive: trailing all-None rows (some xlsx builds add
            # a final blank line).
            if all(c is None for c in row_list):
                continue
            rec: dict[str, object] = {}
            for col_name, value in zip(header, row_list, strict=False):
                rec[col_name] = value
            if year is not None:
                try:
                    row_year = (
                        int(rec.get("year"))
                        if rec.get("year") is not None
                        else None
                    )
                except (TypeError, ValueError):
                    continue
                if row_year != int(year):
                    continue
            rows.append(rec)
    finally:
        wb.close()

    if header is None:
        raise ValueError(
            f"Maddison Project xlsx {path} is empty (no header row)."
        )

    long_records: list[dict[str, object]] = []

    for rec in rows:
        countrycode_raw = rec.get("countrycode")
        if not isinstance(countrycode_raw, str) or not countrycode_raw:
            # Defensive: skip rows with missing/empty ISO3. The live
            # xlsx has no such rows for the 169 real countries but
            # the read is robust to them.
            continue
        iso3 = countrycode_raw.strip()

        try:
            year_value = int(rec.get("year")) if rec.get("year") is not None else None  # type: ignore[arg-type]
        except (TypeError, ValueError):
            continue
        if year_value is None:
            continue

        country = str(rec.get("country") or "")
        region = str(rec.get("region") or "")

        gdppc_cell = rec.get("gdppc")
        pop_cell = rec.get("pop")

        gdppc_present = (
            gdppc_cell is not None
            and not (isinstance(gdppc_cell, float) and pd.isna(gdppc_cell))
        )
        pop_present = (
            pop_cell is not None
            and not (isinstance(pop_cell, float) and pd.isna(pop_cell))
        )

        # gdppc observation (raw indicator #1)
        if gdppc_present:
            long_records.append(
                {
                    "countrycode": iso3,
                    "year": int(year_value),
                    "country": country,
                    "region": region,
                    "variable_name": "maddison_project_gdp_per_capita_2011_intl",
                    "raw_column": "gdppc",
                    "raw_value": str(gdppc_cell),
                    "normalized_value": float(gdppc_cell),  # type: ignore[arg-type]
                }
            )

        # pop observation (raw indicator #2)
        if pop_present:
            long_records.append(
                {
                    "countrycode": iso3,
                    "year": int(year_value),
                    "country": country,
                    "region": region,
                    "variable_name": "maddison_project_population_thousands",
                    "raw_column": "pop",
                    "raw_value": str(pop_cell),
                    "normalized_value": float(pop_cell),  # type: ignore[arg-type]
                }
            )

        # Derived GDP total (indicator #3) -- only emitted when BOTH
        # gdppc and pop are non-NULL for the same country-year.
        if gdppc_present and pop_present:
            derived_total = (
                float(gdppc_cell)  # type: ignore[arg-type]
                * float(pop_cell)  # type: ignore[arg-type]
                * 1000.0
            )
            long_records.append(
                {
                    "countrycode": iso3,
                    "year": int(year_value),
                    "country": country,
                    "region": region,
                    "variable_name": (
                        "maddison_project_gdp_total_2011_intl_derived"
                    ),
                    "raw_column": MADDISON_PROJECT_DERIVED_GDP_RAW_COLUMN,
                    "raw_value": f"{derived_total:.6f}",
                    "normalized_value": derived_total,
                }
            )

    if not long_records:
        # No rows matched. Return an empty narrow frame with the
        # expected column shape so downstream code does not have
        # to special-case an empty result.
        empty_cols = [
            "countrycode", "year", "country", "region",
            "variable_name", "raw_column",
            "raw_value", "normalized_value",
        ]
        return pd.DataFrame(columns=empty_cols)

    narrow = pd.DataFrame.from_records(long_records)
    # Sort for deterministic insertion order. mergesort is stable;
    # year ASC, countrycode ASC, variable_name ASC is the canonical
    # Stage 2 ordering.
    narrow = narrow.sort_values(
        by=["year", "countrycode", "variable_name"],
        ascending=[True, True, True],
        kind="mergesort",
    ).reset_index(drop=True)

    # Carry the year_window + country_count in df.attrs so the
    # orchestrator can surface them on the result model without
    # re-iterating the frame.
    if not narrow.empty:
        year_window: tuple[int, int] = (
            int(narrow["year"].min()),
            int(narrow["year"].max()),
        )
        country_count = int(narrow["countrycode"].nunique())
    else:
        year_window = (0, 0)
        country_count = 0
    narrow.attrs["year_window"] = year_window
    narrow.attrs["country_count"] = country_count
    narrow.attrs["derived_gdp_total_emitted"] = int(
        (narrow["raw_column"] == MADDISON_PROJECT_DERIVED_GDP_RAW_COLUMN).sum()
    )

    return narrow
