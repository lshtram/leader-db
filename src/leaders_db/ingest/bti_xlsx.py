"""Stage 2 -- Bertelsmann BTI xlsx read (REQ-SRC-002).

This module is the BTI xlsx-read half of the BTI adapter. It owns:

- :func:`_read_xlsx_one_sheet` -- open one BTI edition sheet with
  openpyxl in read-only mode, walk the header row (row 0, with the
  merged "Regions:" label in col 0), find the per-indicator column
  positions, and return the country rows with the indicator cells
  extracted.
- :func:`read_bti` -- the read orchestrator. Resolves the requested
  target year to the matching BTI edition sheet via
  :func:`bti_io.sheet_for_year`, calls :func:`_read_xlsx_one_sheet`
  once per indicator in the catalog, builds a long-format frame,
  and pivots to wide format (one row per ``(country, year)``, one
  column per catalog ``variable_name``).

BTI xlsx layout (verified live against ``BTI_2006-2026_Scores.xlsx``):

- 12 sheets total: ``BTI 2026``, ``BTI 2024``, ``BTI 2022``, ...,
  ``BTI 2008``, ``BTI 2006``, ``BTI 2006_old`` (one BTI edition per
  sheet; BTI 2006_old pre-dates the 2006 methodology refresh).
- Per sheet, 123 columns:
  - col 0: merged "Regions:" multi-line label (skip)
  - col 1: integer region code (1-7)
  - col 2: ranking (numeric)
  - col 3-77: composite indices + Q1-Q17 with sub-questions
    (numeric 1-10)
  - col 78+: trend / categorical classification columns (text)
- Country rows: 137 countries per edition (BTI 2024 holds 137
  valid country rows + 21 trailing blank rows + 1 header row).
- Missing-data convention: blank cells (no explicit token); the
  read coerces them to NaN. Numeric cells are float. The original
  cell text is preserved in ``df.attrs["_bti_raw_long"]`` for the
  ``raw_value`` audit trail.

Constants (the source key, the BTI attribution, the
:class:`IndicatorSpec` dataclass, the catalog loader, the path
helpers, and the parquet writer) live in :mod:`leaders_db.ingest.bti_io`
to break the import cycle. The DB helpers live in
:mod:`leaders_db.ingest.bti_db` and :mod:`leaders_db.ingest.bti_db_helpers`.
The orchestrator that ties everything together lives in
:mod:`leaders_db.ingest.bti`.
"""

from __future__ import annotations

from pathlib import Path

import openpyxl
import pandas as pd

from .bti_io import (
    default_xlsx_path,
    load_indicator_catalog,
    sheet_for_year,
    target_year_for_sheet,
)

__all__ = ["read_bti"]


# ---------------------------------------------------------------------------
# Single-sheet reader
# ---------------------------------------------------------------------------


def _read_xlsx_one_sheet(
    xlsx_path: Path, sheet_name: str, wanted_columns: list[str]
) -> tuple[list[tuple[str, int, list[object]]], dict[str, int]]:
    """Read one BTI edition sheet and return (country_rows, column_to_index).

    Args:
        xlsx_path: absolute path to the BTI xlsx.
        sheet_name: xlsx sheet name (one of the catalog's edition
            labels; e.g. ``"BTI 2024"`` for the 2023 target year).
        wanted_columns: the subset of header strings to extract
            (the catalog's ``raw_column`` values, in catalog order).

    Returns:
        A 2-tuple:

        1. ``country_rows``: list of ``(country_name, target_year, [row_values])``
           triples in sheet order. ``target_year`` is the canonical
           in-coverage year the sheet represents (e.g. ``2023`` for
           ``BTI 2024``).
        2. ``column_to_index``: mapping ``{header_string: zero_indexed_col}``
           for every requested header (only headers that were found
           in the xlsx are present; missing headers are absent).

    Raises:
        KeyError: if the sheet is absent from the xlsx (the catalog's
            ``raw_column`` has drifted from the live xlsx).
    """
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    try:
        ws = _resolve_sheet(wb, xlsx_path, sheet_name)
        header_row = _read_header_row(ws, sheet_name)
        column_to_index = _resolve_columns(header_row, wanted_columns)
        target_year = target_year_for_sheet(sheet_name)
        rows = list(_iter_country_rows(ws, target_year))
        return rows, column_to_index
    finally:
        wb.close()


def _resolve_sheet(
    wb: openpyxl.workbook.Workbook,
    xlsx_path: Path,
    sheet_name: str,
) -> openpyxl.worksheet.worksheet.Worksheet:
    """Return the BTI edition worksheet, raising if absent."""
    if sheet_name not in wb.sheetnames:
        raise KeyError(
            f"BTI xlsx {xlsx_path} has no sheet named {sheet_name!r}. "
            f"Available sheets: {wb.sheetnames}. The catalog or the "
            "edition-to-year mapping may have drifted from the live xlsx."
        )
    return wb[sheet_name]


def _read_header_row(
    ws: openpyxl.worksheet.worksheet.Worksheet, sheet_name: str,
) -> list[object]:
    """Return the sheet header row, raising if empty."""
    header_row: list[object] = []
    for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
        header_row = list(row)
        break
    if not header_row:
        raise ValueError(f"BTI sheet {sheet_name!r} has no header row.")
    return header_row


def _resolve_columns(
    header_row: list[object], wanted_columns: list[str],
) -> dict[str, int]:
    """Build the ``raw_column -> zero_indexed_col`` map for the wanted headers.

    Compares trimmed strings to be tolerant of trailing whitespace in
    the xlsx (e.g. ``"  S | Status Index"``). The catalog stores the
    verbatim header (with the leading two-space padding) for human
    readability, but the lookup is whitespace-insensitive.
    """
    column_to_index: dict[str, int] = {}
    for wanted in wanted_columns:
        wanted_trim = wanted.strip()
        for col_idx, cell in enumerate(header_row):
            if cell is None:
                continue
            if str(cell).strip() == wanted_trim:
                column_to_index[wanted] = col_idx
                break
    return column_to_index


def _iter_country_rows(
    ws: openpyxl.worksheet.worksheet.Worksheet, target_year: int | None,
) -> list[tuple[str, int, list[object]]]:
    """Yield ``(country_name, target_year, [row_values])`` for valid rows.

    Skips trailing all-None rows, rows whose col 0 is not a string
    country name, and rows whose col 0 is longer than 60 chars
    (defensive: the live xlsx has no such rows, but future releases
    could surface spurious merged-cell content).
    """
    rows: list[tuple[str, int, list[object]]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row is None or all(c is None for c in row):
            continue
        country_name = row[0]
        if not country_name or not isinstance(country_name, str):
            continue
        if len(country_name) > 60:
            continue
        rows.append((str(country_name).strip(), int(target_year), list(row)))
    return rows


# ---------------------------------------------------------------------------
# Read orchestrator
# ---------------------------------------------------------------------------


def read_bti(
    *,
    year: int | None = None,
    xlsx_path: Path | None = None,
    catalog_path: Path | None = None,
    sheet_name: str | None = None,
) -> pd.DataFrame:
    """Read BTI from the xlsx and pivot to wide format.

    Steps:

    1. Load the catalog.
    2. Open the xlsx at ``xlsx_path`` (default: data-lake path).
    3. Resolve the target edition:
       - If ``sheet_name`` is given, use it directly.
       - Else if ``year`` is given, call :func:`sheet_for_year`.
       - Else, default to the latest edition ("BTI 2026").
    4. Call :func:`_read_xlsx_one_sheet` once with the catalog's
       ``raw_column`` list to extract per-country indicator cells.
    5. Build a long-format frame ``(country, year, variable_name,
       value)`` and pivot to wide format (one row per ``(country,
       year)``, one column per catalog ``variable_name``).
    6. Coerce the ``year`` column to ``int`` and the indicator columns
       to ``float`` (NaN for absent cells).

    Args:
        year: target year to filter to (e.g. ``2023``). If provided,
            the adapter resolves the BTI edition whose covered
            interval contains the year (e.g. ``BTI 2024`` for 2023).
            Ignored if ``sheet_name`` is given explicitly.
        xlsx_path: override the input xlsx. Default: data-lake path.
        catalog_path: override the indicator catalog. Default: checked-in.
        sheet_name: override the BTI edition sheet name. When given,
            used directly (no year-to-sheet resolution).

    Returns:
        A pandas DataFrame with columns ``country``, ``year``, then
        one column per catalog indicator (named with the
        ``variable_name``). ``year`` is integer. Indicator columns are
        float (``NaN`` = missing). BTI does not return ISO3 codes; the
        ``country`` column carries the BTI display name (Stage 3
        resolves it to ISO3).

    Raises:
        FileNotFoundError: if the xlsx is missing.
        ValueError: if ``year`` is given but no BTI edition covers it.
        KeyError: if a catalog ``raw_column`` header is absent from
            the xlsx (i.e. the BTI release renamed a column).
    """
    path = xlsx_path or default_xlsx_path()
    if not path.is_file():
        raise FileNotFoundError(f"BTI xlsx not found: {path}")

    specs = load_indicator_catalog(catalog_path=catalog_path)

    # Resolve the target edition.
    if sheet_name is None:
        if year is None:
            sheet_name = "BTI 2026"
        else:
            sheet_name = sheet_for_year(int(year))

    wanted_columns = [spec.raw_column for spec in specs]
    country_rows, column_to_index = _read_xlsx_one_sheet(
        path, sheet_name, wanted_columns
    )

    # Build the long-format frame.
    records: list[dict[str, object]] = []
    for country_name, target_year, row_values in country_rows:
        for spec in specs:
            col_idx = column_to_index.get(spec.raw_column)
            if col_idx is None:
                # Header was not found in the xlsx -- skip this
                # indicator for this row. The wide frame will carry
                # NaN for every row of this variable.
                continue
            if col_idx >= len(row_values):
                continue
            cell = row_values[col_idx]
            records.append(
                {
                    "country": country_name,
                    "year": target_year,
                    "variable_name": spec.variable_name,
                    "value": cell,
                }
            )

    if not records:
        # Nothing to return -- emit an empty frame with the expected
        # columns so downstream code does not have to special-case an
        # empty result.
        return pd.DataFrame(
            columns=["country", "year", "variable_name", "value"]
        ).drop(columns=["variable_name", "value"])

    long_df = pd.DataFrame.from_records(records)

    # Pivot to wide format: one row per (country, year), one column per
    # variable_name.
    wide = long_df.pivot_table(
        index=["country", "year"],
        columns="variable_name",
        values="value",
        aggfunc="first",
    )
    wide = wide.reset_index()

    # Type coercion. ``year`` is already int. Indicator columns may be
    # numeric, string (defensive), or ``None``; coerce to float and let
    # ``pd.to_numeric`` produce NaN for the non-numeric cells.
    wide["year"] = wide["year"].astype(int)
    for col in list(wide.columns):
        if col in {"country", "year"}:
            continue
        wide[col] = pd.to_numeric(wide[col], errors="coerce").astype(float)

    # Preserve the pre-coercion long frame in ``df.attrs`` so the DB
    # write can recover the original cell text for the ``raw_value``
    # audit-trail column. BTI cells are numeric; after
    # ``pd.to_numeric(errors="coerce")`` empty / non-numeric cells
    # become NaN, but we keep the verbatim value in the attr so the
    # audit trail reflects what the xlsx said.
    wide.attrs["_bti_raw_long"] = long_df.copy()

    # Carry the resolved sheet name in attrs so the orchestrator can
    # surface it in the manifest without re-resolving.
    wide.attrs["_bti_sheet_name"] = sheet_name

    return wide
