"""Stage 2 -- CIRI Human Rights Data Project (CIRIGHTS) xlsx read.

This module is the xlsx-read half of the CIRIGHTS adapter. It owns:

- :func:`_empty_wide` -- build an empty wide frame with the expected
  column shape (used by the no-rows and out-of-range-year short
  circuits).
- :func:`read_xlsx_to_wide_dataframe` -- open the single ``Sheet1``
  sheet with ``openpyxl.read_only=True``; narrow to the 7 catalog
  columns; pivot to wide format. Carries ``_cirights_raw_lookup`` and
  ``year_window`` in ``df.attrs`` for the downstream DB write and
  the audit trail.
- :func:`read_cirights` -- the read orchestrator: thin wrapper over
  :func:`read_xlsx_to_wide_dataframe` that resolves the default
  xlsx path and loads the catalog.

CIRIGHTS xlsx layout (verified live 2026-06-18 against the real
``data/raw/cirights/cirights_v3.12.10.24.xlsx``): 1 sheet ``Sheet1``,
50 columns, 7,931 data rows + 1 header. The Stage 2 adapter narrows
to 9 columns: the 2 identity columns (``country``, ``year``) + the
7 catalog ``raw_column`` indicators.

The cell coercion (``_coerce_cirights_value``) and the audit-trail
text helper (``_raw_cell_text``) live in
:mod:`leaders_db.ingest.cirights_xlsx_pivot` (the long-to-wide
pivot / test-seam module). The in-memory test seam
(:func:`read_cirights_from_dataframe`) also lives there.

Constants (the source key, the CIRIGHTS attribution, the
:class:`IndicatorSpec` dataclass, the catalog loader, the path
helpers, and the parquet writer) live in
:mod:`leaders_db.ingest.cirights_io` to break the import cycle. The
DB helpers live in :mod:`leaders_db.ingest.cirights_db` and
:mod:`leaders_db.ingest.cirights_db_helpers`. The orchestrator
lives in :mod:`leaders_db.ingest.cirights`.
"""

from __future__ import annotations

import logging
from pathlib import Path

import openpyxl
import pandas as pd

from .cirights_io import IndicatorSpec, default_xlsx_path, load_indicator_catalog
from .cirights_xlsx_pivot import (
    _empty_wide,
    read_cirights_from_dataframe,
)

_logger = logging.getLogger(__name__)

__all__ = [
    "read_cirights",
    "read_xlsx_to_wide_dataframe",
]


# ---------------------------------------------------------------------------
# xlsx reader
# ---------------------------------------------------------------------------


def read_xlsx_to_wide_dataframe(
    xlsx_path: Path,
    specs: list[IndicatorSpec],
    *,
    year: int | None = None,
) -> pd.DataFrame:
    """Open the CIRIGHTS xlsx and return the wide-format frame.

    Opens the single sheet ``Sheet1`` with
    ``openpyxl.read_only=True``, walks the 50-column structure in a
    single pass, narrows to the 2 identity columns (``country``,
    ``year``) + the 7 catalog ``raw_column`` indicators, coerces each
    cell, and pivots to wide format. The wide frame carries
    ``_cirights_raw_lookup`` (a per-row, per-variable lookup of
    pre-coercion cell text) and ``year_window`` in ``df.attrs``.

    The xlsx is already in long format per country-year (one row per
    ``(country, year)``, indicator columns in cells). The "pivot" is
    therefore a column rename + per-cell coercion, not a reshape.
    The per-cell coercion, raw_lookup construction, and
    ``_empty_wide`` helper live in
    :mod:`leaders_db.ingest.cirights_xlsx_pivot` to keep this
    module under the 400-line convention from
    :file:`docs/process/coding-guidelines.md`.

    Args:
        xlsx_path: absolute path to the CIRIGHTS xlsx.
        specs: the catalog specs (1+ per indicator). Provides the
            ``raw_column`` -> ``variable_name`` rename and the order
            of the indicator columns in the wide frame.
        year: filter to a single year. Default: all years.

    Returns:
        A pandas DataFrame with columns ``country``, ``year``, then
        one column per catalog ``variable_name`` in catalog order.
        ``year`` is int; indicator columns are ``Int64`` (nullable).
        The wide frame is dense: every ``(country, year)`` row is
        present, even when all 7 indicator cells are missing.

    Raises:
        FileNotFoundError: if the xlsx is missing.
        ValueError: if the sheet name is not ``Sheet1`` or a
            catalog ``raw_column`` is missing from the xlsx header.
    """
    if not xlsx_path.is_file():
        raise FileNotFoundError(f"CIRIGHTS xlsx not found: {xlsx_path}")
    expected_sheet = "Sheet1"
    wb = openpyxl.load_workbook(
        xlsx_path, read_only=True, data_only=True,
    )
    try:
        if expected_sheet not in wb.sheetnames:
            raise ValueError(
                f"CIRIGHTS xlsx {xlsx_path} has no sheet named "
                f"{expected_sheet!r}. Available sheets: {wb.sheetnames}. "
                "The xlsx's sheet name may have drifted from the "
                "expected value."
            )
        ws = wb[expected_sheet]

        # Collect the header + per-row data in a single pass.
        rows: list[dict[str, object]] = []
        header: list[str] | None = None
        raw_columns: set[str] = {spec.raw_column for spec in specs}
        for row in ws.iter_rows(values_only=True):
            row_list = list(row)
            if header is None:
                header = [str(c) if c is not None else "" for c in row_list]
                # Validate that every catalog raw_column is present
                # in the xlsx header. The catalog is the source of
                # truth; if a column has been renamed in the live
                # xlsx, the catalog must be updated to match (drift
                # guard). This check fires before the data walk so
                # the error is surfaced at the first failure.
                missing_cols = raw_columns - set(header)
                if missing_cols:
                    raise ValueError(
                        f"CIRIGHTS xlsx {xlsx_path} is missing "
                        f"expected indicator columns: {sorted(missing_cols)}. "
                        f"Header was: {header}"
                    )
                continue
            # Defensive: trailing all-None rows (some xlsx builds add
            # a final blank line).
            if all(c is None for c in row_list):
                continue
            rec: dict[str, object] = {}
            for col_name, value in zip(header, row_list, strict=False):
                rec[col_name] = value
            if year is not None:
                try:
                    row_year = int(rec.get("year")) if rec.get("year") is not None else None
                except (TypeError, ValueError):
                    continue
                if row_year != int(year):
                    continue
            rows.append(rec)
    finally:
        wb.close()

    if header is None:
        raise ValueError(
            f"CIRIGHTS xlsx {xlsx_path} is empty (no header row)."
        )

    if not rows:
        # No rows matched the filter. Return an empty wide frame with
        # the expected column shape so downstream code does not have
        # to special-case an empty result.
        return _empty_wide(specs)

    long_df = pd.DataFrame.from_records(rows, columns=header)

    # Delegate the per-cell coercion + wide pivot + raw_lookup to
    # the pivot module. The pivot function does the column narrow
    # + rename + Int64 cast + per-cell coercion + raw_lookup in
    # one pass. This keeps the xlsx reader focused on the
    # file-format I/O and the per-cell coercion / audit-trail
    # logic in one place.
    return read_cirights_from_dataframe(long_df, specs)


def read_cirights(
    xlsx_path: Path | None = None,
    *,
    year: int | None = None,
    catalog_path: Path | None = None,
) -> pd.DataFrame:
    """Open the CIRIGHTS xlsx and return the wide-format frame.

    Thin wrapper over :func:`read_xlsx_to_wide_dataframe` that
    resolves the default xlsx path and loads the catalog. Mirrors
    the V-Dem / WGI / UCDP / SIPRI milex / SIPRI Yearbook Ch.7 /
    PTS read orchestrator pattern.

    Args:
        xlsx_path: override the input xlsx. Default: data-lake path.
        year: filter to a single year. Default: all 42 years.
        catalog_path: override the catalog. Default: checked-in.

    Returns:
        A wide-format DataFrame (one row per ``(country, year)``,
        one column per catalog ``variable_name``). Indicator columns
        are ``Int64``; the wide frame carries
        ``_cirights_raw_lookup`` and ``year_window`` in ``df.attrs``.

    Raises:
        FileNotFoundError: if the xlsx is missing.
        ValueError: if the sheet name has drifted from ``Sheet1`` or
            a catalog ``raw_column`` is missing from the xlsx header.
    """
    path = xlsx_path or default_xlsx_path()
    if not path.is_file():
        raise FileNotFoundError(f"CIRIGHTS xlsx not found: {path}")
    specs = load_indicator_catalog(catalog_path=catalog_path)
    return read_xlsx_to_wide_dataframe(path, specs=specs, year=year)
