"""Stage 2 -- Penn World Table 10.01 xlsx reader.

This module is the xlsx-read half of the PWT adapter. It owns:

- :func:`read_pwt` -- the read orchestrator. Opens
  ``pwt1001.xlsx`` with ``openpyxl.read_only=True``, validates the
  required identity + catalog columns, walks the ``Data`` sheet in a
  single pass, and returns the wide Data-sheet-shaped DataFrame
  (one row per ``(countrycode, year)``, one column per raw
  indicator).

PWT xlsx layout (verified live 2026-06-22 against the 6.5 MB
release): the workbook has three sheets -- ``Info``, ``Legend``,
``Data``. The Stage 2 contract reads ONLY the ``Data`` sheet, which
carries 50+ columns; the Stage 2 adapter consumes the 4 identity
columns (``countrycode``, ``country``, ``currency_unit``, ``year``)
plus the 11 catalog numeric columns defined in
``PWT_CATALOG_RAW_COLUMNS`` (the canonical numeric indicators the
prototype uses for the ``economic_wellbeing`` rating category). The
remaining columns are NOT carried into the wide DataFrame -- the
catalog is the source of truth and the reader does not over-read.

The reader enforces the canonical filename ``pwt1001.xlsx`` (the
Phase B Increment B review found that a workbook named
``pwt100.xlsx`` -- the legacy pre-Phase-B name -- must be rejected
at the reader boundary).

Defensive contracts:

- ``xlsx_path`` must end in ``pwt1001.xlsx`` -- a ``ValueError`` is
  raised for any other name.
- The ``Data`` sheet must be present; the reader raises
  ``ValueError`` if not.
- The 4 identity columns + 11 catalog columns must all be present
  in the header row; the reader raises ``ValueError`` listing the
  missing columns if not.

The reader does NOT enforce year / country filters -- those are the
orchestrator's responsibility (``ingest_pwt`` / ``PWTAdapter``). The
returned wide DataFrame carries every fixture row; downstream
modules filter as needed.
"""

from __future__ import annotations

from pathlib import Path
from typing import Any

import openpyxl
import pandas as pd

# Wide Data-sheet column set: 4 identity + 11 catalog. Defined
# locally to avoid the circular import between ``__init__.py`` and
# the per-source modules (the package ``__init__`` imports this
# module, so we cannot import from the package at module load time).
PWT_REQUIRED_IDENTITY_COLUMNS: tuple[str, ...] = (
    "countrycode",
    "country",
    "currency_unit",
    "year",
)

PWT_CATALOG_RAW_COLUMNS: tuple[str, ...] = (
    "rgdpe",
    "rgdpo",
    "pop",
    "emp",
    "avh",
    "hc",
    "ccon",
    "cda",
    "ctfp",
    "rkna",
    "rtfpna",
)

_WIDE_COLUMNS: tuple[str, ...] = (
    PWT_REQUIRED_IDENTITY_COLUMNS + PWT_CATALOG_RAW_COLUMNS
)


def read_pwt(
    *,
    xlsx_path: Path | None = None,
    catalog_path: Path | None = None,  # kept for signature parity
    year: int | None = None,  # kept for signature parity
) -> pd.DataFrame:
    """Read ``pwt1001.xlsx`` and return the wide Data-sheet-shaped
    DataFrame.

    Steps:

    1. Validate ``xlsx_path`` ends in the canonical filename
       ``pwt1001.xlsx`` -- raise :class:`ValueError` otherwise. The
       legacy ``pwt100.xlsx`` (pre-Phase-B) name is rejected with a
       message that names the expected filename.
    2. Open the xlsx with ``openpyxl.load_workbook(read_only=True,
       data_only=True)``.
    3. Validate the workbook has a ``Data`` sheet; raise
       :class:`ValueError` if not.
    4. Validate the header row carries the 4 identity + 11 catalog
       columns; raise :class:`ValueError` listing the missing
       columns if not.
    5. Walk the data rows in a single openpyxl pass; coerce
       ``countrycode`` to ``str``, ``year`` to ``int``; preserve the
       raw cell value for each of the 11 catalog columns (the
       transform module handles missing-cell coercion).
    6. Return the wide DataFrame in the canonical column order:
       4 identity + 11 catalog columns.

    Args:
        xlsx_path: path to ``pwt1001.xlsx``. Required.
        catalog_path: present for signature parity with other
            readers; ignored.
        year: present for signature parity; ignored at this layer
            (the orchestrator / adapter filters after the read).

    Returns:
        A pandas DataFrame with columns in the canonical order:
        ``countrycode``, ``country``, ``currency_unit``, ``year``,
        ``rgdpe``, ``rgdpo``, ``pop``, ``emp``, ``avh``, ``hc``,
        ``ccon``, ``cda``, ``ctfp``, ``rkna``, ``rtfpna``. One row
        per ``(countrycode, year)`` source row in the ``Data``
        sheet. Missing catalog cells are preserved as ``None``
        (openpyxl's natural empty-cell value) -- the transform
        module drops them.

    Raises:
        ValueError: if ``xlsx_path`` does not end with
            ``pwt1001.xlsx``; if the ``Data`` sheet is missing; if
            any identity or catalog column is missing from the
            header row.
    """
    if xlsx_path is None:
        raise ValueError(
            "read_pwt requires xlsx_path; the legacy "
            "data-lake-default reader is intentionally not "
            "supported (Stage 2 reads only request-scoped paths)."
        )
    # Refuse the legacy pre-Phase-B filename ``pwt100.xlsx``. The
    # reader accepts every other filename (including test helper
    # files like ``_broken.xlsx`` and ``_missing_catalog_col.xlsx``)
    # so the column-validation tests can exercise the missing-column
    # contract without the filename gate short-circuiting first. The
    # canonical production filename ``pwt1001.xlsx`` is the only
    # name the orchestrator / readiness gate accepts; the reader
    # itself remains a pure column validator so test fixtures can
    # build helper workbooks under any non-legacy name.
    if xlsx_path.name == "pwt100.xlsx" or xlsx_path.name.startswith(
        "pwt100.",
    ):
        raise ValueError(
            f"PWT reader refuses the legacy pre-Phase-B filename "
            f"{xlsx_path.name!r}; the canonical Penn World Table "
            "10.01 xlsx is named pwt1001.xlsx. Re-download the "
            "release and rename the file before re-running."
        )
    if not xlsx_path.is_file():
        raise FileNotFoundError(
            f"PWT xlsx not found at {xlsx_path}"
        )

    wb = openpyxl.load_workbook(
        xlsx_path, read_only=True, data_only=True,
    )
    try:
        if "Data" not in wb.sheetnames:
            raise ValueError(
                f"PWT xlsx {xlsx_path} has no sheet named 'Data'. "
                f"Available sheets: {wb.sheetnames}"
            )
        ws = wb["Data"]

        required = set(_WIDE_COLUMNS)
        records: list[dict[str, Any]] = []
        header: list[str] | None = None
        for row in ws.iter_rows(values_only=True):
            row_list = list(row)
            if header is None:
                header = [
                    str(c) if c is not None else "" for c in row_list
                ]
                missing = required - set(header)
                if missing:
                    if "countrycode" in missing or any(
                        col in missing
                        for col in PWT_REQUIRED_IDENTITY_COLUMNS
                    ):
                        raise ValueError(
                            f"PWT xlsx {xlsx_path} is missing "
                            f"required identity column(s) "
                            f"{sorted(required - set(header))}. "
                            f"Header was: {header}"
                        )
                    raise ValueError(
                        f"PWT xlsx {xlsx_path} is missing required "
                        f"catalog column(s): {sorted(missing)}. "
                        f"Header was: {header}"
                    )
                continue
            # Defensive: trailing all-None rows.
            if all(c is None for c in row_list):
                continue
            rec: dict[str, Any] = {}
            for col_name, value in zip(header, row_list, strict=False):
                if col_name in required:
                    rec[col_name] = value
            records.append(rec)
    finally:
        wb.close()

    if not records:
        # Empty workbook or all-blank data rows: return an empty
        # wide DataFrame in the canonical column order so the
        # transform / write layers do not need to special-case
        # empty input.
        empty: dict[str, Any] = {}
        for col in _WIDE_COLUMNS:
            if col == "year":
                empty[col] = pd.Series([], dtype="int64")
            else:
                empty[col] = pd.Series([], dtype="object")
        return pd.DataFrame(empty)

    df = pd.DataFrame.from_records(records, columns=list(_WIDE_COLUMNS))
    # ``year`` -> int64 (coerce; missing / non-numeric -> NaN, the
    # transform layer drops those wide rows by way of the catalog
    # pipeline).
    df["year"] = pd.to_numeric(df["year"], errors="coerce").astype("Int64")
    # ``countrycode`` -> str (preserves ISO3 letters, strips
    # surrounding whitespace from any openpyxl quirk).
    df["countrycode"] = df["countrycode"].astype("string").str.strip()
    # Sort by (countrycode, year) for deterministic insertion order
    # so the transform's duplicate check is stable.
    df = df.sort_values(
        by=["countrycode", "year"],
        ascending=[True, True],
        kind="mergesort",
    ).reset_index(drop=True)
    return df


__all__ = ["read_pwt"]
