"""Stage 2 -- Political Terror Scale (PTS) xlsx read (REQ-SRC-007).

This module is the PTS xlsx-read half of the PTS adapter. It owns:

- :func:`_coerce_pts_value` -- apply the §6 4-case sentinel matrix
  to a single ``(PTS_X, NA_Status_X)`` pair. Returns int 1-5 for
  valid cells, ``None`` for missing/inconsistent. Includes the
  §6.5 defensive check (an unknown ``NA_Status`` code is logged and
  treated as missing).
- :func:`_raw_cell_text` -- render the original ``PTS_X`` cell text
  for the ``raw_value`` audit column.
- :func:`read_xlsx_to_long_dataframe` -- open the ``PTS-2025`` sheet
  with ``openpyxl.read_only=True``; return a long-format DataFrame
  with the 10 retained columns.
- :func:`read_pts_from_dataframe` -- test seam: takes a pre-loaded
  DataFrame (matching the 14-column xlsx header shape), applies the
  §6 sentinel matrix, and delegates the pivot to
  :func:`pts_xlsx_pivot.pivot_long_to_wide`. Includes the §6.4
  defensive region check.
- :func:`read_pts` -- the read orchestrator.

PTS xlsx layout (verified live 2026-06-18): 1 sheet ``PTS-2025``,
14 columns, 10,531 country-year rows + 1 header. The Stage 2 adapter
uses the 4 identity columns (``Country, COW_Code_A, Year, Region``)
+ the 6 indicator + NA_Status columns; the other 4 columns are
deferred.

The 4-case sentinel matrix (§6.1):

- Case 1: int 1-5 + NA_Status=0 -> valid (keep the int).
- Case 2: int 1-5 + NA_Status != 0 -> drop (NA_Status takes precedence).
- Case 3: 'NA' + NA_Status != 0 -> drop (expected sentinel path).
- Case 4: 'NA' + NA_Status = 0 -> drop + warning (inconsistency;
  e.g. Bahamas 2017 ``PTS_A='NA'`` with ``NA_Status_A=0``).

The long-to-wide pivot lives in :mod:`pts_xlsx_pivot` to keep this
module under the 400-line cap. Constants in :mod:`pts_io`; DB helpers
in :mod:`pts_db` + :mod:`pts_db_helpers`; orchestrator in :mod:`pts`.
"""

from __future__ import annotations

import logging
from pathlib import Path

import openpyxl
import pandas as pd

from .pts_io import (
    PTS_NA_STATUS_CODES,
    PTS_REGION_CODES,
    IndicatorSpec,
    default_xlsx_path,
    load_indicator_catalog,
)
from .pts_xlsx_pivot import pivot_long_to_wide

_logger = logging.getLogger(__name__)
__all__ = [
    "read_pts",
    "read_pts_from_dataframe",
    "read_xlsx_to_long_dataframe",
]

# Canonical 3-indicator metadata. Mirrored as module-level constants
# (rather than imported) so the test seam can construct a valid wide
# frame without instantiating IndicatorSpec objects.
_PTS_INDICATOR_COLS: tuple[tuple[str, str], ...] = (
    ("PTS_A", "pts_amnesty_score"),
    ("PTS_H", "pts_human_rights_watch_score"),
    ("PTS_S", "pts_state_dept_score"),
)

# NA_Status column per indicator (paired with PTS_X).
_NA_STATUS_COLS: dict[str, str] = {
    "PTS_A": "NA_Status_A",
    "PTS_H": "NA_Status_H",
    "PTS_S": "NA_Status_S",
}


# Single-cell coercion (the §6 sentinel matrix + §6.5 defensive check).
def _coerce_pts_value(
    pts_cell: object,
    na_status: int,
    *,
    country: str,
    year: int,
    indicator: str,
) -> int | None:
    """Apply the 4-case §6 sentinel matrix with the §6.5 defensive check.

    Returns the int 1-5 for valid cells, ``None`` for missing or
    inconsistent cells. Logs a warning for the case-4 inconsistency
    and for the §6.5 unknown ``NA_Status`` code.

    Precedence rule (per §6): **NA_Status takes precedence over
    PTS_X**. A cell is valid iff ``NA_Status == 0`` AND ``PTS_X`` is
    an int in 1-5. Any other combination drops the indicator.

    Cases:
      1. int 1-5 + NA_Status=0  -> valid; return the int.
      2. int 1-5 + NA_Status!=0 -> drop (NA_Status confirms missing).
      3. 'NA' + NA_Status!=0    -> drop (expected sentinel path).
      4. 'NA' + NA_Status=0     -> drop + warning (inconsistency).

    §6.5 defensive check: an unknown ``NA_Status`` code (one not in
    :data:`pts_io._PTS_NA_STATUS_CODES`, the 5 known codes
    ``0/66/77/88/99``) is logged at WARNING and treated as missing.

    Args:
        pts_cell: the ``PTS_X`` cell value (int 1-5 or str 'NA';
            defensive for unexpected types).
        na_status: the paired ``NA_Status_X`` integer.
        country, year, indicator: used in the case-4 and §6.5
            warning messages.

    Returns:
        The int 1-5 for valid cells; ``None`` otherwise.
    """
    # §6.5 defensive check: an unknown NA_Status code is logged and
    # treated as missing. A future xlsx release that introduces a new
    # code (e.g., the hypothetical 55 per architecture §6.5) will
    # surface here rather than silently dropping the cell or raising.
    if na_status not in PTS_NA_STATUS_CODES:
        _logger.warning(
            "PTS unknown NA_Status code: country=%s year=%d "
            "indicator=%s na_status=%s. Treating as missing.",
            country, year, indicator, na_status,
        )
        return None
    if na_status != 0:
        return None  # Cases 2 and 3: NA_Status takes precedence.
    if isinstance(pts_cell, bool):
        # bool is a subclass of int in Python; exclude so True/False
        # are not coerced to 1/0 (a bug, not data).
        _logger.warning(
            "PTS unexpected cell value (bool): country=%s year=%d "
            "indicator=%s pts_cell=%r na_status=%d. Treating as missing.",
            country, year, indicator, pts_cell, na_status,
        )
        return None
    if isinstance(pts_cell, int) and 1 <= pts_cell <= 5:
        return pts_cell  # Case 1: valid data.
    if isinstance(pts_cell, str) and pts_cell == "NA":
        # Case 4: inconsistency. Log and treat as missing.
        _logger.warning(
            "PTS data inconsistency: country=%s year=%d indicator=%s "
            "has PTS_X='NA' with NA_Status=0. Treating as missing.",
            country, year, indicator,
        )
        return None
    # Anything else (float, unexpected string). Log and treat as missing.
    _logger.warning(
        "PTS unexpected cell value: country=%s year=%d indicator=%s "
        "pts_cell=%r na_status=%d. Treating as missing.",
        country, year, indicator, pts_cell, na_status,
    )
    return None


def _raw_cell_text(pts_cell: object) -> str | None:
    """Render the original ``PTS_X`` cell text for the ``raw_value``
    audit column.

    Per the §6.3 audit-trail matrix:

    - int 1-5 -> ``str(int)`` (e.g. ``"3"``).
    - int 1-5 with NA_Status != 0 -> ``str(int)`` (audit shows the
      published value even though the row was dropped).
    - ``'NA'`` (any NA_Status) -> ``"NA"`` (literal sentinel).
    - ``None`` cells -> ``None`` (defensive).

    Args:
        pts_cell: the raw ``PTS_X`` cell value (``int``, ``str 'NA'``,
            ``None``, or defensive for other types).

    Returns:
        The stringified cell text, or ``None`` for ``None`` cells.
    """
    if pts_cell is None:
        return None
    if isinstance(pts_cell, bool):
        return str(pts_cell)
    if isinstance(pts_cell, int):
        return str(pts_cell)
    if isinstance(pts_cell, str):
        return pts_cell  # literal "NA" or any other string
    return str(pts_cell)


# Read orchestrator.
def read_pts(
    xlsx_path: Path | None = None,
    *,
    year: int | None = None,
    catalog_path: Path | None = None,
) -> pd.DataFrame:
    """Open the xlsx and return the wide-format frame.

    Opens the single ``PTS-2025`` sheet, walks the 14-column
    structure, applies the §6 sentinel matrix, and returns the wide
    frame (one row per ``(COW_Code_A, Year)``, one column per catalog
    ``variable_name``).

    Args:
        xlsx_path: absolute path to the PTS xlsx (or ``None`` for the
            data-lake default).
        year: filter to a single year (e.g. ``2023``). Default: all
            years in the xlsx (1976-2024).
        catalog_path: override the indicator catalog. Default: checked-in.

    Returns:
        A wide-format DataFrame (one row per ``(cow_code, year)``, one
        column per catalog ``variable_name``). Indicator columns are
        ``Int64`` (nullable; ``pd.NA`` = missing per §6). The wide
        frame is **dense**: every row is present, even when all 3
        indicator cells are missing.

    Raises:
        FileNotFoundError: if the xlsx is missing.
    """
    if xlsx_path is None:
        xlsx_path = default_xlsx_path()
    if not xlsx_path.is_file():
        raise FileNotFoundError(f"PTS xlsx not found: {xlsx_path}")
    specs = load_indicator_catalog(catalog_path=catalog_path)
    long_df = read_xlsx_to_long_dataframe(xlsx_path, year=year)
    return read_pts_from_dataframe(long_df, specs, year=year)


def read_xlsx_to_long_dataframe(
    xlsx_path: Path,
    *,
    year: int | None = None,
) -> pd.DataFrame:
    """Open the PTS xlsx and return the long-format frame.

    Retains 10 of the 14 columns (the 4 secondary ID columns
    ``Country_OLD, COW_Code_N, WordBank_Code_A, UN_Code_N`` are
    dropped): ``Country, COW_Code_A, Year, Region, PTS_A/H/S,
    NA_Status_A/H/S``. The ``year=`` filter is applied here so the
    long frame stays small for year-filtered runs.

    Args:
        xlsx_path: absolute path to the PTS xlsx.
        year: filter to a single year. Default: all years.

    Returns:
        A pandas DataFrame with the 10 retained columns.

    Raises:
        FileNotFoundError: if the xlsx is missing.
        ValueError: if the sheet name is not ``PTS-2025`` or the
            header does not contain ``PTS_A`` / ``NA_Status_A``.
    """
    if not xlsx_path.is_file():
        raise FileNotFoundError(f"PTS xlsx not found: {xlsx_path}")
    expected_sheet = "PTS-2025"
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    try:
        if expected_sheet not in wb.sheetnames:
            raise ValueError(
                f"PTS xlsx {xlsx_path} has no sheet named {expected_sheet!r}. "
                f"Available sheets: {wb.sheetnames}. The xlsx's sheet name "
                "may have drifted from the expected value."
            )
        ws = wb[expected_sheet]

        rows: list[dict[str, object]] = []
        header: list[str] | None = None
        for row in ws.iter_rows(values_only=True):
            row_list = list(row)
            if header is None:
                header = [str(c) if c is not None else "" for c in row_list]
                continue
            # Defensive: trailing all-None rows (some xlsx builds add
            # a final blank line).
            if all(c is None for c in row_list):
                continue
            rec: dict[str, object] = {}
            for col_name, value in zip(header, row_list, strict=False):
                rec[col_name] = value
            if year is not None and rec.get("Year") != year:
                continue
            rows.append(rec)
    finally:
        wb.close()

    if header is None:
        raise ValueError(
            f"PTS xlsx {xlsx_path} is empty (no header row)."
        )
    if "PTS_A" not in header or "NA_Status_A" not in header:
        raise ValueError(
            f"PTS xlsx {xlsx_path} is missing expected indicator columns "
            f"(PTS_A / NA_Status_A). Header was: {header}"
        )

    return pd.DataFrame.from_records(rows, columns=header)


def read_pts_from_dataframe(
    df: pd.DataFrame,
    specs: list[IndicatorSpec],
    *,
    year: int | None = None,
) -> pd.DataFrame:
    """Apply the §6 sentinel matrix and delegate the pivot.

    Test seam: takes a pre-loaded DataFrame (matching the 14-column
    xlsx header shape) and returns the wide-format frame.

    Steps:

    1. Per row, per indicator pair ``(PTS_X, NA_Status_X)``: apply
       :func:`_coerce_pts_value` (the §6.5 defensive check is
       layered inside) and record the original cell text in the
       ``raw_lookup`` dict for the ``raw_value`` audit trail.
    2. §6.4 defensive region check: per row, if ``Region`` is not in
       :data:`pts_io._PTS_REGION_CODES` (7 World Bank codes) and
       not the ``'mena, ssa'`` anomaly, log a WARNING. The row is
       preserved in the wide frame per the §6.4 spec.
    3. Delegate the long-to-wide pivot to
       :func:`pts_xlsx_pivot.pivot_long_to_wide`.

    Args:
        df: long-format input frame with the 10 retained columns.
        specs: the catalog specs (``raw_column`` -> ``variable_name``
            mapping; one per PTS indicator).
        year: filter to a single year after the pivot. Default: keep
            all years.

    Returns:
        A wide-format DataFrame (one row per ``(cow_code, year)``,
        one column per catalog variable_name). Indicator columns
        are ``Int64``; the wide frame carries ``_pts_raw_lookup``,
        ``regions_covered``, and ``year_window`` in ``df.attrs``.
    """
    if df.empty:
        return pivot_long_to_wide([], {}, year=year)

    # Apply the §6 sentinel matrix per row + build the raw_lookup
    # audit dict in one pass. The §6.4 defensive region check is
    # layered here (per-row context is available).
    long_records: list[dict[str, object]] = []
    raw_lookup: dict[tuple[str, int, str], str] = {}
    for _, raw_row in df.iterrows():
        country = str(raw_row.get("Country") or "")
        cow_code = str(raw_row.get("COW_Code_A") or "")
        year_value = raw_row.get("Year")
        try:
            year_int = int(year_value) if year_value is not None else 0
        except (TypeError, ValueError):
            continue  # defensive: skip rows with non-int years
        region = str(raw_row.get("Region") or "")

        # §6.4 defensive region check: an unknown Region code (one not in
        # the 7 World Bank codes + the ``'mena, ssa'`` data anomaly)
        # is logged at WARNING but the row is preserved per the spec.
        if region and region not in PTS_REGION_CODES and region != "mena, ssa":
            _logger.warning(
                "PTS unknown region code: country=%s year=%d "
                "region=%s. Preserving but flagging.",
                country, year_int, region,
            )

        for raw_col, var_name in _PTS_INDICATOR_COLS:
            pts_cell = raw_row.get(raw_col)
            na_status_cell = raw_row.get(_NA_STATUS_COLS[raw_col])
            try:
                na_status_int = int(na_status_cell) if na_status_cell is not None else 0
            except (TypeError, ValueError):
                na_status_int = 0  # defensive default

            value = _coerce_pts_value(
                pts_cell,
                na_status_int,
                country=country,
                year=year_int,
                indicator=var_name,
            )
            # The raw_lookup key uses the canonical ``variable_name``
            # so the DB writer finds it by the same key it iterates over.
            raw_text = _raw_cell_text(pts_cell)
            if raw_text is not None:
                raw_lookup[(country, year_int, var_name)] = raw_text

            long_records.append(
                {
                    "country": country,
                    "cow_code": cow_code,
                    "year": year_int,
                    "region": region,
                    "variable_name": var_name,
                    "value": value,
                }
            )

    return pivot_long_to_wide(long_records, raw_lookup, year=year)
