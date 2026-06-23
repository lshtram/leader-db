"""Phase B Increment B -- PWT ``read_pwt`` reader boundary.

This file covers the Stage 2 ``read_pwt`` reader module's
boundary contract. The reader returns the wide
Data-sheet-shaped DataFrame (one row per ``(countrycode,
year)``, one column per raw indicator). The transform module
is what pivots the wide frame into the canonical long
format; reader tests assert raw shape only.

PASS-ELIGIBLE / DOMAIN-RED conventions
--------------------------------------

- ``PASS-ELIGIBLE`` -- the test passes against the Phase B
  stub (the stub is a wrong-shaped 0-row wide DataFrame that
  fails every content assertion -- the tests are NOT pass-
  eligible; the marker is unused in this file).
- ``DOMAIN-RED`` -- the test is intentionally RED at the
  domain layer until the production reader lands. Failure
  mode is an assertion failure on the wrong-shaped stub
  output (zero rows, missing columns, or ``DID NOT RAISE``)
  -- NOT ``ModuleNotFoundError``.

Coverage
--------

- The reader returns the wide Data-sheet-shaped DataFrame
  with the documented identity + catalog columns.
- The reader opens the ``Data`` sheet ONLY (NOT ``Info`` /
  ``Legend``).
- A workbook missing the identity columns raises
  ``ValueError`` mentioning the missing column.
- A workbook missing one of the catalog indicator columns
  raises ``ValueError`` mentioning the missing column.
- The reader refuses a workbook named ``pwt100.xlsx`` (the
  legacy pre-Phase-B name).
"""

from __future__ import annotations

from pathlib import Path

import pandas as pd
import pytest

from .conftest import (
    PWT_CATALOG_RAW_COLUMNS,
    PWT_TEST_IDENTITY_COLUMNS,
    PWT_XLSX_NAME,
)

# ---------------------------------------------------------------------------
# 1. Reader
# ---------------------------------------------------------------------------


def test_pwt_reader_returns_wide_data_sheet_dataframe(
    pwt_xlsx_dir: Path,
) -> None:
    """The reader returns the wide Data-sheet-shaped DataFrame.

    DOMAIN-RED: ``read_pwt()`` returns a wrong-shaped wide
    DataFrame (right columns, zero rows) in the Phase B stub.
    The test fails at the ``len(df) == 6`` assertion -- the
    production reader must populate the frame from the staged
    xlsx so the row count passes.
    """
    from leaders_db.ingest.sources.pwt.reader import read_pwt

    df = read_pwt(xlsx_path=pwt_xlsx_dir / PWT_XLSX_NAME)
    assert isinstance(df, pd.DataFrame)

    identity_cols = set(PWT_TEST_IDENTITY_COLUMNS)
    assert identity_cols.issubset(set(df.columns)), (
        f"reader frame missing required identity columns: "
        f"identity_cols - present = "
        f"{sorted(identity_cols - set(df.columns))}"
    )

    catalog_cols = set(PWT_CATALOG_RAW_COLUMNS)
    assert catalog_cols.issubset(set(df.columns)), (
        f"reader frame missing required catalog columns: "
        f"catalog_cols - present = "
        f"{sorted(catalog_cols - set(df.columns))}"
    )

    assert len(df) == 6, f"expected 6 wide rows, got {len(df)}"
    assert set(df["countrycode"].unique()) == {"USA", "MEX", "SWE"}, (
        f"reader frame has unexpected ISO3s: "
        f"{sorted(set(df['countrycode'].unique()))}"
    )
    assert pd.api.types.is_integer_dtype(df["year"]), (
        f"reader frame year column dtype is "
        f"{df['year'].dtype}, expected integer"
    )


def test_pwt_reader_does_not_open_info_or_legend_sheets(
    pwt_xlsx_dir: Path,
) -> None:
    """The reader opens the ``Data`` sheet ONLY.

    DOMAIN-RED: ``read_pwt()`` returns a wrong-shaped wide
    DataFrame (right columns, zero rows) in the Phase B stub.
    The test fails at the ISO3 set assertion (empty set) -- the
    production reader must populate the frame from the staged
    xlsx so the distinct ISO3 set is exactly the 3 fixture
    countries.
    """
    import openpyxl

    from leaders_db.ingest.sources.pwt.reader import read_pwt

    multi_sheet = pwt_xlsx_dir / "_multi.xlsx"
    wb = openpyxl.Workbook()
    default = wb.active
    if default is not None:
        wb.remove(default)

    info_ws = wb.create_sheet(title="Info")
    info_ws.append(["version", "release_date", "note"])
    info_ws.append(["BOGUS_VERSION", "1970-01-01", "wrong sheet!"])

    legend_ws = wb.create_sheet(title="Legend")
    legend_ws.append(["variable_name", "description", "wrong_sheet"])
    legend_ws.append(["BOGUS_VAR", "wrong sheet row", "yes"])

    data_ws = wb.create_sheet(title="Data")
    data_ws.append(
        list(PWT_TEST_IDENTITY_COLUMNS) + list(PWT_CATALOG_RAW_COLUMNS),
    )
    # All 3 fixture countries on the Data sheet (the test asserts
    # the reader returns exactly these ISO3s and not the Info /
    # Legend contamination).
    data_ws.append(
        ["USA", "United States", "US Dollar", 2019, 19.5e12, 19.5e12,
         327.2e6, 158.5e6, 1785.0, 3.65, None, None, None, None, None],
    )
    data_ws.append(
        ["MEX", "Mexico", "Mexican Peso", 2019, 2.2e12, 2.1e12,
         127.6e6, 55.7e6, 2200.0, 2.55, None, None, None, None, None],
    )
    data_ws.append(
        ["SWE", "Sweden", "Swedish Krona", 2019, 5.0e11, 4.8e11,
         10.3e6, None, None, None, None, None, None, None, None],
    )

    wb.save(multi_sheet)

    df = read_pwt(xlsx_path=multi_sheet)

    assert "version" not in df.columns, (
        f"reader pulled the Info sheet column 'version': "
        f"present columns = {sorted(df.columns)}"
    )
    assert "wrong_sheet" not in df.columns, (
        f"reader pulled the Legend sheet column 'wrong_sheet': "
        f"present columns = {sorted(df.columns)}"
    )
    non_null_iso3 = df["countrycode"].dropna()
    assert set(non_null_iso3.unique()) == {"USA", "MEX", "SWE"}, (
        f"reader frame has wrong-sheet ISO3s: "
        f"{sorted(set(non_null_iso3.unique()))}"
    )


def test_pwt_reader_requires_identity_columns(pwt_xlsx_dir: Path) -> None:
    """A workbook missing the identity columns raises a clear
    error.

    DOMAIN-RED: ``read_pwt()`` does NOT raise ``ValueError`` in
    the Phase B stub (it returns a wrong-shaped wide DataFrame
    regardless of the input). The test fails at the
    ``pytest.raises(ValueError)`` block because no exception
    was raised -- the production reader must validate the
    identity columns and raise.
    """
    import openpyxl

    from leaders_db.ingest.sources.pwt.reader import read_pwt

    broken_xlsx = pwt_xlsx_dir / "_broken.xlsx"
    wb = openpyxl.Workbook()
    default = wb.active
    if default is not None:
        wb.remove(default)
    ws = wb.create_sheet(title="Data")
    ws.append(["year", "pop"])
    ws.append([2019, 100.0])
    wb.save(broken_xlsx)

    with pytest.raises(ValueError) as exc_info:
        read_pwt(xlsx_path=broken_xlsx)
    msg = str(exc_info.value)
    assert "countrycode" in msg or "identity" in msg.lower(), (
        f"error must name the missing identity column; got {msg!r}"
    )


def test_pwt_reader_requires_all_catalog_columns(
    pwt_xlsx_dir: Path,
) -> None:
    """A workbook missing one of the catalog indicator columns
    raises ``ValueError`` mentioning the missing column.

    DOMAIN-RED: ``read_pwt()`` does NOT raise ``ValueError`` in
    the Phase B stub. The test fails at the
    ``pytest.raises(ValueError)`` block -- the production
    reader must validate the catalog columns and raise.
    """
    import openpyxl

    from leaders_db.ingest.sources.pwt.reader import read_pwt

    broken_xlsx = pwt_xlsx_dir / "_missing_catalog_col.xlsx"
    wb = openpyxl.Workbook()
    default = wb.active
    if default is not None:
        wb.remove(default)
    ws = wb.create_sheet(title="Data")
    ws.append(
        [
            "countrycode", "country", "currency_unit", "year",
            # rgdpe deliberately dropped
            "rgdpo", "pop", "emp", "avh", "hc",
            "ccon", "cda", "ctfp", "rkna", "rtfpna",
        ],
    )
    ws.append(
        ["USA", "United States", "US Dollar", 2019, 19.5e12,
         327.2e6, 158.5e6, 1785.0, 3.65, None, None, None, None, None],
    )
    wb.save(broken_xlsx)

    with pytest.raises(ValueError) as exc_info:
        read_pwt(xlsx_path=broken_xlsx)
    msg = str(exc_info.value)
    assert "rgdpe" in msg, (
        f"error must name the missing catalog column; got {msg!r}"
    )


def test_pwt_reader_refuses_legacy_pwt100_xlsx(
    pwt_xlsx_dir: Path,
) -> None:
    """The reader refuses a workbook named ``pwt100.xlsx``.

    DOMAIN-RED: ``read_pwt()`` does NOT raise
    ``ValueError`` / ``FileNotFoundError`` in the Phase B stub
    (it returns a wrong-shaped wide DataFrame). The test fails
    at the ``pytest.raises((ValueError, FileNotFoundError))``
    block -- the production reader must enforce the filename
    gate and raise.
    """
    from leaders_db.ingest.sources.pwt.reader import read_pwt

    legacy = pwt_xlsx_dir / "pwt100.xlsx"
    legacy.touch()
    with pytest.raises((ValueError, FileNotFoundError)) as exc_info:
        read_pwt(xlsx_path=legacy)
    msg = str(exc_info.value)
    assert "pwt1001.xlsx" in msg, (
        f"error must name the expected filename pwt1001.xlsx; "
        f"got {msg!r}"
    )


__all__ = []
