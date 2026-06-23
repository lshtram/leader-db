"""Phase B Increment B -- PWT ``SourceAdapter`` (readiness, dispatch, request scoping).

This file covers the per-source PWT package's
:class:`PWTAdapter` and the public-orchestrator surface
(``ingest_pwt``). Tests are split per the
``docs/sources/ingestion-plan.md`` mirrored layout:

- ``test_adapter.py`` -- this file: PWTAdapter readiness gate,
  metadata validation, request-scoped ``raw_root``, and the
  ``STAGE2_ADAPTERS`` dispatch contract.
- ``test_reader.py`` -- the ``read_pwt`` boundary.
- ``test_transform.py`` -- the ``transform_pwt_long_frame``
  boundary (long row schema, locator, temporal_kind,
  attribution, missing-cell emission, catalog-driven
  emission).
- ``test_db_cli.py`` -- the registry E2E, CLI E2E, year
  behavior, and idempotency boundary.

PASS-ELIGIBLE / DOMAIN-RED conventions
--------------------------------------

Each test in this file is tagged with one of:

- ``PASS-ELIGIBLE`` -- the test passes against the Phase B
  stub and must keep passing once the production code lands.
- ``DOMAIN-RED`` -- the test is intentionally RED at the
  domain layer until the production PWT adapter lands.
  Failure mode is an assertion failure on the wrong-shaped
  stub output (e.g. the readiness gate's generic blocker
  does NOT name the missing metadata field). NOT
  ``ModuleNotFoundError``.

Coverage
--------

- ``STAGE2_ADAPTERS['pwt']`` is callable (or ``None`` while
  pending Increment B) but never silently dropped.
- The per-source package exposes a public ``ingest_pwt``
  orchestrator and the documented constants
  (``PWT_SOURCE_KEY``, ``PWT_XLSX_NAME``, ``PWT_ATTRIBUTION``).
- The readiness gate fires BEFORE the reader opens the
  workbook; the gate names the missing metadata field in
  the blocker.
- The readiness gate honors a request-scoped ``raw_root``
  override (the protocol change in Increment A).
- The PWT ``metadata.json`` carries every required field
  (source_url, license_note, checksum_sha256, local_files,
  ingestion_status, coverage).
- Removing any required field flips readiness to
  ``ready=False``; the production gate must name the missing
  field in the blocker.
- The PWT attribution text is a substring of
  ``docs/sources/attributions.md`` (Rule #15 drift guard).
"""

from __future__ import annotations

import hashlib
import json
import shutil
from pathlib import Path

import pytest
from sqlalchemy import select

from leaders_db.db.models import Source
from leaders_db.ingest import STAGE2_ADAPTERS
from leaders_db.ingest.interfaces import IngestRequest
from leaders_db.ingest.sources.pwt import PWTAdapter
from leaders_db.ingest.sources.pwt.db_helpers import (
    register_pwt_source,
)

from .conftest import PWT_METADATA_NAME, PWT_SOURCE_KEY, PWT_XLSX_NAME

# ---------------------------------------------------------------------------
# 1. Dispatch / STAGE2_ADAPTERS gate
# ---------------------------------------------------------------------------


def test_stage2_adapters_pwt_is_callable_not_none() -> None:
    """``STAGE2_ADAPTERS['pwt']`` is now a real callable (no
    longer ``None``).

    Contract: the Increment B implementation registers a PWT
    adapter in the legacy dispatch table; the CLI can find it.

    DOMAIN-RED: ``STAGE2_ADAPTERS['pwt']`` is still ``None`` at
    Phase B (the per-source package exists but the legacy
    dispatch entry is intentionally not flipped until tests
    prove the metadata gate, reader, transform, DB write,
    manifest, attribution, and CLI boundary).
    """
    assert "pwt" in STAGE2_ADAPTERS
    adapter = STAGE2_ADAPTERS["pwt"]
    assert adapter is not None, (
        "STAGE2_ADAPTERS['pwt'] is still None; the Increment B "
        "adapter must be wired before tests can pass."
    )
    assert callable(adapter)


def test_pwt_source_package_public_orchestrator_exists() -> None:
    """The new per-source PWT package exposes a public
    ``ingest_pwt`` orchestrator.

    PASS-ELIGIBLE: the Phase B stub provides a callable
    ``ingest_pwt`` symbol; the test passes against the stub.
    """
    from leaders_db.ingest.sources.pwt import ingest_pwt

    assert callable(ingest_pwt)


def test_pwt_source_package_public_constants() -> None:
    """The new per-source PWT package exposes the spec'd
    constants: ``PWT_SOURCE_KEY``, ``PWT_XLSX_NAME``,
    ``PWT_ATTRIBUTION``.

    PASS-ELIGIBLE: the stub exports the constants with the
    canonical attribution text (Rule #15 drift guard).
    """
    from leaders_db.ingest.sources.pwt import (  # type: ignore[attr-defined]
        PWT_ATTRIBUTION,
        PWT_SOURCE_KEY,
        PWT_XLSX_NAME,
    )

    assert PWT_SOURCE_KEY == "pwt"
    assert PWT_XLSX_NAME == "pwt1001.xlsx"
    assert isinstance(PWT_ATTRIBUTION, str)
    assert "Feenstra" in PWT_ATTRIBUTION
    assert "Inklaar" in PWT_ATTRIBUTION
    assert "Timmer" in PWT_ATTRIBUTION


# ---------------------------------------------------------------------------
# 2. Readiness gate: missing metadata blocks before reader access
# ---------------------------------------------------------------------------


def test_pwt_readiness_blocks_when_metadata_missing(
    pwt_xlsx_no_metadata: Path,
) -> None:
    """``check_ready()`` with no ``metadata.json`` returns
    ``ready=False`` with a blocker mentioning the missing file.

    Contract: the readiness gate fires BEFORE the reader opens
    ``pwt1001.xlsx``. The blocker names the missing artifact so a
    developer can fix it without reading source code.

    DOMAIN-RED: the Phase B stub ``PWTAdapter().check_ready()``
    returns a generic "not implemented" blocker that does NOT
    mention ``metadata``. The test fails at the assertion
    ``"metadata" in readiness.blocker.lower()`` -- the production
    gate must mention the missing field by name.
    """
    from leaders_db.ingest.interfaces import SourceReadiness
    from leaders_db.ingest.sources.pwt import PWTAdapter, ingest_pwt

    adapter = PWTAdapter()
    readiness = adapter.check_ready(
        IngestRequest(
            source_key="pwt", raw_root=pwt_xlsx_no_metadata.parent,
        ),
    )
    assert isinstance(readiness, SourceReadiness)
    assert readiness.ready is False, (
        "PWT adapter should refuse to run when metadata.json is "
        "absent; got ready=True"
    )
    assert readiness.blocker is not None
    assert "metadata" in readiness.blocker.lower(), (
        f"blocker must mention metadata so a developer can act on "
        f"it; got {readiness.blocker!r}"
    )

    # Sanity: the function is also exposed via the public
    # orchestrator (the registry / CLI uses it).
    assert callable(ingest_pwt)


def test_pwt_registry_blocks_before_openpyxl_load_workbook(
    pwt_xlsx_no_metadata: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    """``registry.ingest_source`` for PWT with missing metadata
    raises an actionable error AND never invokes
    ``openpyxl.load_workbook``.

    Strengthened proof (Phase B review feedback): the test
    monkeypatches ``openpyxl.load_workbook`` with a spy that
    records every call. The readiness gate must short-circuit
    BEFORE the reader opens the workbook -- any recorded call
    is a regression.

    PASS-ELIGIBLE: the Phase B stub readiness gate
    short-circuits before the reader opens the workbook; the
    spy on ``openpyxl.load_workbook`` records zero calls.
    Regression guard for the registry's ``check_ready -> read``
    ordering.
    """
    import openpyxl

    from leaders_db.ingest.registry import (
        ingest_source,
        register,
        unregister,
    )
    from leaders_db.ingest.sources.pwt import PWTAdapter

    register("pwt", PWTAdapter())
    try:
        # Spy on ``openpyxl.load_workbook``. Any call means the
        # runner bypassed the readiness gate and reached the
        # reader.
        load_workbook_calls: list[tuple] = []
        real_load_workbook = openpyxl.load_workbook

        def spy_load_workbook(*args, **kwargs):
            load_workbook_calls.append((args, kwargs))
            return real_load_workbook(*args, **kwargs)

        monkeypatch.setattr(openpyxl, "load_workbook", spy_load_workbook)

        with pytest.raises(RuntimeError) as exc_info:
            ingest_source(
                IngestRequest(source_key="pwt", year=2019),
            )

        # The runner NEVER opened the workbook.
        assert load_workbook_calls == [], (
            f"openpyxl.load_workbook was invoked before readiness "
            f"succeeded: {load_workbook_calls!r}"
        )
        # The error names the source so a developer can act on
        # it without reading source code.
        msg = str(exc_info.value)
        assert "pwt" in msg.lower(), (
            f"error must name the source / blocker; got {msg!r}"
        )
    finally:
        unregister("pwt")


# ---------------------------------------------------------------------------
# 3. Metadata validation
# ---------------------------------------------------------------------------


def test_pwt_metadata_required_fields_present(pwt_xlsx_dir: Path) -> None:
    """The PWT ``metadata.json`` carries every required field
    (source_url, license_note, checksum_sha256, local_files,
    ingestion_status, coverage).

    PASS-ELIGIBLE: fixture-validation test. The test fixture
    writes a well-formed ``metadata.json`` (the SHA-256 matches
    the staged xlsx bytes) and the test verifies the fixture
    is valid against the readiness contract.
    """
    import hashlib

    payload = json.loads(
        (pwt_xlsx_dir / PWT_METADATA_NAME).read_text(encoding="utf-8"),
    )
    required = {
        "source_url",
        "license_note",
        "checksum_sha256",
        "local_files",
        "ingestion_status",
        "coverage",
    }
    missing = required - set(payload.keys())
    assert not missing, (
        f"metadata.json missing required fields: {sorted(missing)}"
    )
    assert "pwt1001.xlsx" in payload["local_files"], (
        f"local_files must include pwt1001.xlsx; got "
        f"{payload['local_files']!r}"
    )
    assert payload["ingestion_status"] == "downloaded"
    # coverage must be a non-empty string so downstream
    # consumers can detect gaps (e.g. 1950-2019 vs a 2023
    # request).
    assert isinstance(payload["coverage"], str) and payload["coverage"], (
        f"coverage must be a non-empty string; got "
        f"{payload['coverage']!r}"
    )
    expected_sha = hashlib.sha256(
        (pwt_xlsx_dir / "pwt1001.xlsx").read_bytes(),
    ).hexdigest()
    assert payload["checksum_sha256"] == expected_sha, (
        "checksum_sha256 must match the staged xlsx"
    )


def test_pwt_well_formed_metadata_passes_readiness(
    pwt_xlsx_dir: Path,
) -> None:
    """``check_ready()`` accepts a fully-formed ``metadata.json``
    (returns ``ready=True`` with no blocker).

    PASS-ELIGIBLE: the Phase B readiness stub already accepts the
    well-formed fixture so later production work cannot regress the
    positive readiness path while fixing the negative metadata cases.
    """
    from leaders_db.ingest.sources.pwt import PWTAdapter

    adapter = PWTAdapter()
    readiness = adapter.check_ready(
        IngestRequest(source_key="pwt", raw_root=pwt_xlsx_dir.parent),
    )
    assert readiness.ready is True, (
        f"check_ready() must return ready=True for a well-formed "
        f"bundle; got blocker={readiness.blocker!r}"
    )
    assert readiness.blocker is None


def test_pwt_metadata_missing_source_url_blocks(
    pwt_xlsx_dir: Path,
) -> None:
    """Removing ``source_url`` from ``metadata.json`` flips
    ``check_ready()`` to ``ready=False``.

    DOMAIN-RED: the Phase B stub returns ``ready=False`` with
    a generic "not implemented" blocker. The test fails at
    the assertion that the blocker mentions ``source_url`` --
    the production gate must name the missing field.
    """
    from leaders_db.ingest.sources.pwt import PWTAdapter

    bad_path = pwt_xlsx_dir / PWT_METADATA_NAME
    payload = json.loads(bad_path.read_text(encoding="utf-8"))
    del payload["source_url"]
    bad_path.write_text(json.dumps(payload), encoding="utf-8")

    adapter = PWTAdapter()
    readiness = adapter.check_ready(
        IngestRequest(source_key="pwt", raw_root=pwt_xlsx_dir.parent),
    )
    assert readiness.ready is False
    assert readiness.blocker is not None
    assert "source_url" in readiness.blocker.lower(), (
        f"blocker must mention source_url; got {readiness.blocker!r}"
    )


def test_pwt_metadata_license_note_missing_blocks(
    pwt_xlsx_dir: Path,
) -> None:
    """Removing ``license_note`` blocks the run (Rule #15
    attribution chain starts at metadata).

    DOMAIN-RED: stub returns generic blocker; test fails at
    the assertion that the blocker mentions ``license``.
    """
    from leaders_db.ingest.sources.pwt import PWTAdapter

    bad_path = pwt_xlsx_dir / PWT_METADATA_NAME
    payload = json.loads(bad_path.read_text(encoding="utf-8"))
    del payload["license_note"]
    bad_path.write_text(json.dumps(payload), encoding="utf-8")

    adapter = PWTAdapter()
    readiness = adapter.check_ready(
        IngestRequest(source_key="pwt", raw_root=pwt_xlsx_dir.parent),
    )
    assert readiness.ready is False
    assert "license" in (readiness.blocker or "").lower()


def test_pwt_metadata_checksum_mismatch_blocks(
    pwt_xlsx_dir: Path,
) -> None:
    """A wrong ``checksum_sha256`` blocks the run.

    Contract: the readiness gate recomputes the file's SHA-256
    and refuses to run if it does not match the metadata field.

    DOMAIN-RED: the Phase B readiness stub validates required field
    presence but does not recompute the checksum, so this fails at
    ``readiness.ready is False``. Production must block and mention
    ``checksum``.
    """
    from leaders_db.ingest.sources.pwt import PWTAdapter

    bad_path = pwt_xlsx_dir / PWT_METADATA_NAME
    payload = json.loads(bad_path.read_text(encoding="utf-8"))
    payload["checksum_sha256"] = "0" * 64
    bad_path.write_text(json.dumps(payload), encoding="utf-8")

    adapter = PWTAdapter()
    readiness = adapter.check_ready(
        IngestRequest(source_key="pwt", raw_root=pwt_xlsx_dir.parent),
    )
    assert readiness.ready is False
    assert "checksum" in (readiness.blocker or "").lower()


def test_pwt_metadata_local_files_must_contain_pwt1001_xlsx(
    pwt_xlsx_dir: Path,
) -> None:
    """``local_files`` must include ``pwt1001.xlsx``; a metadata
    listing only ``pwt100.xlsx`` (the legacy pre-Phase-B name)
    blocks the run.

    DOMAIN-RED: stub returns generic blocker; test fails at
    the assertion that the blocker mentions ``pwt1001.xlsx``
    or ``local_files``.
    """
    from leaders_db.ingest.sources.pwt import PWTAdapter

    bad_path = pwt_xlsx_dir / PWT_METADATA_NAME
    payload = json.loads(bad_path.read_text(encoding="utf-8"))
    payload["local_files"] = ["pwt100.xlsx"]
    bad_path.write_text(json.dumps(payload), encoding="utf-8")

    adapter = PWTAdapter()
    readiness = adapter.check_ready(
        IngestRequest(source_key="pwt", raw_root=pwt_xlsx_dir.parent),
    )
    assert readiness.ready is False
    msg = (readiness.blocker or "").lower()
    assert "pwt1001.xlsx" in msg or "local_files" in msg, (
        f"blocker must mention pwt1001.xlsx; got {readiness.blocker!r}"
    )


def test_pwt_metadata_ingestion_status_not_downloaded_blocks(
    pwt_xlsx_dir: Path,
) -> None:
    """``ingestion_status`` must be ``"downloaded"``.

    DOMAIN-RED: stub returns generic blocker; test fails at
    the assertion that the blocker mentions ``ingestion_status``.
    """
    from leaders_db.ingest.sources.pwt import PWTAdapter

    bad_path = pwt_xlsx_dir / PWT_METADATA_NAME
    payload = json.loads(bad_path.read_text(encoding="utf-8"))
    payload["ingestion_status"] = "pending"
    bad_path.write_text(json.dumps(payload), encoding="utf-8")

    adapter = PWTAdapter()
    readiness = adapter.check_ready(
        IngestRequest(source_key="pwt", raw_root=pwt_xlsx_dir.parent),
    )
    assert readiness.ready is False
    assert "ingestion_status" in (readiness.blocker or "").lower()


def test_pwt_metadata_coverage_missing_blocks(
    pwt_xlsx_dir: Path,
) -> None:
    """``coverage`` is a required metadata field (per the
    Increment B readiness contract). Removing it blocks the
    run; the blocker mentions ``coverage`` so a developer can
    fix the metadata without reading source code.

    Contract: a well-formed PWT bundle's ``metadata.json``
    declares its temporal + spatial coverage (e.g.
    ``"country-year economic accounts"``) so downstream
    consumers can detect gaps (year=2023 against a
    1950-2019 coverage) and surface
    ``requested_year_out_of_coverage`` warnings.

    DOMAIN-RED: the Phase B stub ``PWTAdapter().check_ready()``
    returns a generic blocker that does NOT mention
    ``coverage`` once the field is removed (the field IS in
    the required set, so the gate fires, but the blocker is
    still generic). The test fails at the assertion
    ``"coverage" in readiness.blocker.lower()`` -- the
    production gate must mention the missing field by name.
    """
    from leaders_db.ingest.sources.pwt import PWTAdapter

    bad_path = pwt_xlsx_dir / PWT_METADATA_NAME
    payload = json.loads(bad_path.read_text(encoding="utf-8"))
    del payload["coverage"]
    bad_path.write_text(json.dumps(payload), encoding="utf-8")

    adapter = PWTAdapter()
    readiness = adapter.check_ready(
        IngestRequest(source_key="pwt", raw_root=pwt_xlsx_dir.parent),
    )
    assert readiness.ready is False, (
        "PWT adapter should refuse to run when coverage is "
        "missing from metadata.json; got ready=True"
    )
    assert readiness.blocker is not None, (
        "blocker must be a non-None string when ready=False; "
        "got None"
    )
    assert "coverage" in readiness.blocker.lower(), (
        f"blocker must mention the missing 'coverage' field; "
        f"got {readiness.blocker!r}"
    )


def test_pwt_request_scoped_raw_root_overrides_default(
    pwt_xlsx_dir: Path,
    tmp_path: Path,
) -> None:
    """A request-scoped ``raw_root`` override must win over the
    default data-lake path.

    Scenario: the well-formed PWT bundle lives ONLY under a
    request-scoped ``raw_root`` (``<tmp>/custom_raw/pwt/``).
    The adapter must honor ``request.raw_root`` and return
    ``ready=True`` for the request-scoped bundle. A second
    call with a request-scoped raw_root that is intentionally
    EMPTY (no PWT bundle under it) must return ``ready=False``
    -- proving the request-scoped path is the one being
    resolved, not the default data-lake.

    PASS-ELIGIBLE: the Phase B readiness stub already honors
    ``request.raw_root``. This guards the request-scoped data-lake
    behavior while production fills in reader/transform/write.
    """
    import shutil

    from leaders_db.ingest.sources.pwt import PWTAdapter

    custom_raw = tmp_path / "custom_raw"
    custom_pwt = custom_raw / "pwt"
    shutil.copytree(pwt_xlsx_dir, custom_pwt)

    adapter = PWTAdapter()

    # The request-scoped raw_root points to the well-formed
    # bundle; readiness must return ``ready=True``.
    custom_request = IngestRequest(
        source_key="pwt", raw_root=str(custom_raw),
    )
    custom_readiness = adapter.check_ready(custom_request)
    assert custom_readiness.ready is True, (
        f"request-scoped raw_root={custom_raw!r} points to a "
        f"well-formed PWT bundle; check_ready must return "
        f"ready=True; got blocker={custom_readiness.blocker!r}"
    )

    # The request-scoped raw_root points to a NON-EXISTENT
    # bundle. If the adapter resolves to the default
    # data-lake (which IS well-formed), this assertion will
    # fail -- proving the request-scoped path is honored.
    empty_request = IngestRequest(
        source_key="pwt", raw_root=str(tmp_path / "empty_raw"),
    )
    empty_readiness = adapter.check_ready(empty_request)
    assert empty_readiness.ready is False, (
        f"request-scoped raw_root={empty_request.raw_root!r} "
        f"points to a non-existent PWT bundle; check_ready "
        f"must return ready=False (proving the request-scoped "
        f"path is honored, not the default data-lake); got "
        f"readiness={empty_readiness!r}"
    )


# ---------------------------------------------------------------------------
# 4. Public-surface hygiene
# ---------------------------------------------------------------------------


def test_pwt_register_source_uses_request_scoped_bundle_metadata(
    isolated_data_lake: Path,
    tmp_path: Path,
) -> None:
    """Phase B Increment B second-pass regression proof:
    ``register_pwt_source(session, request=...)`` reads the bundle
    ``metadata.json`` from ``request.raw_root`` -- NOT from the
    default data-lake path. A custom raw-root run whose
    ``metadata.json`` carries a distinct ``source_url`` and
    ``license_note`` MUST be reflected in the DB ``sources``
    row.

    The previous Increment B pass always read
    ``data/raw/pwt/metadata.json`` and silently ignored
    ``request.raw_root``, so a custom raw-root run wrote DB
    provenance from the default data-lake bundle (or from the
    fallback literal URL / license text). The fix forwards
    ``request.raw_root`` into the metadata read path.

    DOMAIN-RED until the request-scoped metadata read lands.
    """
    from leaders_db.db.engine import init_database
    from leaders_db.db.session import session_scope

    custom_raw = tmp_path / "raw-bundle-metadata"
    fixtures_dir = Path(__file__).resolve().parents[3] / "fixtures" / "pwt"
    target = custom_raw / PWT_SOURCE_KEY
    target.mkdir(parents=True, exist_ok=True)
    # Copy the fixture xlsx + recompute the SHA-256.
    shutil.copy2(fixtures_dir / "sample.xlsx", target / PWT_XLSX_NAME)
    sha = hashlib.sha256(
        (target / PWT_XLSX_NAME).read_bytes(),
    ).hexdigest()
    # Distinct ``source_url`` and ``license_note`` so we can
    # prove the DB row reflects the request-scoped bundle and
    # NOT the default data-lake metadata (which would carry
    # the canonical Feenstra URL + CC BY 4.0 license).
    custom_source_url = (
        "https://example.test/custom-pwt-bundle/pwt1001.xlsx"
    )
    custom_license_note = (
        "Custom test license note for Phase B Increment B "
        "second-pass regression proof (NOT the canonical CC BY "
        "4.0 license; the DB row MUST reflect this value when "
        "the request-scoped raw_root is honored)."
    )
    metadata_payload = {
        "source_name": "Penn World Table",
        "source_version": "10.01",
        "download_date": "2026-06-22",
        "coverage": "country-year economic accounts",
        "years_available": "1950-2019",
        "license_note": custom_license_note,
        "local_files": [PWT_XLSX_NAME],
        "ingestion_status": "downloaded",
        "source_url": custom_source_url,
        "checksum_sha256": sha,
    }
    (target / PWT_METADATA_NAME).write_text(
        json.dumps(metadata_payload, indent=2), encoding="utf-8",
    )

    db_path = tmp_path / "db-bundle-metadata" / "leaders_db.sqlite"
    db_path.parent.mkdir(parents=True, exist_ok=True)
    db_url = f"sqlite:///{db_path.as_posix()}"
    init_database(db_url)

    request = IngestRequest(
        source_key="pwt",
        year=2019,
        raw_root=custom_raw,
        database_url=db_url,
    )

    # Sanity: with the request-scoped raw_root, the readiness
    # gate must accept the custom bundle (Phase B Increment B
    # reviewer contract: the same request flows through
    # check_ready -> read -> write).
    readiness = PWTAdapter().check_ready(request)
    assert readiness.ready is True, (
        f"check_ready(request) must accept the request-scoped "
        f"bundle; got blocker={readiness.blocker!r}"
    )

    # Register the source via the helper with the request so
    # the metadata read uses the request-scoped path. The DB
    # row MUST carry the custom source_url + license_note.
    with session_scope(db_url) as session:
        source_id = register_pwt_source(session, request=request)
    assert source_id > 0

    with session_scope(db_url) as session:
        pwt_sources = session.execute(
            select(Source).where(
                Source.source_name == "Penn World Table",
                Source.version == "10.01",
            ),
        ).scalars().all()
    assert len(pwt_sources) == 1
    pwt_source = pwt_sources[0]
    assert pwt_source.source_url == custom_source_url, (
        f"DB sources.source_url must reflect the request-scoped "
        f"bundle metadata ({custom_source_url!r}); got "
        f"{pwt_source.source_url!r}. The Phase B Increment B "
        "second-pass regression: register_pwt_source fell back "
        "to the default data-lake metadata when "
        "request.raw_root was set."
    )
    assert pwt_source.license_note == custom_license_note, (
        f"DB sources.license_note must reflect the request-scoped "
        f"bundle metadata ({custom_license_note!r}); got "
        f"{pwt_source.license_note!r}. The Phase B Increment B "
        "second-pass regression: register_pwt_source fell back "
        "to the default data-lake metadata when "
        "request.raw_root was set."
    )

    # And the full ingest path (registry-style convenience
    # path through ``PWTAdapter.ingest(request)``) MUST also
    # write the custom metadata to the DB sources row.
    request_db_path = tmp_path / "db-bundle-metadata-full" / "leaders_db.sqlite"
    request_db_path.parent.mkdir(parents=True, exist_ok=True)
    request_db_url = f"sqlite:///{request_db_path.as_posix()}"
    init_database(request_db_url)

    PWTAdapter().ingest(
        IngestRequest(
            source_key="pwt",
            year=2019,
            raw_root=custom_raw,
            database_url=request_db_url,
        ),
    )

    with session_scope(request_db_url) as session:
        full_pwt_sources = session.execute(
            select(Source).where(
                Source.source_name == "Penn World Table",
                Source.version == "10.01",
            ),
        ).scalars().all()
    assert len(full_pwt_sources) == 1
    full_source = full_pwt_sources[0]
    assert full_source.source_url == custom_source_url, (
        f"PWTAdapter.ingest(request) must upsert the sources "
        f"row from the request-scoped bundle metadata; got "
        f"source_url={full_source.source_url!r}, expected "
        f"{custom_source_url!r}"
    )
    assert full_source.license_note == custom_license_note, (
        f"PWTAdapter.ingest(request) must upsert the sources "
        f"row from the request-scoped bundle metadata; got "
        f"license_note={full_source.license_note!r}"
    )





def test_pwt_attribution_matches_attributions_doc() -> None:
    """The PWT attribution text must be a substring of
    ``docs/sources/attributions.md`` (Rule #15 drift guard).

    PASS-ELIGIBLE: the stub's ``PWT_ATTRIBUTION`` constant
    matches the canonical citation block in
    ``docs/sources/attributions.md``. Drift guard.
    """
    from leaders_db.ingest.sources.pwt import PWT_ATTRIBUTION

    doc_path = (
        Path(__file__).resolve().parents[4]
        / "docs"
        / "sources/attributions.md"
    )
    doc_path = doc_path.resolve()
    doc_text = doc_path.read_text(encoding="utf-8")
    assert PWT_ATTRIBUTION in doc_text, (
        f"PWT_ATTRIBUTION is not a substring of {doc_path}. "
        "Update both in the same commit (Rule #15)."
    )


__all__ = [
    "PWT_SOURCE_KEY",
]
