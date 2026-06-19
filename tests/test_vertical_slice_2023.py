"""Tests for the 2023 thin downstream vertical slice (VS-2023-001..011).

These tests define what "done" means for the 2023 vertical slice described in
``docs/architecture/vertical-slice-2023.md``. They are written **before** the
production module exists; every test fails until the production seam is
implemented. The tests are organized to mirror the architecture doc sections:

  1. Client parser (section 5)
  2. Country / country-year seeding (sections 6, 4)
  3. Observation linking (section 6)
  4. Client-only leader / ruler-year seeding (section 7)
  5. Provisional scoring (section 8)
  6. Confidence / validation (sections 8, 9)
  7. Outputs (section 9)
  8. CLI boundary (section 10)
  9. Idempotency + non-slice preservation (section 4)

Expected production seam:
    src/leaders_db/vertical_slice/slice_2023.py

Public symbols (per architecture doc section 10):
    - VerticalSliceResult       # Pydantic/dataclass result model
    - ClientSliceRow            # Pydantic/dataclass row model
    - load_vertical_slice_client_rows(path, sheet, year, iso3_scope) -> list[ClientSliceRow]
    - run_vertical_slice_2023(config, *, countries=None, categories=None,
                              run_adapters=True) -> VerticalSliceResult

Expected CLI command:
    leaders-db run-vertical-slice-2023 --config configs/prototype-2023.yaml
        [--countries MEX,NGA,USA] [--categories social_wellbeing,integrity]
        [--no-run-adapters]

Scope:
    - Year:        2023
    - Countries:   MEX, NGA, USA
    - Categories:  social_wellbeing, integrity

Harness adequacy note
---------------------
A pure unit test against an in-memory dict would pass while real use fails.
Therefore the tests in this file:

    * seed the database with realistic Source and SourceObservation rows
      (UNDP HDI and WGI, with source_row_reference like "undp_hdi:MEX")
      so the slice's Stage 3 linking path is exercised end-to-end against
      a real SQLite database;
    * drive the public ``run_vertical_slice_2023`` Python seam (not a
      mock) so missing wiring (no DB writes, no output files) causes the
      assertion to fail;
    * drive the ``leaders-db run-vertical-slice-2023`` CLI command via
      Typer's CliRunner so a no-op stub that exits 0 but produces no
      outputs is caught by file existence assertions.
"""

from __future__ import annotations

import csv
import inspect
import shutil
from collections.abc import Sequence
from pathlib import Path

import pytest
from openpyxl import Workbook
from sqlalchemy import func, select
from typer.testing import CliRunner

from leaders_db.cli import app
from leaders_db.config import RunConfig
from leaders_db.db.engine import init_database
from leaders_db.db.models import (
    Country,
    CountryYear,
    Leader,
    RulerScore,
    RulerSpell,
    RulerYear,
    ScoreCategory,
    Source,
    SourceObservation,
    ValidationResult,
)
from leaders_db.db.session import session_scope
from leaders_db.normalize.countries import normalize_country_name

# ---------------------------------------------------------------------------
# Production seam import (graceful when the module is not implemented yet).
# ---------------------------------------------------------------------------
#
# Try-importing the production module under test. When the module does not
# exist (the state at the time of writing), every public name is set to
# ``None`` and each test asserts ``is not None`` first, producing a clear,
# targeted failure that names the missing symbol.

try:
    from leaders_db.vertical_slice import slice_2023
    from leaders_db.vertical_slice.slice_2023 import (
        ClientSliceRow,
        VerticalSliceResult,
        load_vertical_slice_client_rows,
        run_vertical_slice_2023,
    )
except ImportError:  # pragma: no cover - exercised only when module absent
    slice_2023 = None  # type: ignore[assignment]
    ClientSliceRow = None  # type: ignore[assignment]
    VerticalSliceResult = None  # type: ignore[assignment]
    load_vertical_slice_client_rows = None  # type: ignore[assignment]
    run_vertical_slice_2023 = None  # type: ignore[assignment]


# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

DEFAULT_ISO3_SCOPE: tuple[str, ...] = ("MEX", "NGA", "USA")
DEFAULT_CATEGORIES: tuple[str, ...] = ("social_wellbeing", "integrity")
TARGET_YEAR: int = 2023

# Realistic test inputs (no invented data; values match the live UNDP HDI
# 2022 values for these countries and plausible WGI 2022 z-scores).
EXPECTED_UNDP_HDI_BY_ISO3: dict[str, float] = {
    "MEX": 0.781,
    "NGA": 0.548,
    "USA": 0.927,
}
EXPECTED_WGI_Z_BY_ISO3: dict[str, float] = {
    "MEX": -0.7,
    "NGA": -1.0,
    "USA": 1.5,
}

# Slice-owned markers per architecture doc §4.
SLICE_NOTE_PREFIX = "vertical_slice_2023:"
SLICE_RULER_SPELLS_DATASET = "vertical_slice_client_seed"
SLICE_INCLUSION_REASON = "vertical_slice_2023_selected_country"

# Multi-year time-series output (see section 10 below). The slice writes one
# ``vertical_slice_timeseries.csv`` per run when the caller passes a
# ``years=`` kwarg / ``--years`` CLI flag.
TIMESERIES_OUTPUT_CSV = "vertical_slice_timeseries.csv"

# Realistic multi-year test inputs (no invented data; values match the live
# UNDP HDI 2020/2021/2022 values for these countries and plausible WGI
# z-scores for the same window). 2023 is intentionally NOT seeded here so
# the slice's documented 2022 proxy is exercised end-to-end.
EXPECTED_UNDP_HDI_BY_ISO3_BY_YEAR: dict[int, dict[str, float]] = {
    2020: {"MEX": 0.770, "NGA": 0.530, "USA": 0.920},
    2021: {"MEX": 0.775, "NGA": 0.540, "USA": 0.923},
    2022: {"MEX": 0.781, "NGA": 0.548, "USA": 0.927},
}
EXPECTED_WGI_Z_BY_ISO3_BY_YEAR: dict[int, dict[str, float]] = {
    2020: {"MEX": -0.65, "NGA": -0.95, "USA": 1.45},
    2021: {"MEX": -0.68, "NGA": -0.98, "USA": 1.48},
    2022: {"MEX": -0.70, "NGA": -1.00, "USA": 1.50},
}


# ---------------------------------------------------------------------------
# xlsx fixture builder
# ---------------------------------------------------------------------------


def _build_client_matrix_xlsx(path: Path) -> None:
    """Build a small client matrix xlsx that preserves the weird header/data offsets.

    Layout (matches the real client xlsx enough to prove parser behaviour):

        - xlsx rows 1-5 (pandas 0-based rows 0-4) are noise / metadata /
          column-name hints. The parser must skip them and start reading
          data at pandas row 5 (xlsx row 6).
        - xlsx row 5 carries column-name hints in the documented positions
          (country=col 4 1-based, etc.) so a debugging reader can verify
          the offset.
        - xlsx row 6 onward are data rows. We include MEX, NGA, USA (in
          scope) plus BRA (out of scope) to prove the ISO3-scope filter.

    Column mapping (pandas zero-based, from architecture doc §5):
        country:           3
        population:        4
        leader:           14
        year_started:     15
        social_wellbeing: 22
        integrity:        23
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Main"

    # Rows 1..5 (1-based) = pandas 0..4. Noise + a column-name hint row.
    ws.cell(row=1, column=1, value="LATEST BARR'S POLITICAL MATRIX (fixture)")
    ws.cell(row=2, column=1, value="Generated for vertical-slice-2023 tests only")
    ws.cell(row=3, column=1, value="Synthetic values; do not redistribute")
    ws.cell(row=4, column=1, value="(rows 0-4 are intentionally non-data)")
    # xlsx row 5 = pandas row 4 — column-name hints in documented positions.
    ws.cell(row=5, column=4, value="Country")
    ws.cell(row=5, column=5, value="Population")
    ws.cell(row=5, column=15, value="Leader")
    ws.cell(row=5, column=16, value="Year started")
    ws.cell(row=5, column=23, value="social_wellbeing")
    ws.cell(row=5, column=24, value="integrity")

    # xlsx row 6 onward = pandas row 5 onward. Data rows.
    data_rows = [
        # (country, pop, leader, year_started, social, integrity)
        ("Mexico", 128_000_000, "Andrés Manuel López Obrador", 2018, 7, 5),
        ("Nigeria", 220_000_000, "Bola Tinubu", 2023, 4, 3),
        ("United States", 333_000_000, "Joe Biden", 2021, 8, 8),
        ("Brazil", 215_000_000, "Luiz Inácio Lula da Silva", 2023, 6, 4),
    ]
    for i, row in enumerate(data_rows, start=6):
        country, pop, leader, year_started, social, integrity = row
        ws.cell(row=i, column=4, value=country)
        ws.cell(row=i, column=5, value=pop)
        ws.cell(row=i, column=15, value=leader)
        ws.cell(row=i, column=16, value=year_started)
        ws.cell(row=i, column=23, value=social)
        ws.cell(row=i, column=24, value=integrity)

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------


@pytest.fixture(scope="session")
def client_matrix_xlsx_path() -> Path:
    """Return the path to the committed client matrix fixture xlsx.

    Builds the file on first use (so the test file is self-contained and
    the fixture xlsx is committed alongside it on first CI run).
    """
    fixtures_dir = Path(__file__).resolve().parent / "fixtures" / "vertical_slice"
    fixtures_dir.mkdir(parents=True, exist_ok=True)
    xlsx_path = fixtures_dir / "client_matrix_sample.xlsx"
    if not xlsx_path.exists():
        _build_client_matrix_xlsx(xlsx_path)
    return xlsx_path


@pytest.fixture()
def client_xlsx_in_data_lake(
    isolated_data_lake: Path,
    client_matrix_xlsx_path: Path,
) -> Path:
    """Stage the client xlsx under ``data/raw/client_existing/`` in the test data lake.

    The production seam locates the client xlsx via
    :func:`leaders_db.paths.raw_dir('client_existing')` so this is the
    conventional staging location.
    """
    target = isolated_data_lake / "data" / "raw" / "client_existing"
    target.mkdir(parents=True, exist_ok=True)
    dst = target / "client_matrix_fixture.xlsx"
    shutil.copy2(client_matrix_xlsx_path, dst)
    return dst


def _upsert_source(
    session,
    *,
    source_name: str,
    source_type: str = "official",
    source_url: str | None = None,
    version: str | None = None,
) -> Source:
    existing = session.execute(
        select(Source).where(Source.source_name == source_name)
    ).scalar_one_or_none()
    if existing is not None:
        return existing
    row = Source(
        source_name=source_name,
        source_type=source_type,
        source_url=source_url,
        version=version,
    )
    session.add(row)
    session.flush()
    return row


@pytest.fixture()
def seeded_stage2_observations(database_url: str) -> dict[str, int]:
    """Seed UNDP HDI and WGI observations for MEX/NGA/USA at year=2022 (proxy).

    The slice is expected to link these observations to countries via the
    ``source_row_reference`` pattern (``undp_hdi:<ISO3>``, ``wgi:<ISO3>``).
    """
    init_database(database_url)
    with session_scope(database_url) as session:
        undp = _upsert_source(
            session,
            source_name="UNDP HDR 2023-24 (test fixture)",
            version="2023-24",
        )
        wgi = _upsert_source(
            session,
            source_name="World Bank WGI (test fixture)",
            version="2023",
        )
        for iso3, hdi in EXPECTED_UNDP_HDI_BY_ISO3.items():
            session.add(
                SourceObservation(
                    source_id=undp.id,
                    country_id=None,  # Stage 2 leaves NULL; slice fills it
                    leader_id=None,
                    year=2022,
                    variable_name="undp_hdi_hdi",
                    raw_value=str(hdi),
                    normalized_value=hdi,
                    unit="index",
                    source_row_reference=f"undp_hdi:{iso3}",
                )
            )
        for iso3, z in EXPECTED_WGI_Z_BY_ISO3.items():
            session.add(
                SourceObservation(
                    source_id=wgi.id,
                    country_id=None,
                    leader_id=None,
                    year=2022,
                    variable_name="wgi_control_of_corruption",
                    raw_value=str(z),
                    normalized_value=None,  # raw z-score; slice normalizes
                    unit="z_score",
                    source_row_reference=f"wgi:{iso3}",
                )
            )
    return {"undp_hdi": undp.id, "wgi": wgi.id}


@pytest.fixture()
def seeded_score_categories(database_url: str) -> dict[str, int]:
    """Seed the two in-scope ``score_categories`` rows.

    The production seam may also seed them itself; this fixture exists so
    the ``run_adapters=False`` test path does not depend on the seam
    re-seeding them.
    """
    init_database(database_url)
    out: dict[str, int] = {}
    with session_scope(database_url) as session:
        for key, name in (
            ("social_wellbeing", "Social Wellbeing"),
            ("integrity", "Integrity"),
        ):
            existing = session.execute(
                select(ScoreCategory).where(ScoreCategory.category_key == key)
            ).scalar_one_or_none()
            if existing is not None:
                out[key] = existing.id
                continue
            row = ScoreCategory(category_key=key, category_name=name)
            session.add(row)
            session.flush()
            out[key] = row.id
    return out


@pytest.fixture()
def seeded_multi_year_observations(
    database_url: str,
) -> dict[str, dict[int, dict[str, float]]]:
    """Seed UNDP HDI and WGI observations for MEX/NGA/USA at years 2020-2022.

    No 2023 observation is seeded. The slice's 2023 row is expected to use
    2022 as a documented proxy (architecture §8), while the 2020/2021/2022
    rows are expected to use the exact source year (``source_kind='direct'``).

    Returns a mapping shaped like::

        {
            "undp": {2020: {"MEX": 0.770, ...}, 2021: {...}, 2022: {...}},
            "wgi": {2020: {"MEX": -0.65, ...}, 2021: {...}, 2022: {...}},
        }

    so tests can assert on the exact seeded values.
    """
    init_database(database_url)
    with session_scope(database_url) as session:
        undp = _upsert_source(
            session,
            source_name="UNDP HDR 2023-24 (test fixture)",
            version="2023-24",
        )
        wgi = _upsert_source(
            session,
            source_name="World Bank WGI (test fixture)",
            version="2023",
        )
        for year, year_map in EXPECTED_UNDP_HDI_BY_ISO3_BY_YEAR.items():
            for iso3, hdi in year_map.items():
                session.add(
                    SourceObservation(
                        source_id=undp.id,
                        country_id=None,
                        leader_id=None,
                        year=year,
                        variable_name="undp_hdi_hdi",
                        raw_value=str(hdi),
                        normalized_value=hdi,
                        unit="index",
                        source_row_reference=f"undp_hdi:{iso3}",
                    )
                )
        for year, year_map in EXPECTED_WGI_Z_BY_ISO3_BY_YEAR.items():
            for iso3, z in year_map.items():
                session.add(
                    SourceObservation(
                        source_id=wgi.id,
                        country_id=None,
                        leader_id=None,
                        year=year,
                        variable_name="wgi_control_of_corruption",
                        raw_value=str(z),
                        normalized_value=None,
                        unit="z_score",
                        source_row_reference=f"wgi:{iso3}",
                    )
                )
    return {
        "undp": EXPECTED_UNDP_HDI_BY_ISO3_BY_YEAR,
        "wgi": EXPECTED_WGI_Z_BY_ISO3_BY_YEAR,
    }


# ---------------------------------------------------------------------------
# Helper used by every DB-bound integration test below.
# ---------------------------------------------------------------------------


def _run_slice(*, client_xlsx_in_data_lake: Path) -> object:
    """Invoke ``run_vertical_slice_2023`` with the standard test kwargs.

    The seam resolves the database URL and the client xlsx path through
    the ``LEADERSDB_PROJECT_ROOT`` env var (set by ``isolated_data_lake``)
    and :func:`leaders_db.paths.raw_dir`, so no explicit kwargs are needed.
    ``run_adapters=False`` keeps the test offline and deterministic — the
    seam reads the seeded ``SourceObservation`` rows directly.
    """
    assert run_vertical_slice_2023 is not None, (
        "src/leaders_db/vertical_slice/slice_2023.py is not implemented; "
        "the test-builder cannot drive the seam."
    )
    return run_vertical_slice_2023(
        RunConfig(),
        countries=DEFAULT_ISO3_SCOPE,
        categories=DEFAULT_CATEGORIES,
        run_adapters=False,
    )


def _run_slice_with_years(
    *,
    client_xlsx_in_data_lake: Path,
    years: Sequence[int] | None = None,
) -> object:
    """Invoke ``run_vertical_slice_2023`` with an optional ``years`` kwarg.

    The multi-year time-series output feature is driven by passing a
    sequence of years (e.g. ``(2020, 2021, 2022, 2023)``). The current
    production seam does not yet accept this parameter; the helper
    inspects the signature first and surfaces a clear failure with the
    expected error message rather than the raw ``TypeError`` that an
    unknown kwarg would raise.
    """
    assert run_vertical_slice_2023 is not None, (
        "src/leaders_db/vertical_slice/slice_2023.py is not implemented; "
        "the test-builder cannot drive the seam."
    )
    sig = inspect.signature(run_vertical_slice_2023)
    if "years" not in sig.parameters:
        pytest.fail(
            "run_vertical_slice_2023 does not yet accept a 'years' kwarg; "
            "the multi-year time-series output feature is not implemented "
            "(see test_vertical_slice_2023.py section 10)."
        )
    return run_vertical_slice_2023(
        RunConfig(),
        countries=DEFAULT_ISO3_SCOPE,
        categories=DEFAULT_CATEGORIES,
        run_adapters=False,
        years=years,
    )


# ---------------------------------------------------------------------------
# 0. Module surface (smoke)
# ---------------------------------------------------------------------------


def test_vertical_slice_module_exports() -> None:
    """The slice module exposes the four public symbols from architecture §10."""
    assert slice_2023 is not None, (
        "src/leaders_db/vertical_slice/slice_2023.py is not implemented yet; "
        "expected module to be importable."
    )
    assert hasattr(slice_2023, "VerticalSliceResult")
    assert hasattr(slice_2023, "ClientSliceRow")
    assert hasattr(slice_2023, "load_vertical_slice_client_rows")
    assert hasattr(slice_2023, "run_vertical_slice_2023")


# ---------------------------------------------------------------------------
# 1. Client parser (architecture doc §5)
# ---------------------------------------------------------------------------


def test_load_client_rows_filters_to_iso3_scope(
    client_xlsx_in_data_lake: Path,
) -> None:
    """The parser returns only MEX/NGA/USA rows; the Brazil row is filtered out."""
    assert load_vertical_slice_client_rows is not None
    rows = load_vertical_slice_client_rows(
        path=client_xlsx_in_data_lake,
        sheet="Main",
        year=TARGET_YEAR,
        iso3_scope=DEFAULT_ISO3_SCOPE,
    )
    assert len(rows) == 3, f"expected 3 rows, got {len(rows)}"
    iso3s = {row.iso3 for row in rows}
    assert iso3s == {"MEX", "NGA", "USA"}


def test_load_client_rows_preserves_source_row_number(
    client_xlsx_in_data_lake: Path,
) -> None:
    """Each ClientSliceRow carries the source row number for audit + outputs."""
    assert load_vertical_slice_client_rows is not None
    rows = load_vertical_slice_client_rows(
        path=client_xlsx_in_data_lake,
        sheet="Main",
        year=TARGET_YEAR,
        iso3_scope=DEFAULT_ISO3_SCOPE,
    )
    assert len(rows) == 3
    for row in rows:
        assert isinstance(row.source_row_number, int)
        # pandas zero-based row 5 = first data row. The 3 scoped rows live
        # at pandas rows 5, 6, 7 (the Brazil row at row 8 is filtered).
        assert row.source_row_number >= 5


def test_load_client_rows_validates_social_and_integrity_in_range(
    client_xlsx_in_data_lake: Path,
) -> None:
    """social_wellbeing and integrity scores are int 0..10 or None."""
    assert load_vertical_slice_client_rows is not None
    rows = load_vertical_slice_client_rows(
        path=client_xlsx_in_data_lake,
        sheet="Main",
        year=TARGET_YEAR,
        iso3_scope=DEFAULT_ISO3_SCOPE,
    )
    assert len(rows) == 3
    for row in rows:
        social = row.client_scores.get("social_wellbeing")
        integrity = row.client_scores.get("integrity")
        assert social is None or isinstance(social, int), (
            f"social_wellbeing not int for {row.iso3}: {social!r}"
        )
        assert integrity is None or isinstance(integrity, int), (
            f"integrity not int for {row.iso3}: {integrity!r}"
        )
        if social is not None:
            assert 0 <= social <= 10, f"social_wellbeing OOR for {row.iso3}: {social}"
        if integrity is not None:
            assert 0 <= integrity <= 10, f"integrity OOR for {row.iso3}: {integrity}"


def test_load_client_rows_returns_expected_values(
    client_xlsx_in_data_lake: Path,
) -> None:
    """The fixture yields the documented (iso3 -> scores) map."""
    assert load_vertical_slice_client_rows is not None
    rows = load_vertical_slice_client_rows(
        path=client_xlsx_in_data_lake,
        sheet="Main",
        year=TARGET_YEAR,
        iso3_scope=DEFAULT_ISO3_SCOPE,
    )
    by_iso3 = {row.iso3: row for row in rows}
    assert by_iso3["MEX"].client_scores["social_wellbeing"] == 7
    assert by_iso3["MEX"].client_scores["integrity"] == 5
    assert by_iso3["NGA"].client_scores["social_wellbeing"] == 4
    assert by_iso3["NGA"].client_scores["integrity"] == 3
    assert by_iso3["USA"].client_scores["social_wellbeing"] == 8
    assert by_iso3["USA"].client_scores["integrity"] == 8
    # Leader strings are non-empty (architecture §5: "leader string must be non-empty").
    for iso3 in ("MEX", "NGA", "USA"):
        assert by_iso3[iso3].leader_name and by_iso3[iso3].leader_name.strip()


def test_load_client_rows_missing_country_raises(
    isolated_data_lake: Path,
    client_matrix_xlsx_path: Path,
) -> None:
    """A scoped ISO3 not present in the xlsx raises (single-row requirement)."""
    assert load_vertical_slice_client_rows is not None
    # Build an alt xlsx with only MEX and NGA (no USA).
    alt_dir = isolated_data_lake / "data" / "raw" / "client_existing"
    alt_dir.mkdir(parents=True, exist_ok=True)
    alt_path = alt_dir / "client_matrix_no_usa.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "Main"
    # Mirror the main fixture's layout so the offset logic still applies.
    for col, value in (
        (4, "Country"),
        (5, "Population"),
        (15, "Leader"),
        (16, "Year started"),
        (23, "social_wellbeing"),
        (24, "integrity"),
    ):
        ws.cell(row=5, column=col, value=value)
    for i, row in enumerate(
        [
            ("Mexico", 128_000_000, "Andrés Manuel López Obrador", 2018, 7, 5),
            ("Nigeria", 220_000_000, "Bola Tinubu", 2023, 4, 3),
        ],
        start=6,
    ):
        country, pop, leader, year_started, social, integrity = row
        ws.cell(row=i, column=4, value=country)
        ws.cell(row=i, column=5, value=pop)
        ws.cell(row=i, column=15, value=leader)
        ws.cell(row=i, column=16, value=year_started)
        ws.cell(row=i, column=23, value=social)
        ws.cell(row=i, column=24, value=integrity)
    wb.save(alt_path)

    # A scoped ISO3 not present in the xlsx must raise. The implementation
    # chooses the exception type (ValueError / KeyError / custom); we catch
    # the broad Exception because the contract is "raises", not the type.
    with pytest.raises(Exception):  # noqa: B017 - contract is "raises", not type
        load_vertical_slice_client_rows(
            path=alt_path,
            sheet="Main",
            year=TARGET_YEAR,
            iso3_scope=DEFAULT_ISO3_SCOPE,  # USA missing from xlsx!
        )


def test_load_client_rows_duplicate_iso3_raises(
    isolated_data_lake: Path,
) -> None:
    """A duplicate scoped country row in the xlsx raises instead of overwriting.

    Architecture §5 requires exactly one row per scoped ISO3. The parser
    must surface a duplicate as a ValueError naming the offending ISO3
    and the two row numbers so the caller can fix the input file.
    """
    assert load_vertical_slice_client_rows is not None
    dup_dir = isolated_data_lake / "data" / "raw" / "client_existing"
    dup_dir.mkdir(parents=True, exist_ok=True)
    dup_path = dup_dir / "client_matrix_duplicate_mex.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "Main"
    # Same layout hints as the main fixture so the offset logic still applies.
    for col, value in (
        (4, "Country"),
        (5, "Population"),
        (15, "Leader"),
        (16, "Year started"),
        (23, "social_wellbeing"),
        (24, "integrity"),
    ):
        ws.cell(row=5, column=col, value=value)

    # First MEX row at xlsx row 6.
    ws.cell(row=6, column=4, value="Mexico")
    ws.cell(row=6, column=5, value=128_000_000)
    ws.cell(row=6, column=15, value="Andrés Manuel López Obrador")
    ws.cell(row=6, column=16, value=2018)
    ws.cell(row=6, column=23, value=7)
    ws.cell(row=6, column=24, value=5)

    # Duplicate MEX row at xlsx row 7 (e.g. two rows for the same country).
    ws.cell(row=7, column=4, value="Mexico")
    ws.cell(row=7, column=5, value=128_000_000)
    ws.cell(row=7, column=15, value="Andrés Manuel López Obrador")
    ws.cell(row=7, column=16, value=2018)
    ws.cell(row=7, column=23, value=7)
    ws.cell(row=7, column=24, value=5)

    # Plus NGA and USA so the ISO3 coverage is otherwise complete.
    for i, row in enumerate(
        [
            ("Nigeria", 220_000_000, "Bola Tinubu", 2023, 4, 3),
            ("United States", 333_000_000, "Joe Biden", 2021, 8, 8),
        ],
        start=8,
    ):
        country, pop, leader, year_started, social, integrity = row
        ws.cell(row=i, column=4, value=country)
        ws.cell(row=i, column=5, value=pop)
        ws.cell(row=i, column=15, value=leader)
        ws.cell(row=i, column=16, value=year_started)
        ws.cell(row=i, column=23, value=social)
        ws.cell(row=i, column=24, value=integrity)

    wb.save(dup_path)

    with pytest.raises(ValueError) as exc_info:
        load_vertical_slice_client_rows(
            path=dup_path,
            sheet="Main",
            year=TARGET_YEAR,
            iso3_scope=DEFAULT_ISO3_SCOPE,
        )
    msg = str(exc_info.value)
    assert "MEX" in msg, f"error should name the duplicate ISO3, got: {msg!r}"
    assert "duplicate" in msg.lower(), (
        f"error should describe the issue as a duplicate, got: {msg!r}"
    )


# ---------------------------------------------------------------------------
# 2. Country / country-year seeding (architecture doc §6)
# ---------------------------------------------------------------------------


def test_run_seeds_three_countries_with_iso3(
    client_xlsx_in_data_lake: Path,
    seeded_stage2_observations: dict[str, int],
    seeded_score_categories: dict[str, int],
    database_url: str,
) -> None:
    """The seam upserts MEX/NGA/USA countries with normalized names."""
    init_database(database_url)
    _run_slice(client_xlsx_in_data_lake=client_xlsx_in_data_lake)

    with session_scope(database_url) as session:
        iso3s = sorted(
            row.iso3 for row in session.execute(select(Country)).scalars()
        )
    assert iso3s == ["MEX", "NGA", "USA"]
    with session_scope(database_url) as session:
        for row in session.execute(select(Country)).scalars():
            assert row.country_name and row.country_name_normalized, (
                f"country {row.iso3} missing name/normalized"
            )


def test_run_creates_country_years_for_2023(
    client_xlsx_in_data_lake: Path,
    seeded_stage2_observations: dict[str, int],
    seeded_score_categories: dict[str, int],
    database_url: str,
) -> None:
    """One country_years row per scoped country for year 2023, included_in_project=True."""
    init_database(database_url)
    _run_slice(client_xlsx_in_data_lake=client_xlsx_in_data_lake)

    with session_scope(database_url) as session:
        rows = session.execute(
            select(CountryYear, Country)
            .join(Country, CountryYear.country_id == Country.id)
        ).all()
    assert len(rows) == 3
    for cy, country in rows:
        assert country.iso3 in DEFAULT_ISO3_SCOPE
        assert cy.year == TARGET_YEAR
        assert cy.included_in_project is True
        assert cy.inclusion_reason == SLICE_INCLUSION_REASON, (
            f"inclusion_reason mismatch: {cy.inclusion_reason!r}"
        )


def test_run_idempotent_country_seeding(
    client_xlsx_in_data_lake: Path,
    seeded_stage2_observations: dict[str, int],
    seeded_score_categories: dict[str, int],
    database_url: str,
) -> None:
    """A second run does not duplicate countries or country_years."""
    init_database(database_url)
    _run_slice(client_xlsx_in_data_lake=client_xlsx_in_data_lake)
    _run_slice(client_xlsx_in_data_lake=client_xlsx_in_data_lake)

    with session_scope(database_url) as session:
        country_count = session.execute(
            select(func.count(Country.id))
        ).scalar_one()
        cy_count = session.execute(
            select(func.count(CountryYear.id))
        ).scalar_one()
    assert country_count == 3
    assert cy_count == 3


# ---------------------------------------------------------------------------
# 3. Observation linking (architecture doc §6)
# ---------------------------------------------------------------------------


def test_link_observations_by_source_row_reference(
    client_xlsx_in_data_lake: Path,
    seeded_stage2_observations: dict[str, int],
    seeded_score_categories: dict[str, int],
    database_url: str,
) -> None:
    """Observations with source_row_reference like 'undp_hdi:<ISO3>' get country_id set."""
    init_database(database_url)
    _run_slice(client_xlsx_in_data_lake=client_xlsx_in_data_lake)

    with session_scope(database_url) as session:
        countries = {
            c.iso3: c.id for c in session.execute(select(Country)).scalars()
        }
        assert set(countries) == {"MEX", "NGA", "USA"}, (
            "expected countries seeded before linking observations"
        )
        obs_rows = session.execute(
            select(SourceObservation).where(
                SourceObservation.source_row_reference.like("%:%"),
            )
        ).scalars().all()

    assert len(obs_rows) == 6, (
        f"expected 6 seeded observations (3 UNDP + 3 WGI), got {len(obs_rows)}"
    )
    for obs in obs_rows:
        assert obs.country_id is not None, (
            f"obs {obs.id} source_row_reference={obs.source_row_reference} "
            f"was not linked to a country"
        )
        _prefix, _, iso3 = obs.source_row_reference.partition(":")
        assert iso3 in countries, f"unexpected ISO3 suffix {iso3!r}"
        assert obs.country_id == countries[iso3], (
            f"obs {obs.id} ({obs.source_row_reference}) country_id={obs.country_id} "
            f"but expected country_id for ISO3={iso3} ({countries[iso3]})"
        )


def test_link_observations_does_not_overwrite_conflicting_country_id(
    client_xlsx_in_data_lake: Path,
    seeded_stage2_observations: dict[str, int],
    seeded_score_categories: dict[str, int],
    database_url: str,
) -> None:
    """An observation whose country_id already points elsewhere survives linking unchanged."""
    init_database(database_url)

    # Pre-seed a non-slice country (BRA) and a conflict observation:
    # source_row_reference says undp_hdi:MEX, but country_id points to BRA.
    # The slice must not overwrite this row (architecture §6:
    # "Only update rows whose country_id is NULL or already points to the same ISO3").
    with session_scope(database_url) as session:
        bra = Country(
            iso3="BRA",
            country_name="Brazil",
            country_name_normalized=normalize_country_name("Brazil"),
        )
        session.add(bra)
        session.flush()
        bra_id = bra.id

        undp_source = session.execute(
            select(Source).where(Source.source_name.like("UNDP HDR%"))
        ).scalar_one()
        session.add(
            SourceObservation(
                source_id=undp_source.id,
                country_id=bra_id,
                leader_id=None,
                year=2022,
                variable_name="undp_hdi_hdi",
                raw_value="0.999",
                normalized_value=0.999,
                unit="index",
                source_row_reference="undp_hdi:MEX",
            )
        )

    _run_slice(client_xlsx_in_data_lake=client_xlsx_in_data_lake)

    with session_scope(database_url) as session:
        rows = session.execute(
            select(SourceObservation).where(
                SourceObservation.source_row_reference == "undp_hdi:MEX",
            )
        ).scalars().all()

    # At least one row must still point at BRA (the conflict row was preserved).
    bra_rows = [r for r in rows if r.country_id == bra_id]
    assert bra_rows, (
        "conflict observation (source_row_reference='undp_hdi:MEX', "
        "country_id=BRA) was overwritten; the slice must preserve conflicting "
        "non-slice country links (architecture §6)."
    )


# ---------------------------------------------------------------------------
# 4. Client-only leader / ruler-year seeding (architecture doc §7)
# ---------------------------------------------------------------------------


def test_run_seeds_client_only_ruler_year(
    client_xlsx_in_data_lake: Path,
    seeded_stage2_observations: dict[str, int],
    seeded_score_categories: dict[str, int],
    database_url: str,
) -> None:
    """ruler_years carry match_status='client_only' and review_status='manual_review_required'."""
    init_database(database_url)
    _run_slice(client_xlsx_in_data_lake=client_xlsx_in_data_lake)

    with session_scope(database_url) as session:
        rows = session.execute(
            select(RulerYear, Country)
            .join(Country, RulerYear.country_id == Country.id)
        ).all()
    assert len(rows) == 3
    for ry, country in rows:
        assert ry.match_status == "client_only", (
            f"ruler_year for {country.iso3} match_status={ry.match_status!r}"
        )
        assert ry.review_status == "manual_review_required"
        assert ry.review_note is not None
        assert ry.review_note.startswith(SLICE_NOTE_PREFIX)


def test_run_separate_client_and_system_leader_names(
    client_xlsx_in_data_lake: Path,
    seeded_stage2_observations: dict[str, int],
    seeded_score_categories: dict[str, int],
    database_url: str,
) -> None:
    """client_matrix_leader_name and system_selected_leader_name are both populated."""
    init_database(database_url)
    _run_slice(client_xlsx_in_data_lake=client_xlsx_in_data_lake)

    with session_scope(database_url) as session:
        rows = session.execute(
            select(RulerYear, Country)
            .join(Country, RulerYear.country_id == Country.id)
        ).all()
    by_iso3 = {country.iso3: ry for ry, country in rows}

    assert by_iso3["MEX"].client_matrix_leader_name == "Andrés Manuel López Obrador"
    assert by_iso3["NGA"].client_matrix_leader_name == "Bola Tinubu"
    assert by_iso3["USA"].client_matrix_leader_name == "Joe Biden"

    # system_selected_leader_name is also populated (separate string column).
    for iso3 in ("MEX", "NGA", "USA"):
        assert by_iso3[iso3].system_selected_leader_name, (
            f"system_selected_leader_name empty for {iso3}"
        )


def test_run_creates_ruler_spells_with_slice_marker(
    client_xlsx_in_data_lake: Path,
    seeded_stage2_observations: dict[str, int],
    seeded_score_categories: dict[str, int],
    database_url: str,
) -> None:
    """ruler_spells.source_dataset='vertical_slice_client_seed' and notes start with the marker."""
    init_database(database_url)
    _run_slice(client_xlsx_in_data_lake=client_xlsx_in_data_lake)

    with session_scope(database_url) as session:
        rows = session.execute(select(RulerSpell)).scalars().all()
    assert len(rows) == 3
    for spell in rows:
        assert spell.source_dataset == SLICE_RULER_SPELLS_DATASET
        assert spell.notes is not None
        assert spell.notes.startswith(SLICE_NOTE_PREFIX)


# ---------------------------------------------------------------------------
# 5. Provisional scoring (architecture doc §8)
# ---------------------------------------------------------------------------


def test_run_writes_ruler_scores_for_social_and_integrity(
    client_xlsx_in_data_lake: Path,
    seeded_stage2_observations: dict[str, int],
    seeded_score_categories: dict[str, int],
    database_url: str,
) -> None:
    """The seam writes 3 social_wellbeing + 3 integrity rows (no other categories)."""
    init_database(database_url)
    _run_slice(client_xlsx_in_data_lake=client_xlsx_in_data_lake)

    with session_scope(database_url) as session:
        rows = session.execute(
            select(RulerScore, ScoreCategory)
            .join(ScoreCategory, RulerScore.category_id == ScoreCategory.id)
        ).all()
    cats = {cat.category_key for _, cat in rows}
    assert cats == {"social_wellbeing", "integrity"}, (
        f"expected only the two in-scope categories, got {cats}"
    )
    assert len(rows) == 6


def test_run_carries_client_system_final_delta_separately(
    client_xlsx_in_data_lake: Path,
    seeded_stage2_observations: dict[str, int],
    seeded_score_categories: dict[str, int],
    database_url: str,
) -> None:
    """client_score + system_proposed_score populated; final_score NULL; delta computed."""
    init_database(database_url)
    _run_slice(client_xlsx_in_data_lake=client_xlsx_in_data_lake)

    with session_scope(database_url) as session:
        rows = session.execute(select(RulerScore)).scalars().all()
    assert len(rows) == 6
    for rs in rows:
        assert rs.client_score is not None, "client_score NULL (client invariant)"
        assert rs.system_proposed_score is not None, "system_proposed_score NULL"
        assert rs.final_score is None, (
            f"final_score must remain NULL on slice output, got {rs.final_score}"
        )
        expected_delta = rs.system_proposed_score - rs.client_score
        assert rs.score_delta_vs_client == expected_delta, (
            f"score_delta_vs_client={rs.score_delta_vs_client} but "
            f"system - client = {expected_delta}"
        )


def test_run_does_not_write_categories_outside_scope(
    client_xlsx_in_data_lake: Path,
    seeded_stage2_observations: dict[str, int],
    seeded_score_categories: dict[str, int],
    database_url: str,
) -> None:
    """Adding a third category does not produce ruler_scores rows for it.

    The original assertion compared the full sorted list of category_keys
    returned by the join against ``["integrity", "social_wellbeing"]``
    (2 items). With 6 ruler_scores rows (3 countries x 2 categories — the
    same 6 rows asserted by ``test_run_writes_ruler_scores_for_social_and_integrity``),
    the inner join returns 6 rows, so the list-equality assertion cannot
    hold. The intent here is clearly set-membership (no score may exist
    for the out-of-scope category, and the in-scope categories must be
    represented), so we assert set equality instead.
    """
    init_database(database_url)

    # Add a third category (political_freedom) that is NOT in scope. The
    # slice must not write ruler_scores rows for it.
    with session_scope(database_url) as session:
        session.add(
            ScoreCategory(category_key="political_freedom", category_name="Political Freedom")
        )

    _run_slice(client_xlsx_in_data_lake=client_xlsx_in_data_lake)

    with session_scope(database_url) as session:
        cat_keys = {
            key
            for (key,) in session.execute(
                select(ScoreCategory.category_key)
                .join(RulerScore, RulerScore.category_id == ScoreCategory.id)
            ).all()
        }
    assert cat_keys == {"integrity", "social_wellbeing"}
    # And explicitly: the out-of-scope category must have no score rows.
    assert "political_freedom" not in cat_keys


def test_social_score_formula_hdi_times_10_clipped(
    client_xlsx_in_data_lake: Path,
    seeded_stage2_observations: dict[str, int],
    seeded_score_categories: dict[str, int],
    database_url: str,
) -> None:
    """social_wellbeing = round(10 * hdi), clipped to 0..10 (architecture §8)."""
    init_database(database_url)
    _run_slice(client_xlsx_in_data_lake=client_xlsx_in_data_lake)

    expected = {
        "MEX": max(0, min(10, round(10 * 0.781))),
        "NGA": max(0, min(10, round(10 * 0.548))),
        "USA": max(0, min(10, round(10 * 0.927))),
    }

    with session_scope(database_url) as session:
        rows = session.execute(
            select(RulerScore.system_proposed_score, Country.iso3)
            .join(RulerYear, RulerScore.ruler_year_id == RulerYear.id)
            .join(Country, RulerYear.country_id == Country.id)
            .join(ScoreCategory, RulerScore.category_id == ScoreCategory.id)
            .where(ScoreCategory.category_key == "social_wellbeing")
        ).all()
    by_iso3 = {iso3: score for score, iso3 in rows}
    assert by_iso3 == expected


def test_integrity_score_formula_zscore_to_zero_one(
    client_xlsx_in_data_lake: Path,
    seeded_stage2_observations: dict[str, int],
    seeded_score_categories: dict[str, int],
    database_url: str,
) -> None:
    """integrity = round(10 * clamp((z + 2.5) / 5.0, 0, 1)) (architecture §8)."""
    init_database(database_url)
    _run_slice(client_xlsx_in_data_lake=client_xlsx_in_data_lake)

    def expected_integrity(z: float) -> int:
        n = max(0.0, min(1.0, (z + 2.5) / 5.0))
        return max(0, min(10, round(10 * n)))

    expected = {
        iso3: expected_integrity(EXPECTED_WGI_Z_BY_ISO3[iso3])
        for iso3 in DEFAULT_ISO3_SCOPE
    }

    with session_scope(database_url) as session:
        rows = session.execute(
            select(RulerScore.system_proposed_score, Country.iso3)
            .join(RulerYear, RulerScore.ruler_year_id == RulerYear.id)
            .join(Country, RulerYear.country_id == Country.id)
            .join(ScoreCategory, RulerScore.category_id == ScoreCategory.id)
            .where(ScoreCategory.category_key == "integrity")
        ).all()
    by_iso3 = {iso3: score for score, iso3 in rows}
    assert by_iso3 == expected


# ---------------------------------------------------------------------------
# 6. Confidence / validation (architecture doc §8 + §9)
# ---------------------------------------------------------------------------


def test_run_writes_validation_row_per_score(
    client_xlsx_in_data_lake: Path,
    seeded_stage2_observations: dict[str, int],
    seeded_score_categories: dict[str, int],
    database_url: str,
) -> None:
    """One validation_results row per written ruler_scores row."""
    init_database(database_url)
    _run_slice(client_xlsx_in_data_lake=client_xlsx_in_data_lake)

    with session_scope(database_url) as session:
        score_count = session.execute(
            select(func.count(RulerScore.id))
        ).scalar_one()
        val_count = session.execute(
            select(func.count(ValidationResult.id))
        ).scalar_one()
    assert score_count == 6
    assert val_count == 6


def test_validation_row_carries_confidence_components(
    client_xlsx_in_data_lake: Path,
    seeded_stage2_observations: dict[str, int],
    seeded_score_categories: dict[str, int],
    database_url: str,
) -> None:
    """validation_results rows carry the four fixed-weight component scores.

    The final_confidence_score must equal round(0.35*agreement + 0.25*authority
    + 0.25*specificity + 0.15*temporal_fit) within a small tolerance.
    """
    init_database(database_url)
    _run_slice(client_xlsx_in_data_lake=client_xlsx_in_data_lake)

    with session_scope(database_url) as session:
        rows = session.execute(select(ValidationResult)).scalars().all()
    assert len(rows) == 6
    for v in rows:
        assert v.item_type == "ruler_score"
        for col in (
            v.source_agreement_score,
            v.source_authority_score,
            v.specificity_score,
            v.temporal_fit_score,
            v.final_confidence_score,
        ):
            assert col is not None
            assert 0 <= col <= 100, f"component out of 0..100: {col}"
        weighted = (
            0.35 * v.source_agreement_score
            + 0.25 * v.source_authority_score
            + 0.25 * v.specificity_score
            + 0.15 * v.temporal_fit_score
        )
        assert abs(weighted - v.final_confidence_score) <= 1, (
            f"fixed-weight formula mismatch: weighted={weighted} "
            f"final_confidence_score={v.final_confidence_score}"
        )


def test_validation_note_starts_with_slice_marker(
    client_xlsx_in_data_lake: Path,
    seeded_stage2_observations: dict[str, int],
    seeded_score_categories: dict[str, int],
    database_url: str,
) -> None:
    """Each validation_results.validation_note starts with 'vertical_slice_2023:'."""
    init_database(database_url)
    _run_slice(client_xlsx_in_data_lake=client_xlsx_in_data_lake)

    with session_scope(database_url) as session:
        notes = session.execute(
            select(ValidationResult.validation_note)
        ).scalars().all()
    assert notes, "no validation notes were written"
    for note in notes:
        assert note is not None
        assert note.startswith(SLICE_NOTE_PREFIX)


def test_validation_status_set_from_delta(
    client_xlsx_in_data_lake: Path,
    seeded_stage2_observations: dict[str, int],
    seeded_score_categories: dict[str, int],
    database_url: str,
) -> None:
    """validation_status is 'match' when |delta| <= 1, 'high_delta' when |delta| > 2."""
    init_database(database_url)
    _run_slice(client_xlsx_in_data_lake=client_xlsx_in_data_lake)

    with session_scope(database_url) as session:
        rows = session.execute(
            select(ValidationResult, RulerScore)
            .where(ValidationResult.item_type == "ruler_score")
            .where(ValidationResult.item_id == RulerScore.id)
        ).all()

    for v, rs in rows:
        delta = abs(rs.score_delta_vs_client or 0)
        if delta <= 1:
            assert v.validation_status == "match", (
                f"|delta|={delta} should yield 'match', got {v.validation_status}"
            )
        elif delta > 2:
            assert v.validation_status == "high_delta", (
                f"|delta|={delta} should yield 'high_delta', got {v.validation_status}"
            )


def test_rationale_short_starts_with_slice_marker(
    client_xlsx_in_data_lake: Path,
    seeded_stage2_observations: dict[str, int],
    seeded_score_categories: dict[str, int],
    database_url: str,
) -> None:
    """ruler_scores.rationale_short starts with 'vertical_slice_2023:'."""
    init_database(database_url)
    _run_slice(client_xlsx_in_data_lake=client_xlsx_in_data_lake)

    with session_scope(database_url) as session:
        rationales = session.execute(
            select(RulerScore.rationale_short)
        ).scalars().all()
    assert rationales
    for r in rationales:
        assert r is not None
        assert r.startswith(SLICE_NOTE_PREFIX)


# ---------------------------------------------------------------------------
# 7. Outputs (architecture doc §9)
# ---------------------------------------------------------------------------


def test_outputs_write_vertical_slice_scores_csv(
    client_xlsx_in_data_lake: Path,
    seeded_stage2_observations: dict[str, int],
    seeded_score_categories: dict[str, int],
    isolated_data_lake: Path,
    database_url: str,
) -> None:
    """vertical_slice_scores.csv is written with the required columns."""
    init_database(database_url)
    _run_slice(client_xlsx_in_data_lake=client_xlsx_in_data_lake)

    csv_path = (
        isolated_data_lake
        / "data"
        / "outputs"
        / "vertical_slice_2023"
        / "vertical_slice_scores.csv"
    )
    assert csv_path.exists(), f"missing output file: {csv_path}"
    with csv_path.open(newline="", encoding="utf-8") as fh:
        reader = csv.DictReader(fh)
        required_columns = {
            "iso3", "country", "leader", "year", "category_key",
            "client_score", "system_proposed_score", "final_score",
            "score_delta_vs_client", "confidence_score", "source_variable",
            "source_year", "source_raw_value", "review_status", "rationale_short",
        }
        assert required_columns.issubset(set(reader.fieldnames or [])), (
            f"missing required columns: {required_columns - set(reader.fieldnames or [])}"
        )
        rows = list(reader)
    assert len(rows) == 6, f"expected 6 score rows, got {len(rows)}"


def test_outputs_write_comparison_csv(
    client_xlsx_in_data_lake: Path,
    seeded_stage2_observations: dict[str, int],
    seeded_score_categories: dict[str, int],
    isolated_data_lake: Path,
    database_url: str,
) -> None:
    """data/outputs/vertical_slice_2023/vertical_slice_comparison.csv is written."""
    init_database(database_url)
    _run_slice(client_xlsx_in_data_lake=client_xlsx_in_data_lake)

    csv_path = (
        isolated_data_lake
        / "data"
        / "outputs"
        / "vertical_slice_2023"
        / "vertical_slice_comparison.csv"
    )
    assert csv_path.exists(), f"missing output file: {csv_path}"
    with csv_path.open(newline="", encoding="utf-8") as fh:
        reader = csv.DictReader(fh)
        required_columns = {
            "validation_status", "missing_reason", "manual_review_required",
        }
        assert required_columns.issubset(set(reader.fieldnames or [])), (
            f"missing required columns: {required_columns - set(reader.fieldnames or [])}"
        )
        rows = list(reader)
    assert len(rows) >= 3, (
        f"expected at least 3 comparison rows (one per scoped country), got {len(rows)}"
    )


def test_outputs_write_summary_md_with_attribution_and_caveat(
    client_xlsx_in_data_lake: Path,
    seeded_stage2_observations: dict[str, int],
    seeded_score_categories: dict[str, int],
    isolated_data_lake: Path,
    database_url: str,
) -> None:
    """vertical_slice_summary.md mentions UNDP HDI and WGI sources and a 'not final' caveat."""
    init_database(database_url)
    _run_slice(client_xlsx_in_data_lake=client_xlsx_in_data_lake)

    md_path = (
        isolated_data_lake
        / "data"
        / "outputs"
        / "vertical_slice_2023"
        / "vertical_slice_summary.md"
    )
    assert md_path.exists(), f"missing output file: {md_path}"
    body = md_path.read_text(encoding="utf-8")
    body_lower = body.lower()

    # Source attribution for UNDP HDI and WGI (Rule #15).
    assert "undp" in body_lower, (
        "summary md should reference UNDP HDI (Rule #15 attribution)"
    )
    assert "wgi" in body_lower, (
        "summary md should reference WGI (Rule #15 attribution)"
    )

    # Explicit 'not final scoring' caveat (architecture doc §1 + §9).
    assert any(
        phrase in body_lower
        for phrase in (
            "provisional",
            "not final",
            "not the final",
            "experimental",
            "preview",
            "not for redistribution",
        )
    ), (
        "summary md must contain a provisional / not-final caveat "
        "(architecture doc §1 and §9)"
    )


# ---------------------------------------------------------------------------
# 8. CLI boundary (architecture doc §10)
# ---------------------------------------------------------------------------


def test_cli_run_vertical_slice_2023_writes_outputs(
    client_xlsx_in_data_lake: Path,
    seeded_stage2_observations: dict[str, int],
    seeded_score_categories: dict[str, int],
    isolated_data_lake: Path,
) -> None:
    """The CLI command invokes the seam and writes the expected output files.

    A stub that just prints a "not implemented yet" message and exits 0
    would NOT write the three expected output files, so this test catches
    missing runtime wiring.
    """
    # Provide a minimal config file under the isolated data lake so the
    # CLI's --config flag resolves. The seam reads target_year from this
    # file (or falls back to the default).
    config_path = isolated_data_lake / "configs" / "test-vertical-slice.yaml"
    config_path.parent.mkdir(parents=True, exist_ok=True)
    config_path.write_text(
        "project:\n  target_year: 2023\n", encoding="utf-8",
    )

    runner = CliRunner()
    result = runner.invoke(
        app,
        [
            "run-vertical-slice-2023",
            "--config", str(config_path),
            "--countries", "MEX,NGA,USA",
            "--categories", "social_wellbeing,integrity",
            "--no-run-adapters",
        ],
    )
    assert result.exit_code == 0, (
        f"CLI exited with code {result.exit_code}; "
        f"the seam did not run end-to-end.\nOutput:\n{result.output}"
    )

    outputs_dir = isolated_data_lake / "data" / "outputs" / "vertical_slice_2023"
    assert outputs_dir.exists(), (
        f"CLI did not create {outputs_dir} -- the seam may be a stub. "
        f"CLI output:\n{result.output}"
    )
    expected = [
        outputs_dir / "vertical_slice_scores.csv",
        outputs_dir / "vertical_slice_comparison.csv",
        outputs_dir / "vertical_slice_summary.md",
    ]
    missing = [p for p in expected if not p.exists()]
    assert not missing, (
        f"CLI did not write expected output files: {missing}. "
        f"CLI output:\n{result.output}"
    )


# ---------------------------------------------------------------------------
# 9. Idempotency + non-slice preservation (architecture doc §4)
# ---------------------------------------------------------------------------


def test_run_twice_keeps_row_counts_stable(
    client_xlsx_in_data_lake: Path,
    seeded_stage2_observations: dict[str, int],
    seeded_score_categories: dict[str, int],
    database_url: str,
) -> None:
    """Running the seam twice keeps ruler_scores/ruler_years/validation counts stable."""
    init_database(database_url)

    _run_slice(client_xlsx_in_data_lake=client_xlsx_in_data_lake)
    with session_scope(database_url) as session:
        first = {
            "scores": session.execute(
                select(func.count(RulerScore.id))
            ).scalar_one(),
            "years": session.execute(
                select(func.count(RulerYear.id))
            ).scalar_one(),
            "validations": session.execute(
                select(func.count(ValidationResult.id))
            ).scalar_one(),
            "spells": session.execute(
                select(func.count(RulerSpell.id))
            ).scalar_one(),
        }

    _run_slice(client_xlsx_in_data_lake=client_xlsx_in_data_lake)
    with session_scope(database_url) as session:
        second = {
            "scores": session.execute(
                select(func.count(RulerScore.id))
            ).scalar_one(),
            "years": session.execute(
                select(func.count(RulerYear.id))
            ).scalar_one(),
            "validations": session.execute(
                select(func.count(ValidationResult.id))
            ).scalar_one(),
            "spells": session.execute(
                select(func.count(RulerSpell.id))
            ).scalar_one(),
        }

    assert first == second, (
        f"row counts changed across reruns: first={first}, second={second}"
    )
    # Specifically: 3 countries x 2 categories = 6 score rows + 6 validations.
    assert first["scores"] == 6
    assert first["years"] == 3
    assert first["validations"] == 6
    assert first["spells"] == 3


def test_non_slice_sentinel_ruler_scores_row_survives_rerun(
    client_xlsx_in_data_lake: Path,
    seeded_stage2_observations: dict[str, int],
    seeded_score_categories: dict[str, int],
    database_url: str,
) -> None:
    """A manually inserted non-slice RulerScore row survives two reruns unchanged.

    The sentinel is for (Canada, 2020, political_freedom) — outside the
    slice's (year, country, category) scope. The architecture doc §4
    guarantees that non-slice rows are not deleted by the idempotency
    sequence.
    """
    init_database(database_url)

    sentinel_marker = "non-slice sentinel -- must survive reruns"

    # Seed the sentinel: a political_freedom ruler_score for a non-scoped
    # country / year. Set a unique rationale so we can verify it survives.
    with session_scope(database_url) as session:
        can = Country(
            iso3="CAN",
            country_name="Canada",
            country_name_normalized=normalize_country_name("Canada"),
        )
        session.add(can)
        session.flush()
        ld = Leader(full_name="Justin Trudeau", normalized_name="justin trudeau")
        session.add(ld)
        session.flush()
        ry = RulerYear(
            leader_id=ld.id,
            country_id=can.id,
            year=2020,
            match_status="external_only",
            review_status="manual_review_required",
        )
        session.add(ry)
        session.flush()
        cat = session.execute(
            select(ScoreCategory).where(
                ScoreCategory.category_key == "political_freedom"
            )
        ).scalar_one_or_none()
        if cat is None:
            cat = ScoreCategory(
                category_key="political_freedom",
                category_name="Political Freedom",
            )
            session.add(cat)
            session.flush()
        sentinel = RulerScore(
            ruler_year_id=ry.id,
            category_id=cat.id,
            client_score=8,
            system_proposed_score=8,
            final_score=None,
            score_delta_vs_client=0,
            confidence_score=80,
            source_agreement="medium",
            human_review_required=False,
            rationale_short=sentinel_marker,
            review_status="manual_review_required",
        )
        session.add(sentinel)
        session.flush()
        sentinel_id = sentinel.id

    # Run the seam twice.
    _run_slice(client_xlsx_in_data_lake=client_xlsx_in_data_lake)
    _run_slice(client_xlsx_in_data_lake=client_xlsx_in_data_lake)

    with session_scope(database_url) as session:
        row = session.execute(
            select(RulerScore).where(RulerScore.id == sentinel_id)
        ).scalar_one_or_none()
    assert row is not None, (
        f"non-slice sentinel RulerScore id={sentinel_id} was deleted by the seam; "
        f"the slice must preserve non-slice rows on idempotent rerun."
    )
    assert row.rationale_short == sentinel_marker, (
        f"sentinel rationale was modified: {row.rationale_short!r}"
    )


def test_non_slice_client_only_sentinel_ruler_year_survives_rerun(
    client_xlsx_in_data_lake: Path,
    seeded_stage2_observations: dict[str, int],
    seeded_score_categories: dict[str, int],
    database_url: str,
) -> None:
    """A non-slice ``client_only`` row for a scoped ISO3/year survives rerun.

    Reviewer blocker 2: the deletion sweep used to select rows by
    ``match_status == "client_only"`` only, which would mistakenly delete
    a legitimate non-slice ``client_only`` row for one of MEX/NGA/USA at
    year 2023. The fix constrains deletion by the slice marker
    (``vertical_slice_2023:`` in ``review_note``), so this sentinel —
    which has ``match_status="client_only"`` but a non-slice ``review_note``
    — must survive two reruns unchanged.
    """
    init_database(database_url)

    sentinel_marker = (
        "non-slice client_only sentinel -- must survive slice reruns"
    )

    with session_scope(database_url) as session:
        # Reuse a scoped country (MEX) and the same target year (2023), but
        # with a NON-slice ``review_note`` so the deletion sweep's marker
        # filter must skip this row.
        mex = session.execute(
            select(Country).where(Country.iso3 == "MEX")
        ).scalar_one_or_none()
        if mex is None:
            mex = Country(
                iso3="MEX",
                country_name="Mexico",
                country_name_normalized=normalize_country_name("Mexico"),
            )
            session.add(mex)
            session.flush()
        ld = session.execute(
            select(Leader).where(Leader.normalized_name == "non-slice sentinel")
        ).scalar_one_or_none()
        if ld is None:
            ld = Leader(
                full_name="Non-slice Sentinel",
                normalized_name="non-slice sentinel",
                notes="non-slice leader for client_only sentinel",
            )
            session.add(ld)
            session.flush()
        ry = RulerYear(
            leader_id=ld.id,
            country_id=mex.id,
            year=TARGET_YEAR,
            match_status="client_only",
            review_status="manual_review_required",
            review_note=sentinel_marker,
            client_matrix_leader_name="Non-slice Sentinel",
            system_selected_leader_name="Non-slice Sentinel",
        )
        session.add(ry)
        session.flush()
        sentinel_ry_id = ry.id

    _run_slice(client_xlsx_in_data_lake=client_xlsx_in_data_lake)
    _run_slice(client_xlsx_in_data_lake=client_xlsx_in_data_lake)

    with session_scope(database_url) as session:
        row = session.execute(
            select(RulerYear).where(RulerYear.id == sentinel_ry_id)
        ).scalar_one_or_none()
    assert row is not None, (
        f"non-slice client_only sentinel RulerYear id={sentinel_ry_id} was "
        f"deleted by the seam; the slice must constrain deletion by the "
        f"slice marker, not just by match_status='client_only'."
    )
    assert row.review_note == sentinel_marker, (
        f"non-slice sentinel review_note was modified: {row.review_note!r}"
    )
    assert row.match_status == "client_only", (
        f"non-slice sentinel match_status was modified: {row.match_status!r}"
    )


# ---------------------------------------------------------------------------
# 10. Multi-year source-only time-series output (extending the 2023 slice).
#
# The user wants to keep the existing 2023 client-comparison behavior
# (ruler_years / ruler_scores / validation_results) intact, and add a
# separate, source-only time-series output for arbitrary years
# (e.g. 2020, 2021, 2022, 2023). The DB stays 2023-scope; the new file
# is the only place multi-year system_proposed_score rows live.
# ---------------------------------------------------------------------------


def _timeseries_path(isolated_data_lake: Path) -> Path:
    """Return the canonical time-series CSV path under the isolated data lake."""
    return (
        isolated_data_lake
        / "data"
        / "outputs"
        / "vertical_slice_2023"
        / TIMESERIES_OUTPUT_CSV
    )


def test_timeseries_csv_written_when_years_provided(
    client_xlsx_in_data_lake: Path,
    seeded_multi_year_observations: dict[str, dict[int, dict[str, float]]],
    seeded_score_categories: dict[str, int],
    isolated_data_lake: Path,
    database_url: str,
) -> None:
    """``vertical_slice_timeseries.csv`` is written when ``years=`` is provided.

    A run that ignores the ``years`` kwarg (or that doesn't write the new
    file) is caught by the explicit file-existence assertion. The fixture
    seeds UNDP HDI and WGI for 2020/2021/2022 so the seam has real inputs
    to surface in the time-series.
    """
    init_database(database_url)
    _run_slice_with_years(
        client_xlsx_in_data_lake=client_xlsx_in_data_lake,
        years=(2020, 2021, 2022, 2023),
    )

    ts_path = _timeseries_path(isolated_data_lake)
    assert ts_path.exists(), (
        f"missing time-series output: {ts_path}. The multi-year feature "
        "must write data/outputs/vertical_slice_2023/vertical_slice_timeseries.csv "
        "when years= is provided."
    )


def test_timeseries_has_rows_for_all_years_countries_categories(
    client_xlsx_in_data_lake: Path,
    seeded_multi_year_observations: dict[str, dict[int, dict[str, float]]],
    seeded_score_categories: dict[str, int],
    isolated_data_lake: Path,
    database_url: str,
) -> None:
    """The time-series CSV has the required columns and one row per
    ``(year, country, category)`` tuple for which an observation exists.

    With the seeded fixture (UNDP HDI + WGI for MEX/NGA/USA at 2020/2021/2022
    and the documented 2022 -> 2023 proxy), the expected row count is::

        4 years * 3 countries * 2 categories == 24 rows.
    """
    init_database(database_url)
    _run_slice_with_years(
        client_xlsx_in_data_lake=client_xlsx_in_data_lake,
        years=(2020, 2021, 2022, 2023),
    )

    ts_path = _timeseries_path(isolated_data_lake)
    assert ts_path.exists(), f"missing time-series output: {ts_path}"

    with ts_path.open(newline="", encoding="utf-8") as fh:
        reader = csv.DictReader(fh)
        rows = list(reader)
        fieldnames = reader.fieldnames or []

    expected_columns = [
        "iso3", "country", "year", "category_key",
        "system_proposed_score", "source_variable",
        "source_year", "source_raw_value", "source_kind",
        "confidence_score", "note",
    ]
    assert fieldnames == expected_columns, (
        "time-series CSV must remain source-only. It must not grow "
        "client-comparison or leader-identity columns such as client_score, "
        "score_delta_vs_client, or leader. "
        f"Expected {expected_columns}, got {fieldnames}"
    )

    expected_count = (
        4 * len(DEFAULT_ISO3_SCOPE) * len(DEFAULT_CATEGORIES)
    )
    assert len(rows) == expected_count, (
        f"expected {expected_count} time-series rows "
        f"(4 years x {len(DEFAULT_ISO3_SCOPE)} countries x "
        f"{len(DEFAULT_CATEGORIES)} categories), got {len(rows)}"
    )


def test_timeseries_2023_rows_use_2022_proxy(
    client_xlsx_in_data_lake: Path,
    seeded_multi_year_observations: dict[str, dict[int, dict[str, float]]],
    seeded_score_categories: dict[str, int],
    isolated_data_lake: Path,
    database_url: str,
) -> None:
    """2023 rows carry ``source_year=2022`` and ``source_kind='proxy'``.

    No 2023 observation is seeded, so the seam must fall back to the
    documented 2022 proxy for both UNDP HDI and WGI (architecture §8)
    and label the row accordingly.
    """
    init_database(database_url)
    _run_slice_with_years(
        client_xlsx_in_data_lake=client_xlsx_in_data_lake,
        years=(2020, 2021, 2022, 2023),
    )

    ts_path = _timeseries_path(isolated_data_lake)
    assert ts_path.exists(), f"missing time-series output: {ts_path}"

    with ts_path.open(newline="", encoding="utf-8") as fh:
        rows = list(csv.DictReader(fh))

    rows_2023 = [r for r in rows if r.get("year") == "2023"]
    expected_2023_count = len(DEFAULT_ISO3_SCOPE) * len(DEFAULT_CATEGORIES)
    assert len(rows_2023) == expected_2023_count, (
        f"expected {expected_2023_count} rows for year=2023 "
        f"({len(DEFAULT_ISO3_SCOPE)} countries x "
        f"{len(DEFAULT_CATEGORIES)} categories), got {len(rows_2023)}"
    )
    for r in rows_2023:
        assert r["source_year"] == "2022", (
            f"2023 row {r!r} must have source_year='2022' "
            f"(the documented 2022 proxy for 2023), got {r['source_year']!r}"
        )
        assert r["source_kind"] == "proxy", (
            f"2023 row {r!r} must have source_kind='proxy', "
            f"got {r['source_kind']!r}"
        )


def test_timeseries_pre_2023_rows_are_direct(
    client_xlsx_in_data_lake: Path,
    seeded_multi_year_observations: dict[str, dict[int, dict[str, float]]],
    seeded_score_categories: dict[str, int],
    isolated_data_lake: Path,
    database_url: str,
) -> None:
    """2020/2021/2022 rows use the exact source year and ``source_kind='direct'``.

    The fixture seeds a direct observation for each (year, country,
    variable) pair, so the seam must surface those exact rows without
    falling back to a proxy.
    """
    init_database(database_url)
    _run_slice_with_years(
        client_xlsx_in_data_lake=client_xlsx_in_data_lake,
        years=(2020, 2021, 2022, 2023),
    )

    ts_path = _timeseries_path(isolated_data_lake)
    assert ts_path.exists(), f"missing time-series output: {ts_path}"

    with ts_path.open(newline="", encoding="utf-8") as fh:
        rows = list(csv.DictReader(fh))

    pre_2023_rows = [r for r in rows if int(r["year"]) < 2023]
    expected_pre_count = (
        3 * len(DEFAULT_ISO3_SCOPE) * len(DEFAULT_CATEGORIES)
    )
    assert len(pre_2023_rows) == expected_pre_count, (
        f"expected {expected_pre_count} pre-2023 rows "
        f"(3 years x {len(DEFAULT_ISO3_SCOPE)} countries x "
        f"{len(DEFAULT_CATEGORIES)} categories), got {len(pre_2023_rows)}"
    )
    for r in pre_2023_rows:
        assert r["source_year"] == r["year"], (
            f"pre-2023 row {r!r} must have source_year == year, "
            f"got source_year={r['source_year']!r} vs year={r['year']!r}"
        )
        assert r["source_kind"] == "direct", (
            f"pre-2023 row {r!r} must have source_kind='direct', "
            f"got {r['source_kind']!r}"
        )


def test_non_2023_years_do_not_create_db_rows(
    client_xlsx_in_data_lake: Path,
    seeded_multi_year_observations: dict[str, dict[int, dict[str, float]]],
    seeded_score_categories: dict[str, int],
    database_url: str,
) -> None:
    """Passing ``years=`` does NOT widen the DB write scope.

    The client-comparison invariant stays 2023-only: the multi-year
    feature is source-only (CSV output), not DB writes. Concretely:

    - ``ruler_years`` rows remain the three 2023 client_only rows;
    - ``ruler_scores`` rows remain the six 2023 score rows;
    - ``validation_results`` rows remain the six 2023 validation rows.

    A run that accidentally upserts per-year ``ruler_years`` /
    ``ruler_scores`` / ``validation_results`` rows is caught by the
    strict counts below.
    """
    init_database(database_url)
    _run_slice_with_years(
        client_xlsx_in_data_lake=client_xlsx_in_data_lake,
        years=(2020, 2021, 2022, 2023),
    )

    with session_scope(database_url) as session:
        ruler_years_count = session.execute(
            select(func.count(RulerYear.id))
        ).scalar_one()
        ruler_scores_count = session.execute(
            select(func.count(RulerScore.id))
        ).scalar_one()
        validation_count = session.execute(
            select(func.count(ValidationResult.id))
        ).scalar_one()
        years_in_db = sorted(
            session.execute(select(RulerYear.year)).scalars().all()
        )

    assert ruler_years_count == 3, (
        f"expected 3 ruler_years (2023 client_only), got {ruler_years_count}. "
        "The multi-year feature must not widen the DB write scope."
    )
    assert ruler_scores_count == 6, (
        f"expected 6 ruler_scores (3 countries x 2 categories at 2023), "
        f"got {ruler_scores_count}. The multi-year feature must not widen "
        "the DB write scope."
    )
    assert validation_count == 6, (
        f"expected 6 validation_results (one per 2023 score), got "
        f"{validation_count}. The multi-year feature must not widen the "
        "DB write scope."
    )
    # The slice writes one ruler_year per scoped country (3 rows here),
    # each at year=2023. Set-equality asserts the regression we care about
    # ("no non-2023 years leak into ruler_years") without conflating it
    # with the unrelated row-count assertion above.
    assert set(years_in_db) == {TARGET_YEAR}, (
        f"ruler_years must be 2023-only, got year(s)={years_in_db!r}"
    )


def test_cli_accepts_years_flag_and_writes_timeseries(
    client_xlsx_in_data_lake: Path,
    seeded_multi_year_observations: dict[str, dict[int, dict[str, float]]],
    seeded_score_categories: dict[str, int],
    isolated_data_lake: Path,
) -> None:
    """The ``run-vertical-slice-2023`` CLI accepts ``--years`` and writes the CSV.

    The CLI must surface the new ``--years`` flag, parse the comma-separated
    list of integers, pass them through to the seam, and the seam must
    then write ``vertical_slice_timeseries.csv``. A stub CLI that rejects
    the flag (exit code 2) or a seam that ignores it (no file written) is
    caught by the explicit exit-code + file-existence assertions.
    """
    config_path = isolated_data_lake / "configs" / "test-vertical-slice.yaml"
    config_path.parent.mkdir(parents=True, exist_ok=True)
    config_path.write_text(
        "project:\n  target_year: 2023\n", encoding="utf-8",
    )

    runner = CliRunner()
    result = runner.invoke(
        app,
        [
            "run-vertical-slice-2023",
            "--config", str(config_path),
            "--countries", "MEX,NGA,USA",
            "--categories", "social_wellbeing,integrity",
            "--years", "2020,2021,2022,2023",
            "--no-run-adapters",
        ],
    )
    assert result.exit_code == 0, (
        f"CLI exited with code {result.exit_code}; the --years flag may "
        f"be missing or rejected by the seam.\nOutput:\n{result.output}"
    )

    ts_path = _timeseries_path(isolated_data_lake)
    assert ts_path.exists(), (
        f"CLI did not write {ts_path}. CLI output:\n{result.output}"
    )


def test_summary_mentions_timeseries_years_and_2023_only_caveat(
    client_xlsx_in_data_lake: Path,
    seeded_multi_year_observations: dict[str, dict[int, dict[str, float]]],
    seeded_score_categories: dict[str, int],
    isolated_data_lake: Path,
    database_url: str,
) -> None:
    """The summary md names every requested year and states that client
    comparison is 2023-only.

    The summary is the human-facing artifact, so the user must be able to
    read directly from it which years the time-series covers and that the
    client comparison is still scoped to 2023.
    """
    init_database(database_url)
    _run_slice_with_years(
        client_xlsx_in_data_lake=client_xlsx_in_data_lake,
        years=(2020, 2021, 2022, 2023),
    )

    md_path = (
        isolated_data_lake
        / "data"
        / "outputs"
        / "vertical_slice_2023"
        / "vertical_slice_summary.md"
    )
    assert md_path.exists(), f"missing summary markdown: {md_path}"
    body = md_path.read_text(encoding="utf-8")

    for year in (2020, 2021, 2022, 2023):
        assert str(year) in body, (
            f"summary md must mention year {year} (the requested time-series "
            f"year), got body:\n{body}"
        )

    body_lower = body.lower()
    assert any(
        phrase in body_lower
        for phrase in (
            "client comparison is 2023-only",
            "client comparison 2023 only",
            "client comparison is limited to 2023",
            "client comparison limited to 2023",
            "2023-only",
        )
    ), (
        "summary md must contain a caveat that the client comparison is "
        "2023-only (the multi-year feature is source-only). "
        f"Got body:\n{body}"
    )
