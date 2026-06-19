#!/usr/bin/env python3
"""Build tests/fixtures/cirights/sample.xlsx by slicing the real CIRIGHTS xlsx.

This script is idempotent: re-running overwrites the fixture with the
same sliced data. It is committed so the fixture is reproducible
without access to the original download.

Fixture shape (5 countries x 2 years, real-format CIRIGHTS xlsx):

  - Mexico 2021        (8 physint + 11 repression + 4 civpol; well-coded)
  - Mexico 2022        (the proxy data year for 2023)
  - Norway 2021        (high-rights country; physint=8, civpol=6)
  - Norway 2022        (high-rights country; physint=8, civpol=6)
  - China 2021         (low-rights country; physint=0)
  - China 2022         (low-rights country; physint=0)
  - Brazil 2021        (mid-rights country; physint=4)
  - Brazil 2022        (mid-rights country; physint=4)
  - United States of America 2021  (high-rights country; physint=8)
  - United States of America 2022  (high-rights country; physint=8)

The fixture includes a "United States of America" entry (the real
xlsx uses that display name, not "United States") to exercise the
``safe_country_token`` URL-safe substitution: the
``source_row_reference`` should be
``cirights:United_States_of_America:2022:...``.

The fixture exercises:

  - The 7 catalog indicator columns (Physical Integrity Rights Index,
    Repression Index, Civil and Political Rights Index, Disappearances,
    Extrajudicial Killings, Political Imprisonment, Torture).
  - The 2 identity columns (country, year).
  - 2-year coverage (2021 + 2022) to support the
    ``year_window=(2021, 2022)`` test assertion.
  - All 7 indicators populated for all 10 country-year rows (no
    missing cells in the fixture; the missing-cell tests are
    covered by an in-memory DataFrame injected directly into
    :func:`read_cirights_from_dataframe`).

All values are real CIRIGHTS values sliced from the live
``data/raw/cirights/cirights_v3.12.10.24.xlsx`` with openpyxl -- no
invented data.
"""

from __future__ import annotations

import sys
from pathlib import Path

# openpyxl is a project dependency (same as WGI / SIPRI milex / PTS)
import openpyxl


def main() -> None:
    # Resolve the project root relative to this file: tests/fixtures/cirights/
    # -> tests/fixtures/ -> tests/ -> <project_root>/.
    project_root = Path(__file__).resolve().parents[3]
    src_path = (
        project_root
        / "data"
        / "raw"
        / "cirights"
        / "cirights_v3.12.10.24.xlsx"
    )
    dst_path = Path(__file__).resolve().parent / "sample.xlsx"

    if not src_path.exists():
        print(
            f"ERROR: source xlsx not found at {src_path}. Place "
            "cirights_v3.12.10.24.xlsx under data/raw/cirights/ first.",
            file=sys.stderr,
        )
        sys.exit(1)

    print(f"Reading {src_path} ...")
    wb_src = openpyxl.load_workbook(
        str(src_path), read_only=True, data_only=True,
    )
    ws_src = wb_src["Sheet1"]

    # (country, year) pairs we want
    wanted: set[tuple[str, int]] = {
        ("Mexico", 2021),
        ("Mexico", 2022),
        ("Norway", 2021),
        ("Norway", 2022),
        ("China", 2021),
        ("China", 2022),
        ("Brazil", 2021),
        ("Brazil", 2022),
        ("United States of America", 2021),
        ("United States of America", 2022),
    }

    all_rows: list[tuple[object, ...]] = []
    header_row: tuple[object, ...] | None = None
    selected_indices: set[int] = set()
    expected_count = len(wanted)

    for i, row in enumerate(ws_src.iter_rows(values_only=True)):
        row_tuple = tuple(row)  # type: ignore[arg-type]
        if i == 0:
            header_row = row_tuple
            all_rows.append(row_tuple)
            continue
        country = row[0]
        year = row[1]
        if (country, year) in wanted:
            selected_indices.add(i)
            all_rows.append(row_tuple)
            wanted.discard((country, year))
            print(
                f"  Selected row {i + 1}: {country} {year} "
                f"(physint={row[9]}, disap={row[16]})"
            )
        if not wanted:
            break

    if header_row is None:
        print("ERROR: could not read header row from source xlsx", file=sys.stderr)
        sys.exit(1)

    if selected_indices:
        print(
            f"\nTotal selected data rows: {len(selected_indices)} "
            f"(expected {expected_count})"
        )
    else:
        print(
            "WARNING: no rows selected — fixture will contain only header",
            file=sys.stderr,
        )
        sys.exit(1)

    # Write the new workbook. We preserve the full 50-column header
    # (not just the 9 columns the catalog needs) so the
    # ``read_xlsx_to_wide_dataframe`` function can find every column
    # it expects and the drift-guard schema validation passes.
    wb_dst = openpyxl.Workbook()
    ws_dst = wb_dst.active
    ws_dst.title = "Sheet1"

    # Copy header
    ws_dst.append(list(header_row))  # type: ignore[arg-type]

    # Copy selected data rows
    for row in all_rows[1:]:  # skip header (index 0)
        ws_dst.append(list(row))

    wb_dst.save(str(dst_path))
    print(f"\nWritten: {dst_path}")
    print(f"  Rows: {len(all_rows) - 1} data rows + 1 header")
    print(f"  Columns: {len(header_row)}")

    # Verify what we wrote
    wb_verify = openpyxl.load_workbook(str(dst_path), data_only=True)
    ws_verify = wb_verify.active
    verify_rows = list(ws_verify.iter_rows(values_only=True))
    print(f"\nVerification — {len(verify_rows)} total rows in fixture:")
    for j, vr in enumerate(verify_rows):
        if j == 0:
            print(f"  Header ({len(vr)} cols): {list(vr)[:10]} ...")
        else:
            print(
                f"  Row {j}: {vr[0]} {vr[1]} "
                f"physint={vr[9]} "
                f"disap={vr[16]} kill={vr[17]} "
                f"polpris={vr[18]} tort={vr[19]}"
            )


if __name__ == "__main__":
    main()
