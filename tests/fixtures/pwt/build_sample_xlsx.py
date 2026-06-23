"""Build ``tests/fixtures/pwt/sample.xlsx`` for the Phase B PWT Stage 2 tests.

This script writes a small real-format Penn World Table 10.01 xlsx
fixture that the Stage 2 adapter can read end-to-end. The fixture
shape mirrors the live ``data/raw/pwt/pwt1001.xlsx`` workbook
at the level the Phase B tests exercise (sheets + Data sheet
header + 6 data rows); the cell values are NOT real PWT values
(see "Public-domain verification" below).

Fixture layout:

- Sheets: ``Info``, ``Legend``, ``Data`` (three sheets, matching
  the live file's sheet names so the reader's sheet-selection
  test has something to verify it skipped).
- ``Data`` sheet header row: the 4 identity columns
  (``countrycode``, ``country``, ``currency_unit``, ``year``) +
  the 11 catalog numeric columns (``rgdpe``, ``rgdpo``, ``pop``,
  ``emp``, ``avh``, ``hc``, ``ccon``, ``cda``, ``ctfp``,
  ``rkna``, ``rtfpna``) = 15 columns total.
- Body: 3 countries (``USA``, ``MEX``, ``SWE``) x 2 years
  (2018, 2019).

Missing-cell emission semantics (documented per Phase B
review): the Stage 2 transform drops a cell -- does NOT emit
an observation row -- when the raw cell is blank, whitespace,
a sentinel (``NA`` / ``N/A`` / ``NaN`` / ``null``), or any
non-numeric string. Numeric-like strings (e.g. ``"1234.5"``)
are coerced to ``float``.

Expected per-row non-blank catalog cell counts (the tests
assert these explicitly so the totals are derivable):

- USA 2018: 2 cells  (rgdpe, pop)
- USA 2019: 6 cells  (rgdpe, rgdpo, pop, emp, avh, hc)
- MEX 2018: 0 cells  (all catalog cells blank)
- MEX 2019: 6 cells  (rgdpe, rgdpo, pop, emp, avh, hc)
- SWE 2018: 0 cells  (all catalog cells blank)
- SWE 2019: 3 cells  (rgdpe, rgdpo, pop)

Per-year totals:

- 2018: 2 cells  (USA only)
- 2019: 15 cells (USA 6 + MEX 6 + SWE 3)

All-years total: 17 cells.

PWT 10.01 covers 1950-2019; the fixture deliberately uses
2018 + 2019 so the Phase B "out-of-coverage" tests can assert
that ``year=2023`` produces zero observations.

Why a self-contained fixture builder (not a slice-from-bundle):

- The fixture must be buildable in CI without the 6.5 MB source
  xlsx being present. Embedding the values inline removes the
  "fetch the upstream xlsx" dependency from the test build path.
- The fixture is small (3 countries x 2 years x 15 columns) and
  exercises the adapter's reader, transform, and locator logic
  without leaking real PWT values into the test tree.
- Re-running the script is idempotent and never emits print()
  output (CI / pre-commit hooks would otherwise see spurious
  stdout).

Fixture exercises:

- The 4 identity + 11 catalog columns the reader must validate
  + carry into the wide Data-sheet DataFrame.
- The locator pattern ``pwt:Data:<countrycode>:<year>:<raw_column>``
  (e.g. ``pwt:Data:USA:2019:rgdpe``) at the transform layer.
- The ``Data`` sheet as the only data source; ``Info`` and
  ``Legend`` are staged but the Stage 2 reader never opens them.

Public-domain verification: the xlsx is a fake fixture for tests;
no real PWT values are embedded. The numeric magnitudes match
plausible PWT 10.01 shapes (USA 2019 real GDP ~ 20 trillion 2017
USD, etc.) but the actual numbers are placeholder values chosen
to exercise numeric coercion rather than represent real PWT 10.01
measurements. The fixture shape (sheets, headers, ISO3s, year
range) mirrors the live ``pwt1001.xlsx`` workbook so a developer
can review the fixture against the live file's header row + sheet
layout.
"""

from __future__ import annotations

import hashlib
from pathlib import Path

import openpyxl

# Three countries x two years. ``(countrycode, country,
# currency_unit, year, rgdpe, rgdpo, pop, emp, avh, hc, ccon, cda,
# ctfp, rkna, rtfpna)``. Most numeric cells are blank (None) so the
# Phase B tests can pin which observations are emitted; the cells
# that ARE filled match plausible PWT 10.01 magnitudes (USA 2019
# real GDP ~ 20 trillion 2017 USD; MEX 2019 real GDP ~ 2.2
# trillion 2017 USD; SWE 2019 ~ 0.5 trillion 2017 USD). They are NOT
# real PWT 10.01 values -- they are placeholder magnitudes chosen
# to exercise numeric coercion without leaking real data into the
# test tree.
#
# Non-blank catalog cell counts (the per-row totals drive the
# expected observation counts; the test file asserts these
# explicitly so the counts are derivable):
#   USA 2018: 2 (rgdpe, pop)
#   USA 2019: 6 (rgdpe, rgdpo, pop, emp, avh, hc)
#   MEX 2018: 0 (all blank)
#   MEX 2019: 6 (rgdpe, rgdpo, pop, emp, avh, hc)
#   SWE 2018: 0 (all blank)
#   SWE 2019: 3 (rgdpe, rgdpo, pop)
# Per-year totals: 2018 -> 2, 2019 -> 15, all-years -> 17.
SAMPLE_ROW_TYPE = tuple[
    str, str, str, int,
    float | None, float | None, float | None, float | None,
    float | None, float | None, float | None, float | None,
    float | None, float | None, float | None,
]

# USA 2018: 2 non-blank cells (rgdpe, pop).
_USA_2018: SAMPLE_ROW_TYPE = (
    "USA", "United States", "US Dollar", 2018,
    19.5e12, None, 327.2e6, None,
    None, None, None, None,
    None, None, None,
)
# USA 2019: 6 non-blank cells (rgdpe, rgdpo, pop, emp, avh, hc).
_USA_2019: SAMPLE_ROW_TYPE = (
    "USA", "United States", "US Dollar", 2019,
    20.0e12, 19.5e12, 328.3e6, 158.5e6,
    1785.0, 3.65, None, None,
    None, None, None,
)
# MEX 2018: 0 non-blank cells (all catalog cells blank).
_MEX_2018: SAMPLE_ROW_TYPE = (
    "MEX", "Mexico", "Mexican Peso", 2018,
    None, None, None, None,
    None, None, None, None,
    None, None, None,
)
# MEX 2019: 6 non-blank cells (rgdpe, rgdpo, pop, emp, avh, hc).
_MEX_2019: SAMPLE_ROW_TYPE = (
    "MEX", "Mexico", "Mexican Peso", 2019,
    2.2e12, 2.1e12, 127.6e6, 55.7e6,
    2200.0, 2.55, None, None,
    None, None, None,
)
# SWE 2018: 0 non-blank cells (all catalog cells blank).
_SWE_2018: SAMPLE_ROW_TYPE = (
    "SWE", "Sweden", "Swedish Krona", 2018,
    None, None, None, None,
    None, None, None, None,
    None, None, None,
)
# SWE 2019: 3 non-blank cells (rgdpe, rgdpo, pop).
_SWE_2019: SAMPLE_ROW_TYPE = (
    "SWE", "Sweden", "Swedish Krona", 2019,
    5.0e11, 4.8e11, 10.3e6, None,
    None, None, None, None,
    None, None, None,
)

SAMPLE_ROWS: tuple[SAMPLE_ROW_TYPE, ...] = (
    _USA_2018,
    _USA_2019,
    _MEX_2018,
    _MEX_2019,
    _SWE_2018,
    _SWE_2019,
)

# The 4 identity + 11 catalog columns the Stage 2 reader validates
# + emits into the wide Data-sheet-shaped DataFrame (per
# ``docs/sources/ingestion-plan.md`` PWT section, verified against
# the live ``pwt1001.xlsx`` ``Data`` sheet on 2026-06-22). The
# 11 catalog columns are the canonical numeric indicators the
# Stage 2 plan calls out; the live workbook carries additional
# derived / share / price-level columns, but those are not in the
# Phase B catalog and the reader must not over-read them.
FIXTURE_COLUMNS: tuple[str, ...] = (
    "countrycode",
    "country",
    "currency_unit",
    "year",
    "rgdpe",
    "rgdpo",
    "pop",
    "emp",
    "avh",
    "hc",
    "ccon",
    "cda",
    "ctfp",
    "rkna",
    "rtfpna",
)

# Sheet names mirror the live ``pwt1001.xlsx`` exactly.
PWT_INFO_SHEET_NAME: str = "Info"
PWT_LEGEND_SHEET_NAME: str = "Legend"
PWT_DATA_SHEET_NAME: str = "Data"


def build_sample_xlsx(out_xlsx: Path) -> Path:
    """Write ``tests/fixtures/pwt/sample.xlsx``.

    Creates a fresh xlsx with the three canonical sheets (``Info``,
    ``Legend``, ``Data``), the 15 canonical columns in the ``Data``
    header row, and one data row per ``SAMPLE_ROWS`` entry. The
    function is idempotent: an existing file at ``out_xlsx`` is
    overwritten. No print output is emitted (CI / pre-commit
    friendly).
    """
    out_xlsx = Path(out_xlsx)
    out_xlsx.parent.mkdir(parents=True, exist_ok=True)
    if out_xlsx.exists():
        out_xlsx.unlink()

    wb = openpyxl.Workbook()
    default = wb.active
    if default is not None:
        wb.remove(default)

    # The Info and Legend sheets are present so the workbook matches
    # the live ``pwt1001.xlsx`` layout. The Stage 2 reader must not
    # open either of them; the Phase B tests assert that behavior.
    wb.create_sheet(title=PWT_INFO_SHEET_NAME)
    wb.create_sheet(title=PWT_LEGEND_SHEET_NAME)
    data_ws = wb.create_sheet(title=PWT_DATA_SHEET_NAME)

    data_ws.append(list(FIXTURE_COLUMNS))
    for row in SAMPLE_ROWS:
        data_ws.append(list(row))

    wb.save(out_xlsx)
    return out_xlsx


def sha256_of(path: Path) -> str:
    """Return the lowercase hex SHA-256 of ``path``.

    Used by the Phase B metadata validation test to build the
    ``checksum_sha256`` field for the fixture's ``metadata.json``.
    """
    digest = hashlib.sha256()
    with open(path, "rb") as fh:
        for chunk in iter(lambda: fh.read(65536), b""):
            digest.update(chunk)
    return digest.hexdigest()


def main() -> Path:
    """Build the fixture xlsx at ``tests/fixtures/pwt/sample.xlsx``."""
    fixture_path = (
        Path(__file__).resolve().parent / "sample.xlsx"
    )
    return build_sample_xlsx(fixture_path)


if __name__ == "__main__":
    main()
    # The build script returns the fixture path. CI / pre-commit
    # do not consume stdout; the explicit no-print contract is
    # enforced by the absence of any print() calls.
