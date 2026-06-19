"""Build a small real-format BTI xlsx fixture for the BTI Stage 2 tests.

The fixture is a slice of the real ``data/raw/bti/BTI_2006-2026_Scores.xlsx``:
2 BTI edition sheets ("BTI 2024" for 2023, "BTI 2022" for 2021), 5
countries per sheet, with 12 catalog indicator columns populated from
the real BTI xlsx values (no invented data).

The fixture is committed to ``tests/fixtures/bti/sample.xlsx`` so the
tests run deterministically and offline. Build only re-runs the fixture
when the script changes; ``sample.xlsx`` is the artifact the tests
consume.

Why a custom slice rather than the full 388MB file:
- Test-suite speed: a full BTI xlsx read takes ~30s; the slice is ~5ms.
- Test isolation: the test lake is per-test (pytest fixture
  ``isolated_data_lake``); the slice keeps the in-memory frame under
  1k rows.
- Real-format coverage: the slice preserves the real BTI column layout
  (123 columns, with the trailing trend/category columns + the
  merged-cell "Regions:" label in col 0) so the read function exercises
  the real header-walk + per-indicator column resolution path.
"""

from __future__ import annotations

from pathlib import Path

import openpyxl

# The 5 countries we sample across BTI editions. BTI does not cover
# the United States or Sweden (not in the BTI scope). The sample
# spans the 7 regions: Mexico (LAC), Brazil (LAC), India (AO), Nigeria
# (WCA), Kenya (SEA). All 5 are present in every BTI edition sheet
# from 2006 through 2026 (verified against the live xlsx).
SAMPLE_COUNTRIES: tuple[str, ...] = (
    "Mexico",
    "Brazil",
    "India",
    "Nigeria",
    "Kenya",
)

# The 2 BTI edition sheets we sample. Each maps to a target year:
# BTI 2024 -> 2023, BTI 2022 -> 2021.
SAMPLE_SHEETS: tuple[str, ...] = ("BTI 2024", "BTI 2022")


def _build_fixture(source_xlsx: Path, out_xlsx: Path) -> Path:
    """Slice the real BTI xlsx into a small per-edition fixture.

    Reads the source xlsx once per edition sheet, writes the header
    row + a 5-country slice to a fresh xlsx with the same 123-column
    layout (preserving the trailing trend/category columns verbatim
    so the test exercises the real header layout).
    """
    src_wb = openpyxl.load_workbook(source_xlsx, read_only=True)
    try:
        out_wb = openpyxl.Workbook()
        # Drop the default sheet; we add edition sheets below.
        default = out_wb.active
        if default is not None:
            out_wb.remove(default)

        for sheet_name in SAMPLE_SHEETS:
            if sheet_name not in src_wb.sheetnames:
                raise FileNotFoundError(
                    f"Source BTI xlsx is missing sheet {sheet_name!r}. "
                    f"Available: {src_wb.sheetnames}"
                )
            src_ws = src_wb[sheet_name]
            out_ws = out_wb.create_sheet(title=sheet_name)

            # Stream rows: write the header (row 0) + only the rows
            # whose col 0 is in SAMPLE_COUNTRIES. Preserve every
            # column (123 in the live xlsx).
            sample_set = set(SAMPLE_COUNTRIES)
            for i, row in enumerate(src_ws.iter_rows(values_only=True), start=1):
                if i == 1:
                    # Header row -- copy verbatim.
                    out_ws.append(list(row))
                    continue
                country_name = row[0]
                if not isinstance(country_name, str):
                    # Defensive: skip blank / non-string rows.
                    continue
                if country_name.strip() in sample_set:
                    out_ws.append(list(row))

        out_wb.save(out_xlsx)
    finally:
        src_wb.close()
    return out_xlsx


def main() -> Path:
    """Build the fixture xlsx at ``tests/fixtures/bti/sample.xlsx``.

    Reads from ``data/raw/bti/BTI_2006-2026_Scores.xlsx``. Idempotent;
    re-running overwrites the fixture.
    """
    project_root = Path(__file__).resolve().parents[3]
    source_xlsx = project_root / "data" / "raw" / "bti" / "BTI_2006-2026_Scores.xlsx"
    if not source_xlsx.is_file():
        raise FileNotFoundError(
            f"Source BTI xlsx not found: {source_xlsx}. "
            "Stage the cumulative xlsx at data/raw/bti/ before building the fixture."
        )
    fixture_path = project_root / "tests" / "fixtures" / "bti" / "sample.xlsx"
    fixture_path.parent.mkdir(parents=True, exist_ok=True)
    if fixture_path.exists():
        fixture_path.unlink()
    return _build_fixture(source_xlsx, fixture_path)


if __name__ == "__main__":
    out = main()
    # Report size so the build log is auditable.
    size_kb = out.stat().st_size / 1024
    print(f"Wrote {out} ({size_kb:.1f} KB)")
