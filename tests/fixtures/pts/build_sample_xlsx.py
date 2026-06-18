#!/usr/bin/env python3
"""Build tests/fixtures/pts/sample.xlsx by slicing the real PTS-2025.xlsx.

This script is idempotent: re-running overwrites the fixture with the
same sliced data. It is committed so the fixture is reproducible without
access to the original download.

Fixture shape (5 rows):
  - Afghanistan 2022  (case 1: all 3 indicators valid)
  - Afghanistan 2023  (case 1: all 3 indicators valid)
  - United States 2022 (case 3: PTS_S='NA', NA_Status_S=88 → dropped)
  - United States 2023 (case 3: PTS_S='NA', NA_Status_S=88 → dropped)
  - Andorra 2022      (case 2: PTS_A=1, NA_Status_A=88 → dropped;
                       PTS_H='NA', NA_Status_H=88 → dropped;
                       PTS_S=1, NA_Status_S=0 → valid)

The Bahamas 2017 inconsistency case (PTS_A='NA', NA_Status_A=0) is not
included in the file fixture because no Bahamas row falls in the 2022-2023
window. It is tested via an in-memory DataFrame injected directly into
read_pts_from_dataframe in test_ingest_pts.py.
"""

from __future__ import annotations

import sys
from pathlib import Path

# openpyxl is a project dependency (same as WGI / SIPRI milex)
import openpyxl


def main() -> None:
    src_path = Path("data/raw/political_terror_scale/PTS-2025.xlsx")
    if not src_path.exists():
        src_path = (
            Path(__file__).resolve().parents[2]
            / "data"
            / "raw"
            / "political_terror_scale"
            / "PTS-2025.xlsx"
        )
    dst_path = Path(__file__).resolve().parent / "sample.xlsx"

    print(f"Reading {src_path} ...")
    wb_src = openpyxl.load_workbook(
        str(src_path), read_only=True, data_only=True,
    )
    ws_src = wb_src["PTS-2025"]

    # Collect all rows; header is index 0
    all_rows: list[tuple[int, ...]] = []
    header_row: tuple[int, ...] | None = None
    selected_indices: set[int] = set()

    # (country, year) pairs we want
    wanted: set[tuple[str, int]] = {
        ("Afghanistan", 2022),
        ("Afghanistan", 2023),
        ("United States", 2022),
        ("United States", 2023),
        ("Andorra", 2022),
    }

    for i, row in enumerate(ws_src.iter_rows(values_only=True)):
        row_tuple = tuple(row)  # type: ignore[arg-type]
        if i == 0:
            header_row = row_tuple
            all_rows.append(row_tuple)
            continue
        country = row[0]
        year = row[2]
        if (country, year) in wanted:
            selected_indices.add(i)
            all_rows.append(row_tuple)
            wanted.discard((country, year))
            print(
                f"  Selected row {i + 1}: {country} {year} "
                f"(PTS_A={row[8]}, NA_A={row[11]}, "
                f"PTS_H={row[9]}, NA_H={row[12]}, "
                f"PTS_S={row[10]}, NA_S={row[13]})"
            )
        if not wanted:
            break  # all found

    if header_row is None:
        print("ERROR: could not read header row from source xlsx", file=sys.stderr)
        sys.exit(1)

    if selected_indices:
        print(
            f"\nTotal selected data rows: {len(selected_indices)} "
            f"(expected 5)"
        )
    else:
        print("WARNING: no rows selected — fixture will contain only header", file=sys.stderr)

    # Write new workbook
    wb_dst = openpyxl.Workbook()
    ws_dst = wb_dst.active
    ws_dst.title = "PTS-2025"

    # Copy header
    ws_dst.append(list(header_row))  # type: ignore[arg-type]

    # Copy selected data rows
    for row in all_rows[1:]:  # skip header (index 0)
        ws_dst.append(list(row))

    wb_dst.save(str(dst_path))
    print(f"\nWritten: {dst_path}")
    print(f"  Rows: {len(all_rows) - 1} data rows + 1 header")
    print(f"  Columns: {len(header_row)}")

    # Verify what we wrote
    wb_verify = openpyxl.load_workbook(str(dst_path), data_only=True)
    ws_verify = wb_verify.active
    verify_rows = list(ws_verify.iter_rows(values_only=True))
    print(f"\nVerification — {len(verify_rows)} total rows in fixture:")
    for j, vr in enumerate(verify_rows):
        if j == 0:
            print(f"  Header: {list(vr)}")
        else:
            print(
                f"  Row {j}: {vr[0]} {vr[2]} "
                f"PTS_A={vr[8]}/{vr[11]} "
                f"PTS_H={vr[9]}/{vr[12]} "
                f"PTS_S={vr[10]}/{vr[13]}"
            )


if __name__ == "__main__":
    main()
